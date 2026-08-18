import streamlit as st
import gspread
import pandas as pd
from datetime import datetime
from pathlib import Path
import base64
import io
import json
import hashlib
import re
import requests
import uuid
import xlsxwriter
from openpyxl import load_workbook
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account
from PIL import Image, ImageOps

# ==================== 1. 網頁基本與連線設定 ====================
st.set_page_config(page_title="鴻伍裝機日報系統", layout="wide")

# 初始化 Session State (記憶登入狀態與暫存資料)
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'user_name' not in st.session_state:
    st.session_state.user_name = ""
if 'user_role' not in st.session_state:
    st.session_state.user_role = ""
if 'user_id' not in st.session_state:
    st.session_state.user_id = ""
if 'user_permissions' not in st.session_state:
    st.session_state.user_permissions = {}
if 'report_add_form_key' not in st.session_state:
    st.session_state.report_add_form_key = 0
if 'report_edit_form_key' not in st.session_state:
    st.session_state.report_edit_form_key = 0
if 'report_flash_message' not in st.session_state:
    st.session_state.report_flash_message = ""

if 'tab1_search_active' not in st.session_state:
    st.session_state.tab1_search_active = False
if 'tab1_filtered_df' not in st.session_state:
    st.session_state.tab1_filtered_df = pd.DataFrame()
    
if 'tab3_search_active' not in st.session_state:
    st.session_state.tab3_search_active = False
if 'tab3_filtered_df' not in st.session_state:
    st.session_state.tab3_filtered_df = pd.DataFrame()
if 'tab3_edit_requested' not in st.session_state:
    st.session_state.tab3_edit_requested = False
if 'tab3_edit_confirmed' not in st.session_state:
    st.session_state.tab3_edit_confirmed = False

if 'tab1_grid_key' not in st.session_state:
    st.session_state.tab1_grid_key = 0
if 'tab3_grid_key' not in st.session_state:
    st.session_state.tab3_grid_key = 0
if 'tab4_grid_key' not in st.session_state:
    st.session_state.tab4_grid_key = 0
if 'add_form_key' not in st.session_state:
    st.session_state.add_form_key = 0
if 'dev_add_form_key' not in st.session_state:
    st.session_state.dev_add_form_key = 0
if 'dev_add_preview' not in st.session_state:
    st.session_state.dev_add_preview = None
if 'dev_plant_options' not in st.session_state:
    st.session_state.dev_plant_options = []
if 'dev_case_options' not in st.session_state:
    st.session_state.dev_case_options = []
if 'dev_option_manager_key' not in st.session_state:
    st.session_state.dev_option_manager_key = 0
if 'dev_case_checklists' not in st.session_state:
    st.session_state.dev_case_checklists = {}
if 'dev_loaded_case' not in st.session_state:
    st.session_state.dev_loaded_case = None
if 'dev_checklist_key' not in st.session_state:
    st.session_state.dev_checklist_key = 0
if 'dev_test_records' not in st.session_state:
    st.session_state.dev_test_records = []
if 'dev_sales_records' not in st.session_state:
    st.session_state.dev_sales_records = []
if 'dev_sales_form_key' not in st.session_state:
    st.session_state.dev_sales_form_key = 0
if 'dev_development_records' not in st.session_state:
    st.session_state.dev_development_records = []
if 'dev_development_form_key' not in st.session_state:
    st.session_state.dev_development_form_key = 0
if 'installation_excel_last_result' not in st.session_state:
    st.session_state.installation_excel_last_result = None
if 'installation_excel_uploader_key' not in st.session_state:
    st.session_state.installation_excel_uploader_key = 0
if 'installation_unrecorded_grid_key' not in st.session_state:
    st.session_state.installation_unrecorded_grid_key = 0
if 'installation_unrecorded_flash' not in st.session_state:
    st.session_state.installation_unrecorded_flash = ""
if 'dev_results_grid_key' not in st.session_state:
    st.session_state.dev_results_grid_key = 0
if 'dev_results_edit_form_key' not in st.session_state:
    st.session_state.dev_results_edit_form_key = 0
if 'new_installation_migration_checked' not in st.session_state:
    st.session_state.new_installation_migration_checked = False
if 'dev_pending_preview' not in st.session_state:
    st.session_state.dev_pending_preview = None
if 'dev_reason_dialog_key' not in st.session_state:
    st.session_state.dev_reason_dialog_key = 0
if 'dev_flash_message' not in st.session_state:
    st.session_state.dev_flash_message = ""
if 'dev_flash_level' not in st.session_state:
    st.session_state.dev_flash_level = "success"
if 'dev_cloud_initialized' not in st.session_state:
    st.session_state.dev_cloud_initialized = False
if 'dev_data_dirty' not in st.session_state:
    st.session_state.dev_data_dirty = False
if 'dev_pending_sync_message' not in st.session_state:
    st.session_state.dev_pending_sync_message = ""
if 'dev_pending_delete' not in st.session_state:
    st.session_state.dev_pending_delete = None
if 'dev_checklist_navigation_case' not in st.session_state:
    st.session_state.dev_checklist_navigation_case = ""
if 'dev_delete_dialog_key' not in st.session_state:
    st.session_state.dev_delete_dialog_key = 0
if 'dev_pending_previous_record' not in st.session_state:
    st.session_state.dev_pending_previous_record = None
if 'dev_identity_draft' not in st.session_state:
    st.session_state.dev_identity_draft = None
if 'dev_previous_prefill' not in st.session_state:
    st.session_state.dev_previous_prefill = None
if 'dev_manual_only_initialized' not in st.session_state:
    st.session_state.dev_plant_options = []
    st.session_state.dev_case_options = []
    st.session_state.dev_case_checklists = {}
    st.session_state.dev_loaded_case = None
    st.session_state.dev_identity_draft = None
    st.session_state.dev_previous_prefill = None
    st.session_state.dev_pending_previous_record = None
    st.session_state.dev_add_preview = None
    st.session_state.dev_test_records = []
    st.session_state.dev_sales_records = []
    st.session_state.dev_development_records = []
    st.session_state.dev_manual_only_initialized = True

# ==================== Google API 連線與快取優化 ====================
@st.cache_resource(ttl=3600)  # 快取 1 小時，避免頻繁呼叫 Google API 觸發限制
def init_google_sheets():
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            gc = gspread.service_account_from_dict(creds_dict)
        else:
            gc = gspread.service_account(filename='credentials.json')
        
        # 將 open 與 worksheet 的動作全部包在快取函數內，只執行一次
        sheet = gc.open("control table")
        ws_1 = sheet.worksheet("裝機人員進廠時間")
        ws_2 = sheet.worksheet("帳號管理")
        ws_3 = sheet.worksheet("修改紀錄")
        return sheet, ws_1, ws_2, ws_3
        
    except gspread.exceptions.WorksheetNotFound as e:
        st.error(f"⚠️ 系統錯誤：找不到必要的分頁 ({e})。請確認已建立『裝機人員進廠時間』、『帳號管理』與『修改紀錄』。")
        st.stop()
    except Exception as e:
        st.error(f"⚠️ Google API 連線異常，請稍後再試。詳細錯誤：{e}")
        st.stop()

# 取得快取後的試算表物件
sh, worksheet, ws_accounts, ws_log = init_google_sheets()
DEV_WORKSHEET_NAME = "開發區測試資料"
NEW_INSTALLATION_WORKSHEET_NAME = "新版裝機紀錄"
INSTALLATION_EXCEL_WORKSHEET_NAME = "裝機Excel版本紀錄"
INSTALLATION_COMPARISON_WORKSHEET_NAME = "裝機Excel比較紀錄"
REPORT_ENTRY_WORKSHEET_NAME = "報告新增資料"
REPORT_SETTINGS_WORKSHEET_NAME = "報告月份設定"
REPORT_DATA_FILE = Path(__file__).resolve().with_name("data.xlsx")
REPORT_AREA_ORDER = ["北", "中", "南", "國外"]
REPORT_COUNT_COLUMNS = ["訂單數量", "已出貨", "未出貨", "已安裝", "已出貨待安裝"]
REPORT_MONTH_COLUMNS = [f"{month}月份完成率" for month in range(4, 13)]
REPORT_ENTRY_HEADERS = [
    "建立時間", "建立者", "區域", "廠區", "工程名稱",
    *REPORT_COUNT_COLUMNS,
    *REPORT_MONTH_COLUMNS,
]
REPORT_SETTINGS_HEADERS = ["設定鍵", "設定值", "更新時間", "更新者", "操作"]

NEW_INSTALLATION_HEADERS = [
    "紀錄ID",
    "建立時間",
    "更新時間",
    "裝機日期",
    "廠別",
    "案件",
    "機台名稱",
    "項目確認",
    "安裝人員",
    "狀態",
    "未完成或缺貨原因",
    "Remark",
    "建立者",
    "最後修改者",
    "來源版本",
    "來源鍵",
    "照片檔名",
    "照片連結",
    "照片ID",
]


@st.cache_data(ttl=60, show_spinner=False)
def load_production_installation_records():
    """共用正式裝機紀錄快取，避免每個 Streamlit 分頁重複讀取同一工作表。"""
    return worksheet.get_all_records()


@st.cache_data(show_spinner=False)
def load_report_data(file_path, modified_time_ns):
    """讀取報告 Excel；modified_time_ns 只用來讓檔案更新後自動失效快取。"""
    del modified_time_ns
    workbook_sheets = pd.read_excel(file_path, sheet_name=None, header=0)
    report_frames = []

    for area_name in REPORT_AREA_ORDER:
        if area_name not in workbook_sheets:
            continue
        area_df = workbook_sheets[area_name].copy()
        area_df.columns = [
            re.sub(r"\s+", "", str(column))
            for column in area_df.columns
        ]
        required_columns = {"廠區", "工程名稱"}
        if not required_columns.issubset(area_df.columns):
            continue

        area_df["廠區"] = area_df["廠區"].fillna("").astype(str).str.strip()
        area_df["工程名稱"] = area_df["工程名稱"].fillna("").astype(str).str.strip()
        area_df = area_df[
            area_df["廠區"].ne("") & area_df["工程名稱"].ne("")
        ].copy()
        if area_df.empty:
            continue

        area_df.insert(0, "區域", area_name)
        for count_column in REPORT_COUNT_COLUMNS:
            if count_column in area_df.columns:
                area_df[count_column] = pd.to_numeric(
                    area_df[count_column], errors="coerce"
                ).fillna(0)
        report_frames.append(area_df)

    if not report_frames:
        return pd.DataFrame()
    return pd.concat(report_frames, ignore_index=True)


def get_report_entry_worksheet():
    """取得報告新增資料分頁；不存在時自動建立並設定完成率格式。"""
    worksheet_created = False
    try:
        report_worksheet = sh.worksheet(REPORT_ENTRY_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        report_worksheet = sh.add_worksheet(
            title=REPORT_ENTRY_WORKSHEET_NAME,
            rows=1000,
            cols=len(REPORT_ENTRY_HEADERS),
        )
        worksheet_created = True

    if report_worksheet.col_count < len(REPORT_ENTRY_HEADERS):
        report_worksheet.resize(cols=len(REPORT_ENTRY_HEADERS))
    existing_headers = report_worksheet.row_values(1)
    if not existing_headers:
        report_worksheet.append_row(REPORT_ENTRY_HEADERS)
        worksheet_created = True
    elif existing_headers != REPORT_ENTRY_HEADERS:
        raise ValueError(
            f"「{REPORT_ENTRY_WORKSHEET_NAME}」欄位格式不符，請確認標題列。"
        )

    if worksheet_created:
        first_month_column = chr(ord("A") + REPORT_ENTRY_HEADERS.index(REPORT_MONTH_COLUMNS[0]))
        last_month_column = chr(ord("A") + REPORT_ENTRY_HEADERS.index(REPORT_MONTH_COLUMNS[-1]))
        report_worksheet.format(
            f"{first_month_column}2:{last_month_column}",
            {"numberFormat": {"type": "PERCENT", "pattern": "0%"}},
        )
    return report_worksheet


@st.cache_data(ttl=60, show_spinner=False)
def load_report_entry_records():
    """載入由 App 新增、保存在 Google Sheets 的報告資料。"""
    report_worksheet = get_report_entry_worksheet()
    report_records = report_worksheet.get_all_records()
    if not report_records:
        return pd.DataFrame()

    report_df = pd.DataFrame(report_records)
    for text_column in ["區域", "廠區", "工程名稱"]:
        if text_column in report_df.columns:
            report_df[text_column] = (
                report_df[text_column].fillna("").astype(str).str.strip()
            )
    report_df = report_df[
        report_df["區域"].ne("")
        & report_df["廠區"].ne("")
        & report_df["工程名稱"].ne("")
    ].copy()
    for count_column in REPORT_COUNT_COLUMNS:
        if count_column in report_df.columns:
            report_df[count_column] = pd.to_numeric(
                report_df[count_column], errors="coerce"
            ).fillna(0)
    return report_df


@st.cache_resource(show_spinner=False)
def get_report_settings_worksheet():
    """取得報告月份設定分頁；不存在時自動建立。"""
    try:
        settings_worksheet = sh.worksheet(REPORT_SETTINGS_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        settings_worksheet = sh.add_worksheet(
            title=REPORT_SETTINGS_WORKSHEET_NAME,
            rows=200,
            cols=len(REPORT_SETTINGS_HEADERS),
        )
    if settings_worksheet.col_count < len(REPORT_SETTINGS_HEADERS):
        settings_worksheet.resize(cols=len(REPORT_SETTINGS_HEADERS))
    existing_headers = settings_worksheet.row_values(1)
    if not existing_headers:
        settings_worksheet.append_row(REPORT_SETTINGS_HEADERS)
    elif existing_headers != REPORT_SETTINGS_HEADERS:
        raise ValueError(
            f"「{REPORT_SETTINGS_WORKSHEET_NAME}」欄位格式不符，請確認標題列。"
        )
    return settings_worksheet


@st.cache_data(ttl=60, show_spinner=False)
def load_report_active_month():
    """讀取目前統計月份；首次使用時預設為系統月份（限定 4 至 12 月）。"""
    settings_worksheet = get_report_settings_worksheet()
    for row in reversed(settings_worksheet.get_all_records()):
        if str(row.get("設定鍵", "")).strip() != "目前統計月份":
            continue
        try:
            month_value = int(row.get("設定值", 0))
        except (TypeError, ValueError):
            continue
        if 4 <= month_value <= 12:
            return month_value

    default_month = min(max(datetime.now().month, 4), 12)
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    settings_worksheet.append_row(
        [
            "目前統計月份",
            default_month,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            operator,
            "建立預設設定",
        ],
        value_input_option="USER_ENTERED",
    )
    return default_month


def set_report_active_month(month_value, action):
    """以歷史追加方式保存目前統計月份。"""
    month_value = int(month_value)
    if not 4 <= month_value <= 12:
        raise ValueError("統計月份必須介於 4 月至 12 月。")
    settings_worksheet = get_report_settings_worksheet()
    update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    settings_worksheet.append_row(
        ["目前統計月份", month_value, update_time, operator, action],
        value_input_option="USER_ENTERED",
    )
    load_report_active_month.clear()
    try:
        ws_log.append_row(
            [
                update_time,
                operator,
                action,
                "",
                f"{month_value} 月",
            ],
            table_range="A:E",
        )
    except Exception:
        pass
    return month_value


@st.dialog("🔒 確認月份結算")
def show_report_month_settlement_dialog(active_month):
    """確認結算目前月份並切換到下一個月份。"""
    next_month = int(active_month) + 1
    st.warning(
        f"確認結算 {active_month} 月並切換至 {next_month} 月嗎？"
        f"之後新增或修改只會重新計算 {next_month} 月完成率，"
        f"{active_month} 月與更早月份會沿用既有快照。"
    )
    confirm_col, cancel_col = st.columns(2)
    with confirm_col:
        if st.button(
            f"確認結算 {active_month} 月",
            type="primary",
            use_container_width=True,
            key=f"confirm_report_month_{active_month}",
        ):
            try:
                set_report_active_month(
                    next_month,
                    f"結算 {active_month} 月並切換至 {next_month} 月",
                )
                st.session_state.report_flash_message = (
                    f"已結算 {active_month} 月，目前統計月份為 {next_month} 月。"
                )
                st.rerun()
            except Exception as error:
                st.error(f"月份結算失敗：{error}")
    with cancel_col:
        if st.button(
            "取消",
            use_container_width=True,
            key=f"cancel_report_month_{active_month}",
        ):
            st.rerun()


def append_report_entry(
    area,
    plant,
    project,
    order_count,
    shipped_count,
    installed_count,
    existing_record=None,
    action="新增",
):
    """新增報告快照，並在 Google Sheets 寫入衍生欄位公式。"""
    report_worksheet = get_report_entry_worksheet()
    next_row = len(report_worksheet.col_values(1)) + 1
    current_month = load_report_active_month()
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    row_values = {
        "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "建立者": operator,
        "區域": area,
        "廠區": plant,
        "工程名稱": project,
        "訂單數量": int(order_count),
        "已出貨": int(shipped_count),
        "未出貨": f"=F{next_row}-G{next_row}",
        "已安裝": int(installed_count),
        "已出貨待安裝": f"=G{next_row}-I{next_row}",
    }
    existing_record = existing_record or {}
    for month_column in REPORT_MONTH_COLUMNS:
        existing_value = existing_record.get(month_column, "")
        row_values[month_column] = "" if pd.isna(existing_value) else existing_value
    if 4 <= current_month <= 12:
        row_values[f"{current_month}月份完成率"] = (
            f'=IF(G{next_row}=0,"待料",I{next_row}/G{next_row})'
        )

    report_worksheet.append_row(
        [row_values.get(header, "") for header in REPORT_ENTRY_HEADERS],
        value_input_option="USER_ENTERED",
        table_range=f"A:{chr(ord('A') + len(REPORT_ENTRY_HEADERS) - 1)}",
    )
    load_report_entry_records.clear()
    try:
        ws_log.append_row(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                operator,
                f"{action}報告資料：{area}／{plant}／{project}",
                "",
                f"訂單 {int(order_count)}、已出貨 {int(shipped_count)}、已安裝 {int(installed_count)}",
            ],
            table_range="A:E",
        )
    except Exception:
        pass


def format_report_completion_rate(value):
    """將 Excel 完成率小數轉為百分比，並保留「待料」等文字狀態。"""
    if pd.isna(value) or str(value).strip() == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{float(value):.0%}"
    return str(value).strip()


def render_report_area():
    """顯示 data.xlsx 的區域、廠區及工程名稱連動搜尋報告。"""
    st.subheader("📊 報告專區")
    st.caption("資料來源：data.xlsx；區域、廠區與工程名稱會依序連動篩選。")

    if not REPORT_DATA_FILE.exists():
        st.error("找不到報告來源 data.xlsx，請確認檔案已放在 app.py 同一資料夾。")
        return

    try:
        base_report_df = load_report_data(
            str(REPORT_DATA_FILE),
            REPORT_DATA_FILE.stat().st_mtime_ns,
        )
        added_report_df = load_report_entry_records()
        active_report_month = load_report_active_month()
        report_df = pd.concat(
            [frame for frame in [base_report_df, added_report_df] if not frame.empty],
            ignore_index=True,
            sort=False,
        )
        report_df = report_df.drop_duplicates(
            subset=["區域", "廠區", "工程名稱"],
            keep="last",
        ).reset_index(drop=True)
    except Exception as e:
        st.error(f"報告資料讀取失敗：{e}")
        return

    if report_df.empty:
        st.info("Excel 中沒有可顯示的工程明細。")
        return

    if st.session_state.report_flash_message:
        st.success(st.session_state.report_flash_message)
        st.session_state.report_flash_message = ""

    with st.expander("🗓️ 統計月份設定", expanded=True):
        st.info(
            f"目前統計月份：{active_report_month} 月。新增或修改資料時，"
            f"只會重新計算 {active_report_month} 月完成率。"
        )
        month_setting_col1, month_setting_col2 = st.columns([2, 1])
        with month_setting_col1:
            selected_report_month = st.selectbox(
                "目前統計月份",
                list(range(4, 13)),
                index=list(range(4, 13)).index(active_report_month),
                format_func=lambda month: f"{month} 月",
                key=f"report_active_month_selector_{active_report_month}",
            )
        with month_setting_col2:
            st.write("")
            st.write("")
            if st.button(
                "儲存月份設定",
                use_container_width=True,
                disabled=selected_report_month == active_report_month,
                key=f"report_save_active_month_{active_report_month}",
            ):
                try:
                    set_report_active_month(
                        selected_report_month,
                        f"將目前統計月份由 {active_report_month} 月調整為 {selected_report_month} 月",
                    )
                    st.session_state.report_flash_message = (
                        f"目前統計月份已設定為 {selected_report_month} 月。"
                    )
                    st.rerun()
                except Exception as error:
                    st.error(f"月份設定保存失敗：{error}")

        if active_report_month < 12:
            if st.button(
                f"🔒 結算 {active_report_month} 月並切換至 {active_report_month + 1} 月",
                type="primary",
                use_container_width=True,
                key=f"report_settle_month_{active_report_month}",
            ):
                show_report_month_settlement_dialog(active_report_month)
        else:
            st.caption("目前已是 12 月；下一年度開始時請使用上方選單手動設定月份。")

    with st.expander("➕ 新增報告資料", expanded=False):
        st.caption(
            f"請輸入基本數量；未出貨、已出貨待安裝及 {active_report_month} 月完成率"
            "會由 Google Sheets 公式自動計算。"
        )
        report_form_key = st.session_state.report_add_form_key
        add_row1_col1, add_row1_col2, add_row1_col3 = st.columns(3)
        with add_row1_col1:
            add_area = st.selectbox(
                "區域 *",
                REPORT_AREA_ORDER,
                key=f"report_add_area_{report_form_key}",
            )
        area_plant_options = sorted(
            {
                str(value).strip()
                for value in report_df.loc[report_df["區域"].eq(add_area), "廠區"]
                if str(value).strip()
            },
            key=natural_plant_sort_key,
        )
        with add_row1_col2:
            add_plant = st.selectbox(
                "廠區 *",
                area_plant_options,
                index=None,
                placeholder="選擇既有廠區或輸入新廠區",
                accept_new_options=True,
                key=f"report_add_plant_{report_form_key}_{add_area}",
            )
        area_project_options = sorted(
            {
                str(value).strip()
                for value in report_df.loc[
                    report_df["區域"].eq(add_area)
                    & (
                        report_df["廠區"].eq(str(add_plant).strip())
                        if add_plant else True
                    ),
                    "工程名稱",
                ]
                if str(value).strip()
            },
            key=str.casefold,
        )
        with add_row1_col3:
            add_project = st.selectbox(
                "工程名稱 *",
                area_project_options,
                index=None,
                placeholder="選擇既有工程或輸入新工程名稱",
                accept_new_options=True,
                key=f"report_add_project_{report_form_key}_{add_area}_{str(add_plant)}",
            )

        add_row2_col1, add_row2_col2, add_row2_col3 = st.columns(3)
        with add_row2_col1:
            add_order_count = st.number_input(
                "訂單數量 *",
                min_value=0,
                step=1,
                key=f"report_add_order_{report_form_key}",
            )
        with add_row2_col2:
            add_shipped_count = st.number_input(
                "已出貨 *",
                min_value=0,
                step=1,
                key=f"report_add_shipped_{report_form_key}",
            )
        with add_row2_col3:
            add_installed_count = st.number_input(
                "已安裝 *",
                min_value=0,
                step=1,
                key=f"report_add_installed_{report_form_key}",
            )

        if st.button(
            "儲存報告資料",
            type="primary",
            use_container_width=True,
            key=f"report_add_submit_{report_form_key}",
        ):
            cleaned_plant = str(add_plant or "").strip()
            cleaned_project = str(add_project or "").strip()
            if not cleaned_plant or not cleaned_project:
                st.error("請輸入廠區與工程名稱。")
            elif add_shipped_count > add_order_count:
                st.error("已出貨數量不可大於訂單數量。")
            elif add_installed_count > add_shipped_count:
                st.error("已安裝數量不可大於已出貨數量。")
            else:
                try:
                    with st.spinner("正在儲存報告資料並建立公式..."):
                        append_report_entry(
                            add_area,
                            cleaned_plant,
                            cleaned_project,
                            add_order_count,
                            add_shipped_count,
                            add_installed_count,
                        )
                    st.session_state.report_flash_message = (
                        f"已新增：{add_area}／{cleaned_plant}／{cleaned_project}。"
                    )
                    st.session_state.report_add_form_key += 1
                    st.rerun()
                except Exception as e:
                    st.error(f"報告資料新增失敗：{e}")

    with st.expander("✏️ 修改報告資料", expanded=False):
        st.caption("依序選擇區域、廠區與工程名稱後，系統會自動帶入目前數量。")
        edit_form_key = st.session_state.report_edit_form_key
        edit_select_col1, edit_select_col2, edit_select_col3 = st.columns(3)
        with edit_select_col1:
            edit_area = st.selectbox(
                "選擇區域",
                REPORT_AREA_ORDER,
                key=f"report_edit_area_{edit_form_key}",
            )
        edit_plant_options = sorted(
            report_df.loc[report_df["區域"].eq(edit_area), "廠區"]
            .dropna().astype(str).unique(),
            key=natural_plant_sort_key,
        )
        with edit_select_col2:
            edit_plant = st.selectbox(
                "選擇廠區",
                edit_plant_options,
                index=0 if edit_plant_options else None,
                placeholder="此區域沒有廠區資料",
                key=f"report_edit_plant_{edit_form_key}_{edit_area}",
            )
        edit_project_options = sorted(
            report_df.loc[
                report_df["區域"].eq(edit_area)
                & report_df["廠區"].eq(str(edit_plant or "")),
                "工程名稱",
            ].dropna().astype(str).unique(),
            key=str.casefold,
        )
        with edit_select_col3:
            edit_project = st.selectbox(
                "選擇工程名稱",
                edit_project_options,
                index=0 if edit_project_options else None,
                placeholder="此廠區沒有工程資料",
                key=f"report_edit_project_{edit_form_key}_{edit_area}_{str(edit_plant)}",
            )

        selected_report_rows = report_df[
            report_df["區域"].eq(edit_area)
            & report_df["廠區"].eq(str(edit_plant or ""))
            & report_df["工程名稱"].eq(str(edit_project or ""))
        ]
        if selected_report_rows.empty:
            st.info("請先選擇一筆工程資料。")
        else:
            selected_report_record = selected_report_rows.iloc[-1]
            selected_identity = hashlib.sha256(
                f"{edit_area}|{edit_plant}|{edit_project}|{edit_form_key}".encode("utf-8")
            ).hexdigest()[:12]
            edit_count_col1, edit_count_col2, edit_count_col3 = st.columns(3)
            with edit_count_col1:
                edited_order_count = st.number_input(
                    "訂單數量 *",
                    min_value=0,
                    value=int(selected_report_record.get("訂單數量", 0) or 0),
                    step=1,
                    key=f"report_edit_order_{selected_identity}",
                )
            with edit_count_col2:
                edited_shipped_count = st.number_input(
                    "已出貨 *",
                    min_value=0,
                    value=int(selected_report_record.get("已出貨", 0) or 0),
                    step=1,
                    key=f"report_edit_shipped_{selected_identity}",
                )
            with edit_count_col3:
                edited_installed_count = st.number_input(
                    "已裝機 *",
                    min_value=0,
                    value=int(selected_report_record.get("已安裝", 0) or 0),
                    step=1,
                    key=f"report_edit_installed_{selected_identity}",
                )

            if st.button(
                "儲存報告修改",
                type="primary",
                use_container_width=True,
                key=f"report_edit_submit_{selected_identity}",
            ):
                if edited_shipped_count > edited_order_count:
                    st.error("已出貨數量不可大於訂單數量。")
                elif edited_installed_count > edited_shipped_count:
                    st.error("已裝機數量不可大於已出貨數量。")
                else:
                    try:
                        with st.spinner("正在儲存修改並重新建立公式..."):
                            append_report_entry(
                                edit_area,
                                str(edit_plant),
                                str(edit_project),
                                edited_order_count,
                                edited_shipped_count,
                                edited_installed_count,
                                existing_record=selected_report_record.to_dict(),
                                action="修改",
                            )
                        st.session_state.report_flash_message = (
                            f"已更新：{edit_area}／{edit_plant}／{edit_project}。"
                        )
                        st.session_state.report_edit_form_key += 1
                        st.rerun()
                    except Exception as e:
                        st.error(f"報告資料修改失敗：{e}")

    filter_col1, filter_col2, filter_col3 = st.columns(3)
    available_areas = [
        area_name
        for area_name in REPORT_AREA_ORDER
        if area_name in set(report_df["區域"].astype(str))
    ]
    with filter_col1:
        selected_area = st.selectbox(
            "區域",
            ["（全部）", *available_areas],
            key="report_search_area",
        )

    area_filtered_df = report_df[
        report_df["區域"].eq(selected_area)
    ].copy() if selected_area != "（全部）" else report_df.copy()
    available_plants = sorted(
        area_filtered_df["廠區"].dropna().astype(str).unique(),
        key=natural_plant_sort_key,
    )
    with filter_col2:
        selected_plant = st.selectbox(
            "廠區",
            ["（全部）", *available_plants],
            key="report_search_plant",
        )

    plant_filtered_df = area_filtered_df[
        area_filtered_df["廠區"].eq(selected_plant)
    ].copy() if selected_plant != "（全部）" else area_filtered_df.copy()
    available_projects = sorted(
        plant_filtered_df["工程名稱"].dropna().astype(str).unique(),
        key=str.casefold,
    )
    with filter_col3:
        selected_project = st.selectbox(
            "工程名稱",
            ["（全部）", *available_projects],
            key="report_search_project",
        )

    filtered_df = plant_filtered_df[
        plant_filtered_df["工程名稱"].eq(selected_project)
    ].copy() if selected_project != "（全部）" else plant_filtered_df.copy()

    metric_values = {}
    for count_column in REPORT_COUNT_COLUMNS:
        metric_values[count_column] = int(
            pd.to_numeric(filtered_df.get(count_column, 0), errors="coerce").fillna(0).sum()
        )
    metric_columns = st.columns(5)
    metric_labels = ["訂單數量", "已出貨", "未出貨", "已安裝", "已出貨待安裝"]
    for metric_column, metric_label in zip(metric_columns, metric_labels):
        with metric_column:
            st.metric(metric_label, f"{metric_values[metric_label]:,} 台")

    display_df = filtered_df.copy()
    report_month_columns = [
        column for column in display_df.columns
        if column.endswith("月份完成率")
    ]
    for month_column in report_month_columns:
        display_df[month_column] = display_df[month_column].apply(
            format_report_completion_rate
        )
    for count_column in REPORT_COUNT_COLUMNS:
        if count_column in display_df.columns:
            display_df[count_column] = display_df[count_column].astype(int)

    preferred_columns = [
        "區域", "廠區", "工程名稱",
        *REPORT_COUNT_COLUMNS,
        *report_month_columns,
    ]
    display_columns = [
        column for column in preferred_columns
        if column in display_df.columns
    ]
    st.markdown(f"#### 搜尋結果（{len(display_df)} 筆工程）")
    st.dataframe(
        display_df[display_columns],
        hide_index=True,
        use_container_width=True,
        height=min(700, 80 + max(len(display_df), 1) * 35),
    )

    render_installation_excel_version_area()


def installation_record_source_key(prefix, record, row_number=None):
    """產生穩定來源鍵，用來避免舊版或開發暫存資料被重複轉入。"""
    source_values = [
        str(record.get(field, "")).strip()
        for field in [
            "建立時間", "日期", "廠別", "案件", "機台名稱", "項目確認",
            "安裝人員", "狀態", "未完成或缺貨原因", "Remark",
        ]
    ]
    if row_number is not None:
        source_values.insert(0, str(row_number))
    digest = hashlib.sha256(
        json.dumps(source_values, ensure_ascii=False).encode("utf-8")
    ).hexdigest()
    return f"{prefix}-{digest}"


def get_new_installation_worksheet():
    """取得或建立新版裝機紀錄分頁，並補齊缺少的欄位。"""
    try:
        new_worksheet = sh.worksheet(NEW_INSTALLATION_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        new_worksheet = sh.add_worksheet(
            title=NEW_INSTALLATION_WORKSHEET_NAME,
            rows=1000,
            cols=len(NEW_INSTALLATION_HEADERS),
        )

    if new_worksheet.col_count < len(NEW_INSTALLATION_HEADERS):
        new_worksheet.resize(cols=len(NEW_INSTALLATION_HEADERS))
    existing_headers = new_worksheet.row_values(1)
    if not existing_headers:
        new_worksheet.append_row(NEW_INSTALLATION_HEADERS)
    else:
        for header in NEW_INSTALLATION_HEADERS:
            if header not in existing_headers:
                new_worksheet.update_cell(1, len(existing_headers) + 1, header)
                existing_headers.append(header)
        if existing_headers[:len(NEW_INSTALLATION_HEADERS)] != NEW_INSTALLATION_HEADERS:
            raise ValueError(
                f"「{NEW_INSTALLATION_WORKSHEET_NAME}」欄位順序不符，請勿手動調整標題列。"
            )
    return new_worksheet


@st.cache_data(ttl=30, show_spinner=False)
def load_new_installation_records():
    """讀取新版裝機資料並轉成畫面共用欄位。"""
    new_worksheet = get_new_installation_worksheet()
    records = []
    for sheet_row_number, row in enumerate(new_worksheet.get_all_records(), start=2):
        records.append({
            "紀錄ID": str(row.get("紀錄ID", "")).strip(),
            "建立時間": str(row.get("建立時間", "")).strip(),
            "更新時間": str(row.get("更新時間", "")).strip(),
            "日期": str(row.get("裝機日期", "")).strip(),
            "廠別": str(row.get("廠別", "")).strip(),
            "案件": str(row.get("案件", "")).strip(),
            "機台名稱": str(row.get("機台名稱", "")).strip(),
            "項目確認": str(row.get("項目確認", "")).strip(),
            "安裝人員": str(row.get("安裝人員", "")).strip(),
            "狀態": str(row.get("狀態", "")).strip(),
            "未完成或缺貨原因": str(row.get("未完成或缺貨原因", "")).strip(),
            "Remark": str(row.get("Remark", "")).strip(),
            "建立者": str(row.get("建立者", "")).strip(),
            "最後修改者": str(row.get("最後修改者", "")).strip(),
            "來源版本": str(row.get("來源版本", "")).strip() or "新版輸入",
            "來源鍵": str(row.get("來源鍵", "")).strip(),
            "照片檔名": str(row.get("照片檔名", "")).strip(),
            "照片連結": str(row.get("照片連結", "")).strip(),
            "照片ID": str(row.get("照片ID", "")).strip(),
            "_new_sheet_row": sheet_row_number,
        })
    return records


def new_installation_row_values(record, existing_record=None):
    """將畫面資料轉成新版工作表的一列。"""
    existing_record = existing_record or {}
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    return [
        str(record.get("紀錄ID") or existing_record.get("紀錄ID") or uuid.uuid4()),
        str(existing_record.get("建立時間") or record.get("建立時間") or now_text),
        now_text,
        str(record.get("日期", "")).strip(),
        str(record.get("廠別", "")).strip(),
        str(record.get("案件", "")).strip(),
        str(record.get("機台名稱", "")).strip(),
        str(record.get("項目確認", "")).strip(),
        str(record.get("安裝人員", "")).strip(),
        str(record.get("狀態", "")).strip(),
        str(record.get("未完成或缺貨原因", "")).strip(),
        str(record.get("Remark", "")).strip(),
        str(existing_record.get("建立者") or record.get("建立者") or operator),
        operator,
        str(record.get("來源版本") or existing_record.get("來源版本") or "新版輸入"),
        str(record.get("來源鍵") or existing_record.get("來源鍵") or ""),
        str(record.get("照片檔名") or existing_record.get("照片檔名") or ""),
        str(record.get("照片連結") or existing_record.get("照片連結") or ""),
        str(record.get("照片ID") or existing_record.get("照片ID") or ""),
    ]


def append_new_installation_record(record):
    """新增一筆新版裝機紀錄。"""
    new_worksheet = get_new_installation_worksheet()
    values = new_installation_row_values(record)
    new_worksheet.append_row(values, value_input_option="USER_ENTERED")
    load_new_installation_records.clear()
    return values[0]


def update_new_installation_record(existing_record, updated_record):
    """以紀錄 ID 更新新版資料；不接觸舊版工作表。"""
    new_worksheet = get_new_installation_worksheet()
    row_number = int(existing_record.get("_new_sheet_row", 0) or 0)
    if row_number < 2:
        raise ValueError("找不到新版資料的工作表列號。")
    values = new_installation_row_values(updated_record, existing_record)
    end_column = gspread.utils.rowcol_to_a1(1, len(NEW_INSTALLATION_HEADERS))[:-1]
    new_worksheet.update(
        range_name=f"A{row_number}:{end_column}{row_number}",
        values=[values],
    )
    load_new_installation_records.clear()


def delete_new_installation_record(record):
    """依紀錄 ID 重新定位並刪除新版資料列，避免使用過期列號。"""
    record_id = str(record.get("紀錄ID", "")).strip()
    if not record_id:
        raise ValueError("找不到新版裝機紀錄 ID。")

    new_worksheet = get_new_installation_worksheet()
    current_records = new_worksheet.get_all_records()
    target_row_number = next(
        (
            row_number
            for row_number, current_record in enumerate(current_records, start=2)
            if str(current_record.get("紀錄ID", "")).strip() == record_id
        ),
        None,
    )
    if target_row_number is None:
        raise ValueError("此筆新版裝機紀錄已不存在，請重新整理後再試。")

    new_worksheet.delete_rows(target_row_number)
    load_new_installation_records.clear()
    return target_row_number


def complete_matching_new_installation_records(record, exclude_record_id=""):
    """將相同廠別、案件與機台的其他新版紀錄一併標示為已完成。"""
    target_identity = (
        str(record.get("廠別", "")).strip().casefold(),
        str(record.get("案件", "")).strip().casefold(),
        str(record.get("機台名稱", "")).strip().casefold(),
    )
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    changed_cells = []
    changed_count = 0
    for existing_record in load_new_installation_records():
        existing_identity = (
            str(existing_record.get("廠別", "")).strip().casefold(),
            str(existing_record.get("案件", "")).strip().casefold(),
            str(existing_record.get("機台名稱", "")).strip().casefold(),
        )
        if (
            existing_identity != target_identity
            or str(existing_record.get("紀錄ID", "")) == str(exclude_record_id)
            or str(existing_record.get("狀態", "")).strip() == "已完成"
        ):
            continue
        row_number = int(existing_record["_new_sheet_row"])
        changed_cells.extend([
            gspread.Cell(row_number, 3, now_text),
            gspread.Cell(row_number, 10, "已完成"),
            gspread.Cell(row_number, 11, ""),
            gspread.Cell(row_number, 14, operator),
        ])
        changed_count += 1
    if changed_cells:
        get_new_installation_worksheet().update_cells(
            changed_cells,
            value_input_option="USER_ENTERED",
        )
        load_new_installation_records.clear()
    return changed_count


def legacy_records_for_new_interface():
    """替舊版資料加上來源鍵；舊版資料只供搜尋與轉換。"""
    records = []
    for row_number, record in enumerate(load_production_installation_records(), start=2):
        converted = dict(record)
        converted["來源鍵"] = installation_record_source_key("OLD", converted, row_number)
        converted["來源版本"] = "舊版"
        converted["_legacy_sheet_row"] = row_number
        records.append(converted)
    return records


def migrate_saved_dev_installation_records():
    """將既有開發裝機紀錄一次性補入新版資料庫。"""
    if st.session_state.new_installation_migration_checked:
        return 0
    existing_records = load_new_installation_records()
    existing_source_keys = {
        str(record.get("來源鍵", "")).strip() for record in existing_records
    }
    rows_to_append = []
    for record_index, record in enumerate(st.session_state.dev_test_records, start=1):
        source_key = installation_record_source_key("DEV", record, record_index)
        if source_key in existing_source_keys:
            continue
        migrated_record = dict(record)
        migrated_record["來源版本"] = "開發資料轉入"
        migrated_record["來源鍵"] = source_key
        rows_to_append.append(new_installation_row_values(migrated_record))
        existing_source_keys.add(source_key)
    if rows_to_append:
        get_new_installation_worksheet().append_rows(
            rows_to_append,
            value_input_option="USER_ENTERED",
        )
        load_new_installation_records.clear()
    st.session_state.new_installation_migration_checked = True
    return len(rows_to_append)

ACCOUNT_PERMISSION_COLUMNS = {
    "sales_access": "可進入業務專區",
    "development_access": "可進入開發專區",
    "handoff_access": "可進入背鍋俠專區",
    "installation_access": "可進入裝機確認區",
    "attachment_upload": "可上傳附件",
    "attachment_download": "可下載附件",
    "attachment_delete": "可刪除附件",
}

ROLE_PERMISSION_DEFAULTS = {
    "管理者": {permission: True for permission in ACCOUNT_PERMISSION_COLUMNS},
    "業務": {
        "sales_access": True,
        "development_access": False,
        "handoff_access": False,
        "installation_access": False,
        "attachment_upload": False,
        "attachment_download": False,
        "attachment_delete": False,
    },
    "工程師": {
        "sales_access": False,
        "development_access": True,
        "handoff_access": False,
        "installation_access": False,
        "attachment_upload": True,
        "attachment_download": True,
        "attachment_delete": False,
    },
    "RD": {
        "sales_access": False,
        "development_access": True,
        "handoff_access": False,
        "installation_access": False,
        "attachment_upload": True,
        "attachment_download": True,
        "attachment_delete": False,
    },
    "背鍋俠": {
        "sales_access": False,
        "development_access": False,
        "handoff_access": True,
        "installation_access": False,
        "attachment_upload": False,
        "attachment_download": True,
        "attachment_delete": False,
    },
    "裝機": {
        "sales_access": False,
        "development_access": False,
        "handoff_access": False,
        "installation_access": True,
        "attachment_upload": False,
        "attachment_download": True,
        "attachment_delete": False,
    },
    "公用": {permission: False for permission in ACCOUNT_PERMISSION_COLUMNS},
}


def natural_plant_sort_key(plant_name):
    """依廠區代號與其中的數字做自然排序，例如 AP01、AP02、T1、T2、T10。"""
    cleaned_name = str(plant_name or "").strip()
    natural_parts = [
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.findall(r"\d+|\D+", cleaned_name)
    ]
    return (*natural_parts, (2, cleaned_name.casefold()))


def find_latest_unfinished_production_record(records, target_identity):
    """從正式裝機紀錄找出相同廠別、案件、機台的最新未完成資料。"""
    completed_statuses = {"已完成", "完成"}
    for record in reversed(records):
        record_identity = (
            str(record.get("廠別", "")).strip().casefold(),
            str(record.get("案件", "")).strip().casefold(),
            str(record.get("機台名稱", "")).strip().casefold(),
        )
        original_status = str(record.get("狀態", "")).strip()
        if (
            record_identity == target_identity
            and original_status
            and original_status not in completed_statuses
        ):
            return {
                "日期": str(record.get("日期", "")).strip(),
                "廠別": str(record.get("廠別", "")).strip(),
                "案件": str(record.get("案件", "")).strip(),
                "機台名稱": str(record.get("機台名稱", "")).strip(),
                "項目確認": "",
                "安裝人員": str(record.get("安裝人員", "")).strip(),
                "狀態": "未完成",
                "原始狀態": original_status,
                "Remark": str(record.get("Remark", "")).strip(),
                "資料來源": "正式裝機紀錄",
            }
    return None


def find_latest_unfinished_new_record(records, target_identity):
    """從新版資料庫找出相同廠別、案件、機台的最新未完成資料。"""
    for record in reversed(records):
        record_identity = (
            str(record.get("廠別", "")).strip().casefold(),
            str(record.get("案件", "")).strip().casefold(),
            str(record.get("機台名稱", "")).strip().casefold(),
        )
        if record_identity == target_identity and str(record.get("狀態", "")).strip() == "未完成":
            previous_record = dict(record)
            previous_record["資料來源"] = "新版裝機紀錄"
            return previous_record
    return None


def parse_permission_flag(value, default=False):
    """將試算表中的常用布林寫法轉換為權限值。"""
    cleaned_value = str(value or "").strip().lower()
    if not cleaned_value:
        return bool(default)
    return cleaned_value in {"是", "true", "1", "yes", "y", "v", "✓"}


def permissions_for_account(account_row=None, role=""):
    """取得帳號細項權限；空白欄位依角色預設值判斷。"""
    account_row = account_row or {}
    defaults = ROLE_PERMISSION_DEFAULTS.get(
        str(role).strip(),
        {permission: False for permission in ACCOUNT_PERMISSION_COLUMNS},
    )
    return {
        permission: parse_permission_flag(
            account_row.get(column_name, ""),
            defaults.get(permission, False),
        )
        for permission, column_name in ACCOUNT_PERMISSION_COLUMNS.items()
    }


def ensure_account_permission_columns():
    """保留既有帳號資料並自動補上附件與專區權限欄位。"""
    headers = ws_accounts.row_values(1)
    if not headers:
        raise ValueError("「帳號管理」分頁沒有標題列。")

    required_headers = ["帳號", "密碼", "姓名", "權限", *ACCOUNT_PERMISSION_COLUMNS.values()]
    if ws_accounts.col_count < len(required_headers):
        ws_accounts.resize(cols=len(required_headers))

    for header in required_headers:
        if header not in headers:
            ws_accounts.update_cell(1, len(headers) + 1, header)
            headers.append(header)

    account_values = ws_accounts.get_all_values()
    if len(account_values) <= 1:
        return

    role_column = headers.index("權限")
    pending_cells = []
    for row_number, row_values in enumerate(account_values[1:], start=2):
        role = row_values[role_column].strip() if role_column < len(row_values) else ""
        defaults = ROLE_PERMISSION_DEFAULTS.get(
            role,
            {permission: False for permission in ACCOUNT_PERMISSION_COLUMNS},
        )
        for permission, column_name in ACCOUNT_PERMISSION_COLUMNS.items():
            column_number = headers.index(column_name) + 1
            current_value = (
                row_values[column_number - 1].strip()
                if column_number - 1 < len(row_values)
                else ""
            )
            if not current_value:
                pending_cells.append(
                    gspread.Cell(row_number, column_number, "是" if defaults[permission] else "否")
                )

    if pending_cells:
        ws_accounts.update_cells(pending_cells, value_input_option="USER_ENTERED")


def get_apps_script_upload_config():
    """讀取不會提交至 GitHub 的 Apps Script 上傳設定。"""
    try:
        config = st.secrets["apps_script"]
        upload_url = str(config["upload_url"]).strip()
        upload_token = str(config["upload_token"]).strip()
    except Exception as e:
        raise RuntimeError("尚未設定 Apps Script 附件上傳資訊。") from e

    if not upload_url or not upload_token:
        raise RuntimeError("Apps Script 上傳網址或 Token 為空白。")
    return upload_url, upload_token


def upload_dev_attachment(uploaded_file, order_number, part_number):
    """透過 Apps Script 將開發附件存入使用者的 Google Drive。"""
    if not st.session_state.get("user_permissions", {}).get(
        "attachment_upload",
        st.session_state.get("user_role") == "管理者",
    ):
        raise PermissionError("目前帳號沒有上傳附件的權限。")
    upload_url, upload_token = get_apps_script_upload_config()
    original_name = str(uploaded_file.name or "attachment").strip()
    context_prefix = f"{order_number}_{part_number}".replace("/", "-").replace("\\", "-")
    stored_name = f"{context_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{original_name}"
    mime_type = uploaded_file.type or "application/octet-stream"
    response = requests.post(
        upload_url,
        json={
            "token": upload_token,
            "action": "upload",
            "fileName": stored_name,
            "mimeType": mime_type,
            "data": base64.b64encode(uploaded_file.getvalue()).decode("ascii"),
        },
        timeout=180,
    )
    response.raise_for_status()
    try:
        drive_file = response.json()
    except ValueError as e:
        raise RuntimeError("Apps Script 未回傳有效的 JSON 資料。") from e
    if not drive_file.get("ok"):
        raise RuntimeError(drive_file.get("error") or "Apps Script 上傳失敗。")
    return {
        "id": drive_file["id"],
        "原始檔名": original_name,
        "Drive檔名": drive_file.get("name", stored_name),
        "連結": drive_file.get("url", ""),
    }


def upload_installation_photo(photo_data, plant_name, case_name, machine_name):
    """將新版裝機照片上傳至既有的私人 Drive 附件資料夾。"""
    if not st.session_state.get("user_permissions", {}).get(
        "attachment_upload",
        st.session_state.get("user_role") == "管理者",
    ):
        raise PermissionError("目前帳號沒有上傳照片的權限。")

    binary_data = photo_data.get("data", b"")
    if not binary_data:
        raise ValueError("照片內容為空白。")
    if len(binary_data) > 10 * 1024 * 1024:
        raise ValueError(f"照片「{photo_data.get('name', '')}」不可超過 10 MB。")

    upload_url, upload_token = get_apps_script_upload_config()
    original_name = str(photo_data.get("name") or "photo.jpg").strip()
    context_prefix = "_".join([plant_name, case_name, machine_name])
    context_prefix = re.sub(r'[\\/:*?"<>|]+', "-", context_prefix).strip() or "installation"
    stored_name = (
        f"{context_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{original_name}"
    )
    response = requests.post(
        upload_url,
        json={
            "token": upload_token,
            "action": "upload",
            "fileName": stored_name,
            "mimeType": photo_data.get("type") or "image/jpeg",
            "data": base64.b64encode(binary_data).decode("ascii"),
        },
        timeout=180,
    )
    response.raise_for_status()
    try:
        drive_file = response.json()
    except ValueError as e:
        raise RuntimeError("Apps Script 未回傳有效的 JSON 資料。") from e
    if not drive_file.get("ok"):
        raise RuntimeError(drive_file.get("error") or "照片上傳失敗。")
    return {
        "id": str(drive_file.get("id", "")).strip(),
        "name": original_name,
        "url": str(drive_file.get("url", "")).strip(),
    }


def parse_installation_photo_list(value):
    """解析新版裝機照片欄位；相容 JSON 陣列與單一舊值。"""
    cleaned_value = str(value or "").strip()
    if not cleaned_value:
        return []
    try:
        parsed_value = json.loads(cleaned_value)
        if isinstance(parsed_value, list):
            return [str(item).strip() for item in parsed_value if str(item).strip()]
    except (TypeError, ValueError, json.JSONDecodeError):
        pass
    return [cleaned_value]


def installation_photos_from_record(record):
    """將三個照片欄位合併成方便顯示與下載的結構。"""
    names = parse_installation_photo_list(record.get("照片檔名", ""))
    links = parse_installation_photo_list(record.get("照片連結", ""))
    file_ids = parse_installation_photo_list(record.get("照片ID", ""))
    photo_count = max(len(names), len(links), len(file_ids), 0)
    return [
        {
            "附件檔名": names[index] if index < len(names) else f"裝機照片 {index + 1}",
            "附件連結": links[index] if index < len(links) else "",
            "附件ID": file_ids[index] if index < len(file_ids) else "",
        }
        for index in range(photo_count)
    ]


def extract_drive_file_id(file_url):
    """從既有 Drive 連結取得檔案 ID，供舊紀錄刪除使用。"""
    cleaned_url = str(file_url or "").strip()
    path_match = re.search(r"/d/([A-Za-z0-9_-]+)", cleaned_url)
    if path_match:
        return path_match.group(1)
    query_match = re.search(r"[?&]id=([A-Za-z0-9_-]+)", cleaned_url)
    return query_match.group(1) if query_match else ""


def delete_dev_attachment(file_id):
    """透過 Apps Script 將指定資料夾中的附件移至 Google Drive 垃圾桶。"""
    if not st.session_state.get("user_permissions", {}).get(
        "attachment_delete",
        st.session_state.get("user_role") == "管理者",
    ):
        raise PermissionError("目前帳號沒有刪除附件的權限。")

    cleaned_file_id = str(file_id or "").strip()
    if not cleaned_file_id:
        raise ValueError("找不到附件的 Drive 檔案 ID。")

    upload_url, upload_token = get_apps_script_upload_config()
    response = requests.post(
        upload_url,
        json={
            "token": upload_token,
            "action": "delete",
            "fileId": cleaned_file_id,
        },
        timeout=60,
    )
    response.raise_for_status()
    try:
        result = response.json()
    except ValueError as e:
        raise RuntimeError("Apps Script 未回傳有效的 JSON 資料。") from e
    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "Apps Script 刪除附件失敗。")
    return result


@st.cache_resource(ttl=3600)
def get_drive_download_session():
    """建立僅具 Google Drive 讀取權限的服務帳號連線。"""
    drive_scope = ["https://www.googleapis.com/auth/drive.readonly"]
    if "gcp_service_account" in st.secrets:
        credentials = service_account.Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]),
            scopes=drive_scope,
        )
    else:
        credentials = service_account.Credentials.from_service_account_file(
            "credentials.json",
            scopes=drive_scope,
        )
    return AuthorizedSession(credentials)


@st.cache_data(ttl=300, show_spinner=False)
def download_private_drive_file(file_id):
    """直接讀取私人 Drive 檔案，避免 Apps Script Base64 下載逾時。"""
    cleaned_file_id = str(file_id or "").strip()
    if not cleaned_file_id:
        raise ValueError("找不到附件的 Drive 檔案 ID。")

    session = get_drive_download_session()
    api_url = f"https://www.googleapis.com/drive/v3/files/{cleaned_file_id}"
    metadata_response = session.get(
        api_url,
        params={"fields": "id,name,mimeType,size"},
        timeout=30,
    )
    metadata_response.raise_for_status()
    metadata = metadata_response.json()
    media_response = session.get(
        api_url,
        params={"alt": "media"},
        timeout=120,
    )
    media_response.raise_for_status()
    return {
        "name": metadata.get("name", "attachment"),
        "mime_type": metadata.get("mimeType", "application/octet-stream"),
        "bytes": media_response.content,
    }


def download_dev_attachment(file_id):
    """使用服務帳號安全讀取私人 Drive 附件。"""
    if not st.session_state.get("user_permissions", {}).get(
        "attachment_download",
        st.session_state.get("user_role") == "管理者",
    ):
        raise PermissionError("目前帳號沒有下載附件的權限。")

    cleaned_file_id = str(file_id or "").strip()
    if not cleaned_file_id:
        raise ValueError("找不到附件的 Drive 檔案 ID。")

    try:
        return download_private_drive_file(cleaned_file_id)
    except Exception as e:
        raise RuntimeError(f"私人 Drive 附件下載失敗：{e}") from e


def sync_dev_data_to_google():
    """將開發區目前內容以完整快照方式同步至獨立工作表。"""
    headers = [
        "同步批次",
        "紀錄類型",
        "操作時間",
        "裝機日期",
        "操作者",
        "廠別",
        "案件",
        "機台名稱",
        "項目確認",
        "安裝人員",
        "狀態",
        "未完成或缺貨原因",
        "Remark",
        "訂單",
        "品號",
        "附件檔名",
        "附件連結",
        "附件ID",
        "背鍋俠確認時間",
        "背鍋俠確認人",
    ]
    sync_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    operator = f"{st.session_state.user_name} ({st.session_state.user_role})"
    rows = [[
        sync_time, "快照資訊", sync_time, "", operator,
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    ]]

    for plant_name in st.session_state.dev_plant_options:
        rows.append([
            sync_time, "廠別選項", sync_time, "", operator,
            plant_name, "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        ])

    for case_name in st.session_state.dev_case_options:
        checklist_items = st.session_state.dev_case_checklists.get(case_name, [])
        rows.append([
            sync_time, "案件設定", sync_time, "", operator,
            "", case_name, "", "\n".join(checklist_items), "", "", "", "", "", "", "", "", "", "", "",
        ])

    for record in st.session_state.dev_test_records:
        rows.append([
            sync_time,
            "裝機測試結果",
            record.get("建立時間", sync_time),
            record.get("日期", ""),
            operator,
            record.get("廠別", ""),
            record.get("案件", ""),
            record.get("機台名稱", ""),
            record.get("項目確認", ""),
            record.get("安裝人員", ""),
            record.get("狀態", ""),
            record.get("未完成或缺貨原因", ""),
            record.get("Remark", ""),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    for record in st.session_state.dev_sales_records:
        rows.append([
            sync_time,
            "業務專區",
            record.get("建立時間", sync_time),
            "",
            record.get("建立者", operator),
            "", "", "", "", "", "", "", "",
            record.get("訂單", ""),
            record.get("品號", ""),
            "",
            "",
            "",
            "",
            "",
        ])

    for record in st.session_state.dev_development_records:
        rows.append([
            sync_time,
            "開發專區",
            record.get("建立時間", sync_time),
            "",
            record.get("建立者", operator),
            "", record.get("案件", ""), "", "", "", "", "", "",
            record.get("訂單", ""),
            record.get("品號", ""),
            record.get("附件檔名", ""),
            record.get("附件連結", ""),
            record.get("附件ID", ""),
            record.get("背鍋俠確認時間", ""),
            record.get("背鍋俠確認人", ""),
        ])

    try:
        dev_worksheet = sh.worksheet(DEV_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        dev_worksheet = sh.add_worksheet(
            title=DEV_WORKSHEET_NAME,
            rows=1000,
            cols=len(headers),
        )

    # Google Sheets 不允許直接在目前格線範圍的尾端之外插入欄位。
    # 舊版工作表只有 13 欄，先擴充至新版所需欄數，再進行欄位遷移。
    if dev_worksheet.col_count < len(headers):
        dev_worksheet.resize(cols=len(headers))

    existing_headers = dev_worksheet.row_values(1)
    if existing_headers and "裝機日期" not in existing_headers:
        dev_worksheet.insert_cols([["裝機日期"]], col=4)
        existing_headers = dev_worksheet.row_values(1)
    if existing_headers and "未完成或缺貨原因" not in existing_headers:
        reason_col = headers.index("未完成或缺貨原因") + 1
        dev_worksheet.insert_cols([["未完成或缺貨原因"]], col=reason_col)
        existing_headers = dev_worksheet.row_values(1)
    for dev_header in [
        "訂單", "品號", "附件檔名", "附件連結", "附件ID",
        "背鍋俠確認時間", "背鍋俠確認人",
    ]:
        if existing_headers and dev_header not in existing_headers:
            dev_worksheet.insert_cols([[dev_header]], col=len(existing_headers) + 1)
            existing_headers = dev_worksheet.row_values(1)
    if not existing_headers:
        dev_worksheet.append_row(headers)
    elif existing_headers != headers:
        raise ValueError(f"「{DEV_WORKSHEET_NAME}」的欄位格式不符，請先確認標題列。")
    dev_worksheet.append_rows(rows, value_input_option="USER_ENTERED")
    return len(rows) - 1


def load_latest_dev_data_from_google():
    """從開發區工作表載入最後一次同步快照。"""
    try:
        dev_worksheet = sh.worksheet(DEV_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        return None

    saved_rows = dev_worksheet.get_all_records()
    if not saved_rows:
        return None

    latest_batch = next(
        (
            str(row.get("同步批次", "")).strip()
            for row in reversed(saved_rows)
            if str(row.get("同步批次", "")).strip()
        ),
        "",
    )
    if not latest_batch:
        return None

    batch_rows = [
        row for row in saved_rows
        if str(row.get("同步批次", "")).strip() == latest_batch
    ]

    plant_options = sorted(
        {
            str(row.get("廠別", "")).strip()
            for row in batch_rows
            if row.get("紀錄類型") == "廠別選項" and str(row.get("廠別", "")).strip()
        },
        key=natural_plant_sort_key,
    )

    case_options = []
    case_checklists = {}
    test_records = []
    sales_records = []
    development_records = []
    for row in batch_rows:
        record_type = str(row.get("紀錄類型", "")).strip()
        if record_type == "案件設定":
            case_name = str(row.get("案件", "")).strip()
            if case_name:
                case_options.append(case_name)
                case_checklists[case_name] = [
                    item.strip()
                    for item in str(row.get("項目確認", "")).splitlines()
                    if item.strip()
                ]
        elif record_type == "裝機測試結果":
            test_records.append({
                "建立時間": str(row.get("操作時間", "")).strip(),
                "日期": str(row.get("裝機日期", "")).strip(),
                "廠別": str(row.get("廠別", "")).strip(),
                "案件": str(row.get("案件", "")).strip(),
                "機台名稱": str(row.get("機台名稱", "")).strip(),
                "項目確認": str(row.get("項目確認", "")).strip(),
                "安裝人員": str(row.get("安裝人員", "")).strip(),
                "狀態": str(row.get("狀態", "")).strip(),
                "未完成或缺貨原因": str(row.get("未完成或缺貨原因", "")).strip(),
                "Remark": str(row.get("Remark", "")).strip(),
            })
        elif record_type == "業務專區":
            sales_records.append({
                "建立時間": str(row.get("操作時間", "")).strip(),
                "訂單": str(row.get("訂單", "")).strip(),
                "品號": str(row.get("品號", "")).strip(),
                "建立者": str(row.get("操作者", "")).strip(),
            })
        elif record_type == "開發專區":
            development_records.append({
                "建立時間": str(row.get("操作時間", "")).strip(),
                "訂單": str(row.get("訂單", "")).strip(),
                "品號": str(row.get("品號", "")).strip(),
                "案件": str(row.get("案件", "")).strip(),
                "附件檔名": str(row.get("附件檔名", "")).strip(),
                "附件連結": str(row.get("附件連結", "")).strip(),
                "附件ID": str(row.get("附件ID", "")).strip(),
                "背鍋俠確認時間": str(row.get("背鍋俠確認時間", "")).strip(),
                "背鍋俠確認人": str(row.get("背鍋俠確認人", "")).strip(),
                "建立者": str(row.get("操作者", "")).strip(),
            })

    st.session_state.dev_plant_options = plant_options
    st.session_state.dev_case_options = sorted(set(case_options))
    st.session_state.dev_case_checklists = case_checklists
    st.session_state.dev_test_records = test_records
    st.session_state.dev_sales_records = sales_records
    st.session_state.dev_development_records = development_records
    st.session_state.dev_loaded_case = None
    st.session_state.dev_identity_draft = None
    st.session_state.dev_previous_prefill = None
    st.session_state.dev_pending_previous_record = None
    st.session_state.dev_add_preview = None
    st.session_state.dev_pending_preview = None
    st.session_state.dev_option_manager_key += 1
    st.session_state.dev_add_form_key += 1
    st.session_state.dev_checklist_key += 1

    return {
        "批次": latest_batch,
        "廠別": len(plant_options),
        "案件": len(case_options),
        "測試結果": len(test_records),
        "業務資料": len(sales_records),
        "開發資料": len(development_records),
    }


def queue_dev_auto_sync(message):
    """標記開發資料已異動，於下一次重跑時自動保存。"""
    st.session_state.dev_data_dirty = True
    st.session_state.dev_pending_sync_message = message


def auto_sync_pending_dev_data():
    """自動保存待同步的開發資料；失敗時保留待同步狀態供下次重試。"""
    if not st.session_state.dev_data_dirty:
        return True

    action_message = st.session_state.dev_pending_sync_message or "開發資料已更新"
    try:
        synced_count = sync_dev_data_to_google()
        st.session_state.dev_data_dirty = False
        st.session_state.dev_pending_sync_message = ""
        st.session_state.dev_flash_level = "success"
        st.session_state.dev_flash_message = (
            f"{action_message}，已自動同步至「{DEV_WORKSHEET_NAME}」"
            f"（目前 {synced_count} 筆資料）。"
        )
        return True
    except Exception as e:
        st.session_state.dev_flash_level = "error"
        st.session_state.dev_flash_message = (
            f"{action_message}，但自動同步失敗；系統會在下次操作時重試。詳細錯誤：{e}"
        )
        return False


def initialize_dev_cloud_data():
    """供業務、開發與管理頁面共用的雲端資料初始化。"""
    if not st.session_state.dev_cloud_initialized:
        st.session_state.dev_cloud_initialized = True
        try:
            load_summary = load_latest_dev_data_from_google()
            if load_summary:
                st.session_state.dev_flash_level = "success"
                st.session_state.dev_flash_message = (
                    f"已自動載入雲端資料：{load_summary['廠別']} 個廠別、"
                    f"{load_summary['案件']} 個案件、"
                    f"{load_summary['測試結果']} 筆裝機紀錄、"
                    f"{load_summary['業務資料']} 筆業務資料、"
                    f"{load_summary['開發資料']} 筆開發資料。"
                )
            else:
                st.session_state.dev_flash_level = "info"
                st.session_state.dev_flash_message = "雲端目前沒有開發區保存資料，將從空白資料開始。"
        except Exception as e:
            st.session_state.dev_flash_level = "error"
            st.session_state.dev_flash_message = f"自動載入開發資料失敗：{e}"

    auto_sync_pending_dev_data()
    if st.session_state.dev_flash_message:
        flash_renderer = getattr(st, st.session_state.dev_flash_level, st.info)
        flash_renderer(st.session_state.dev_flash_message)
        st.session_state.dev_flash_message = ""
        st.session_state.dev_flash_level = "success"


def can_delete_dev_data():
    return bool(
        st.session_state.get("user_permissions", {}).get(
            "attachment_delete",
            st.session_state.get("user_role") == "管理者",
        )
    )


def render_secure_attachment_download(record, key_prefix, label_prefix="下載"):
    """以延遲下載方式由 Apps Script 取得附件，不顯示 Drive 網址。"""
    attachment_name = str(record.get("附件檔名", "")).strip() or "開發附件"
    file_id = str(record.get("附件ID", "")).strip() or extract_drive_file_id(
        record.get("附件連結", "")
    )
    if not file_id:
        st.warning(f"附件「{attachment_name}」缺少 Drive 檔案 ID，暫時無法下載。")
        return

    if not st.session_state.get("user_permissions", {}).get(
        "attachment_download",
        st.session_state.get("user_role") == "管理者",
    ):
        st.write(f"📎 {attachment_name}")
        st.caption("目前帳號沒有下載附件的權限。")
        return

    st.download_button(
        f"📎 {label_prefix} {attachment_name}",
        data=lambda target_file_id=file_id: download_dev_attachment(target_file_id)["bytes"],
        file_name=attachment_name,
        mime="application/octet-stream",
        on_click="ignore",
        use_container_width=True,
        key=f"{key_prefix}_secure_download",
    )


def render_sales_delete_controls(key_prefix):
    """供具刪除權限者選擇業務紀錄，實際刪除在確認對話框進行。"""
    if not can_delete_dev_data() or not st.session_state.dev_sales_records:
        return

    st.markdown("##### 管理資料")
    selected_index = st.selectbox(
        "選擇要刪除的業務紀錄",
        range(len(st.session_state.dev_sales_records)),
        format_func=lambda index: (
            f"{st.session_state.dev_sales_records[index].get('訂單', '')}｜"
            f"{st.session_state.dev_sales_records[index].get('品號', '')}｜"
            f"{st.session_state.dev_sales_records[index].get('建立時間', '')}"
        ),
        key=f"{key_prefix}_sales_delete_select",
    )
    if st.button(
        "🗑️ 刪除選取的業務資料",
        use_container_width=True,
        key=f"{key_prefix}_sales_delete_button",
    ):
        st.session_state.dev_pending_delete = {"type": "sales", "index": selected_index}
        st.rerun()


def render_development_delete_button(record_index, key_prefix):
    """顯示單筆開發紀錄刪除入口。"""
    if not can_delete_dev_data():
        return
    if st.button(
        "🗑️ 刪除開發紀錄與附件",
        use_container_width=True,
        key=f"{key_prefix}_development_delete_{record_index}",
    ):
        st.session_state.dev_pending_delete = {
            "type": "development",
            "index": record_index,
        }
        st.rerun()


def find_matching_case_name(case_name):
    """以去除前後空白且不分英文字母大小寫的方式尋找既有案件。"""
    cleaned_name = str(case_name or "").strip()
    if not cleaned_name:
        return ""
    normalized_name = cleaned_name.casefold()
    return next(
        (
            existing_name
            for existing_name in st.session_state.dev_case_options
            if str(existing_name).strip().casefold() == normalized_name
        ),
        "",
    )


def queue_checklist_navigation(case_name):
    """安排下一次重跑時切換到指定案件的確認項目設定。"""
    st.session_state.dev_checklist_navigation_case = str(case_name or "").strip()
    st.session_state.dev_option_manager_key += 1


def render_sales_area():
    """顯示業務資料輸入；呼叫前必須先完成業務專區權限判斷。"""
    st.markdown("### 業務專區")
    st.caption("輸入訂單與品號後會立即記錄，並自動保存至 Google Sheets。")

    sales_form_key = st.session_state.dev_sales_form_key
    with st.form(f"dev_sales_form_{sales_form_key}"):
        sales_col1, sales_col2 = st.columns(2)
        with sales_col1:
            sales_order = st.text_input(
                "訂單 *",
                placeholder="輸入訂單編號",
                key=f"dev_sales_order_{sales_form_key}",
            )
        with sales_col2:
            sales_part_number = st.text_input(
                "品號 *",
                placeholder="輸入品號",
                key=f"dev_sales_part_{sales_form_key}",
            )

        sales_submitted = st.form_submit_button(
            "新增業務紀錄",
            type="primary",
            use_container_width=True,
        )

    if sales_submitted:
        if not st.session_state.get("user_permissions", {}).get("sales_access", False):
            st.error("目前帳號沒有新增業務資料的權限。")
        else:
            cleaned_order = sales_order.strip()
            cleaned_part_number = sales_part_number.strip()
            missing_sales_fields = []
            if not cleaned_order:
                missing_sales_fields.append("訂單")
            if not cleaned_part_number:
                missing_sales_fields.append("品號")

            if missing_sales_fields:
                st.error(f"請填寫必填欄位：{'、'.join(missing_sales_fields)}")
            else:
                st.session_state.dev_sales_records.append({
                    "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "訂單": cleaned_order,
                    "品號": cleaned_part_number,
                    "建立者": f"{st.session_state.user_name} ({st.session_state.user_role})",
                })
                queue_dev_auto_sync(
                    f"已新增業務紀錄：訂單「{cleaned_order}」、品號「{cleaned_part_number}」"
                )
                st.session_state.dev_sales_form_key += 1
                st.rerun()

    st.divider()
    st.markdown("#### 已記錄資料")
    if st.session_state.dev_sales_records:
        sales_df = pd.DataFrame(st.session_state.dev_sales_records)
        sales_columns = [
            column for column in ["建立時間", "訂單", "品號", "建立者"]
            if column in sales_df.columns
        ]
        st.dataframe(sales_df[sales_columns], hide_index=True, use_container_width=True)
        st.info(f"目前共有 {len(sales_df)} 筆業務資料。")
        render_sales_delete_controls("role_sales")
    else:
        st.info("目前沒有業務資料。")


def render_development_area(can_upload, can_download):
    """顯示開發附件頁面，並在操作點再次驗證上傳與下載權限。"""
    st.markdown("### 開發專區")
    st.caption("從業務專區資料選擇訂單與品號，上傳附件後建立開發紀錄。")

    sales_pairs = []
    for sales_record in st.session_state.dev_sales_records:
        pair = (
            str(sales_record.get("訂單", "")).strip(),
            str(sales_record.get("品號", "")).strip(),
        )
        if all(pair) and pair not in sales_pairs:
            sales_pairs.append(pair)

    if not sales_pairs:
        st.info("目前沒有可選擇的訂單與品號，請先由業務人員建立資料。")
    else:
        development_orders = list(dict.fromkeys(order for order, _ in sales_pairs))
        selected_development_order = st.selectbox(
            "選擇訂單 *",
            development_orders,
            key="dev_development_order",
        )
        development_parts = [
            part_number for order, part_number in sales_pairs
            if order == selected_development_order
        ]
        selected_development_part = st.selectbox(
            "選擇品號 *",
            development_parts,
            key=(
                f"dev_development_part_{st.session_state.dev_development_form_key}_"
                f"{selected_development_order}"
            ),
        )

        development_form_key = st.session_state.dev_development_form_key
        with st.form(f"dev_development_form_{development_form_key}"):
            development_file = st.file_uploader(
                "附加檔案 *",
                accept_multiple_files=False,
                disabled=not can_upload,
                key=f"dev_development_files_{development_form_key}",
            )
            st.caption("每筆紀錄附加一個檔案，單檔上限 10 MB；附件會保存到指定的 Google Drive 資料夾。")
            development_submitted = st.form_submit_button(
                "上傳附件並建立紀錄",
                type="primary",
                use_container_width=True,
                disabled=not can_upload,
            )

        if not can_upload:
            st.info("目前帳號只有檢視權限，無法上傳附件。")
        elif development_submitted:
            if not st.session_state.get("user_permissions", {}).get("attachment_upload", False):
                st.error("目前帳號沒有上傳附件的權限。")
            elif development_file is None:
                st.error("請附加一個檔案。")
            elif development_file.size > 10 * 1024 * 1024:
                st.error("附件不可超過 10 MB。")
            else:
                try:
                    with st.spinner("正在上傳附件並建立開發紀錄..."):
                        uploaded_drive_file = upload_dev_attachment(
                            development_file,
                            selected_development_order,
                            selected_development_part,
                        )
                        st.session_state.dev_development_records.append({
                            "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "訂單": selected_development_order,
                            "品號": selected_development_part,
                            "案件": "",
                            "附件檔名": uploaded_drive_file["原始檔名"],
                            "附件連結": uploaded_drive_file["連結"],
                            "附件ID": uploaded_drive_file["id"],
                            "背鍋俠確認時間": "",
                            "背鍋俠確認人": "",
                            "建立者": (
                                f"{st.session_state.user_name} "
                                f"({st.session_state.user_role})"
                            ),
                        })
                        queue_dev_auto_sync(
                            f"已建立訂單「{selected_development_order}」、"
                            f"品號「{selected_development_part}」的開發紀錄"
                        )
                        st.session_state.dev_development_form_key += 1
                        st.rerun()
                except Exception as e:
                    st.error(f"附件上傳或紀錄建立失敗，本次資料未保存。詳細錯誤：{e}")

    st.divider()
    st.markdown("#### 已建立的開發紀錄")
    if st.session_state.dev_development_records:
        development_df = pd.DataFrame(st.session_state.dev_development_records)
        development_display = development_df.copy()
        development_display["附件數量"] = development_display["附件檔名"].apply(
            lambda value: len([name for name in str(value).splitlines() if name.strip()])
        )
        development_columns = [
            column for column in ["建立時間", "訂單", "品號", "附件數量", "建立者"]
            if column in development_display.columns
        ]
        st.dataframe(
            development_display[development_columns],
            hide_index=True,
            use_container_width=True,
        )

        for record_index, development_record in reversed(
            list(enumerate(st.session_state.dev_development_records, start=1))
        ):
            with st.expander(
                f"{development_record.get('訂單', '')}｜"
                f"{development_record.get('品號', '')}｜"
                f"{development_record.get('建立時間', '')}",
                expanded=False,
            ):
                attachment_names = [
                    name.strip()
                    for name in str(development_record.get("附件檔名", "")).splitlines()
                    if name.strip()
                ]
                attachment_links = [
                    link.strip()
                    for link in str(development_record.get("附件連結", "")).splitlines()
                    if link.strip()
                ]
                for attachment_index, attachment_name in enumerate(attachment_names):
                    if can_download and attachment_index < len(attachment_links):
                        st.link_button(
                            f"📎 {attachment_name}",
                            attachment_links[attachment_index],
                            use_container_width=True,
                            key=(
                                f"dev_attachment_{record_index}_{attachment_index}_"
                                f"{st.session_state.dev_development_form_key}"
                            ),
                        )
                    else:
                        st.write(f"📎 {attachment_name}")
                if not can_download:
                    st.caption("目前帳號沒有下載附件的權限。")
                render_development_delete_button(record_index - 1, "role_development")
        st.info(f"目前共有 {len(development_df)} 筆開發資料。")
    else:
        st.info("目前沒有開發資料。")


def render_handoff_area(can_download):
    """顯示開發上傳成果，供背鍋俠下載附件並確認交接。"""
    st.markdown("### 背鍋俠專區")
    st.caption("確認開發上傳的訂單、品號與附件；確認後資料會進入裝機確認區。")

    if not st.session_state.dev_development_records:
        st.info("目前沒有開發上傳資料。")
        return

    handoff_rows = []
    for record in st.session_state.dev_development_records:
        confirmed_time = str(record.get("背鍋俠確認時間", "")).strip()
        handoff_rows.append({
            "訂單": record.get("訂單", ""),
            "品號": record.get("品號", ""),
            "附件檔名": record.get("附件檔名", ""),
            "狀態": "已確認" if confirmed_time else "待確認",
            "確認時間": confirmed_time,
            "確認人": record.get("背鍋俠確認人", ""),
        })
    st.dataframe(pd.DataFrame(handoff_rows), hide_index=True, use_container_width=True)

    pending_count = sum(1 for row in handoff_rows if row["狀態"] == "待確認")
    st.info(f"目前共 {len(handoff_rows)} 筆，其中 {pending_count} 筆待確認。")

    ordered_records = sorted(
        enumerate(st.session_state.dev_development_records),
        key=lambda item: bool(str(item[1].get("背鍋俠確認時間", "")).strip()),
    )
    for record_index, record in ordered_records:
        confirmed_time = str(record.get("背鍋俠確認時間", "")).strip()
        status_label = "✅ 已確認" if confirmed_time else "⏳ 待確認"
        with st.expander(
            f"{status_label}｜{record.get('訂單', '')}｜{record.get('品號', '')}",
            expanded=not bool(confirmed_time),
        ):
            st.markdown(f"**開發上傳時間：** {record.get('建立時間', '')}")
            st.markdown(f"**上傳者：** {record.get('建立者', '')}")
            attachment_name = str(record.get("附件檔名", "")).strip()
            if can_download:
                render_secure_attachment_download(record, f"handoff_{record_index}")
            else:
                st.write(f"📎 {attachment_name or '（無附件名稱）'}")
                st.caption("目前帳號沒有下載附件的權限。")

            if confirmed_time:
                st.success(
                    f"已由 {record.get('背鍋俠確認人', '') or '未知使用者'} "
                    f"於 {confirmed_time} 確認。"
                )
            elif st.button(
                "✅ 確認並送至裝機確認區",
                type="primary",
                use_container_width=True,
                key=f"handoff_confirm_{record_index}",
            ):
                if not st.session_state.get("user_permissions", {}).get(
                    "handoff_access",
                    st.session_state.get("user_role") == "管理者",
                ):
                    st.error("目前帳號沒有執行背鍋俠確認的權限。")
                else:
                    confirmation_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    confirmer = f"{st.session_state.user_name} ({st.session_state.user_role})"
                    record["背鍋俠確認時間"] = confirmation_time
                    record["背鍋俠確認人"] = confirmer
                    action_message = (
                        f"已確認開發附件並送至裝機確認區："
                        f"訂單「{record.get('訂單', '')}」、品號「{record.get('品號', '')}」"
                    )
                    queue_dev_auto_sync(action_message)
                    log_dev_delete_action(action_message, attachment_name, "已確認")
                    st.rerun()


INSTALLATION_ORDER_FIELDS = [
    "日期", "訂單", "客戶簡稱", "業務員名稱", "品號", "品名",
    "訂單數量", "未交數量", "已交數量",
]
INSTALLATION_IDENTITY_FIELDS = INSTALLATION_ORDER_FIELDS[:6]
INSTALLATION_QUANTITY_FIELDS = INSTALLATION_ORDER_FIELDS[6:]
INSTALLATION_TARGET_PART_NUMBER = "PM-0001-00"
INSTALLATION_HEADER_ALIASES = {
    "日期": ("日期",),
    "訂單": ("訂單",),
    "客戶簡稱": ("客戶簡稱",),
    "業務員名稱": ("業務員名稱",),
    "品號": ("品號",),
    "品名": ("品名",),
    "訂單數量": ("訂單數量",),
    "未交數量": ("未交數量", "未交跟未開發票"),
    "已交數量": ("已交數量",),
}


def normalize_excel_cell_value(value):
    """將訂單明細儲存格轉成可穩定比較及保存的文字。"""
    if value is None:
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        return format(value, ".15g")
    return str(value).strip()


def normalize_order_number(order_number):
    """移除訂單最後的四碼項次，例如 -0002。"""
    cleaned_order = str(order_number or "").strip()
    return re.sub(r"-\d{4}\s*$", "", cleaned_order)


def find_installation_header_indexes(header_values):
    """辨識不同匯出版本的訂單明細欄位，並統一成固定九欄。"""
    normalized_headers = [
        re.sub(r"\s+", "", normalize_excel_cell_value(value))
        for value in header_values
    ]
    header_indexes = {}
    for field_name in INSTALLATION_ORDER_FIELDS:
        normalized_aliases = [
            re.sub(r"\s+", "", alias)
            for alias in INSTALLATION_HEADER_ALIASES.get(field_name, (field_name,))
        ]
        if field_name in {"日期", "訂單"}:
            header_index = next(
                (
                    index for index, header in enumerate(normalized_headers)
                    if any(header.startswith(alias) for alias in normalized_aliases)
                ),
                None,
            )
        else:
            header_index = next(
                (
                    index for index, header in enumerate(normalized_headers)
                    if header in normalized_aliases
                ),
                None,
            )
        if header_index is None:
            return None
        header_indexes[field_name] = header_index
    return header_indexes


def parse_installation_excel(file_bytes):
    """依 PM-0001-00 找出訂單主號，再擷取所有同訂單項目。"""
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    source_rows = []
    matched_sheet_names = []
    try:
        for sheet in workbook.worksheets:
            if sheet.max_row > 20000 or sheet.max_column > 500:
                raise ValueError(
                    f"工作表「{sheet.title}」範圍過大（{sheet.max_row} 列、"
                    f"{sheet.max_column} 欄）；單張工作表上限為 20,000 列、500 欄。"
                )
            header_row_number = None
            header_indexes = None
            for candidate_row_number, header_values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 30),
                    values_only=True,
                ),
                start=1,
            ):
                candidate_indexes = find_installation_header_indexes(header_values)
                if candidate_indexes:
                    header_row_number = candidate_row_number
                    header_indexes = candidate_indexes
                    break

            if not header_indexes:
                continue
            matched_sheet_names.append(sheet.title)
            for row_number, row_values in enumerate(
                sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
                start=header_row_number + 1,
            ):
                record = {
                    field_name: normalize_excel_cell_value(row_values[column_index])
                    if column_index < len(row_values) else ""
                    for field_name, column_index in header_indexes.items()
                }
                if not record["訂單"]:
                    continue
                source_rows.append({
                    **record,
                    "工作表": sheet.title,
                    "Excel列": row_number,
                    "訂單主號": normalize_order_number(record["訂單"]),
                })
    finally:
        workbook.close()

    if not matched_sheet_names:
        raise ValueError(
            "找不到包含「日期、訂單、客戶簡稱、業務員名稱、品號、品名、"
            "訂單數量、未交數量、已交數量」的工作表。"
        )

    target_rows = [
        row for row in source_rows
        if row["品號"].strip().upper() == INSTALLATION_TARGET_PART_NUMBER
    ]
    if not target_rows:
        raise ValueError(f"找不到品號為 {INSTALLATION_TARGET_PART_NUMBER} 的項目。")

    target_order_numbers = {
        row["訂單主號"] for row in target_rows if row["訂單主號"]
    }
    parsed_rows = []
    for row in source_rows:
        if row["訂單主號"] not in target_order_numbers:
            continue
        comparison_data = {
            field_name: row[field_name]
            for field_name in INSTALLATION_ORDER_FIELDS
        }
        signature = json.dumps(
            comparison_data,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        parsed_rows.append({
            **comparison_data,
            "工作表": row["工作表"],
            "Excel列": row["Excel列"],
            "訂單主號": row["訂單主號"],
            "內容": "｜".join(
                f"{field_name}={comparison_data[field_name]}"
                for field_name in INSTALLATION_ORDER_FIELDS
            ),
            "資料JSON": signature,
        })

    return {
        "資料": parsed_rows,
        "目標品號筆數": len(target_rows),
        "訂單主號": sorted(target_order_numbers),
        "資料工作表": matched_sheet_names,
    }


@st.cache_resource(show_spinner=False)
def get_installation_excel_worksheet():
    """取得或建立裝機 Excel 版本紀錄分頁。"""
    headers = [
        "版本ID", "紀錄類型", "上傳時間", "上傳者", "訂單", "品號",
        "檔名", "工作表", "Excel列", "資料JSON", "顯示內容",
    ]
    try:
        excel_worksheet = sh.worksheet(INSTALLATION_EXCEL_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        excel_worksheet = sh.add_worksheet(
            title=INSTALLATION_EXCEL_WORKSHEET_NAME,
            rows=2000,
            cols=len(headers),
        )

    if excel_worksheet.col_count < len(headers):
        excel_worksheet.resize(cols=len(headers))
    existing_headers = excel_worksheet.row_values(1)
    if not existing_headers:
        excel_worksheet.append_row(headers)
    elif existing_headers != headers:
        raise ValueError(
            f"「{INSTALLATION_EXCEL_WORKSHEET_NAME}」的欄位格式不符，請確認標題列。"
        )
    return excel_worksheet


@st.cache_data(ttl=60, show_spinner=False)
def load_installation_excel_versions(order_number, part_number):
    """載入指定訂單、品號的所有 Excel 版本。"""
    excel_worksheet = get_installation_excel_worksheet()
    versions = {}
    for sheet_row_number, row in enumerate(excel_worksheet.get_all_records(), start=2):
        if (
            str(row.get("訂單", "")).strip() != order_number
            or str(row.get("品號", "")).strip() != part_number
        ):
            continue
        version_id = str(row.get("版本ID", "")).strip()
        if not version_id:
            continue
        version = versions.setdefault(version_id, {
            "版本ID": version_id,
            "上傳時間": str(row.get("上傳時間", "")).strip(),
            "上傳者": str(row.get("上傳者", "")).strip(),
            "檔名": str(row.get("檔名", "")).strip(),
            "資料": [],
            "工作表列": [],
        })
        version["工作表列"].append(sheet_row_number)
        if str(row.get("紀錄類型", "")).strip() == "資料":
            version["資料"].append({
                "工作表": str(row.get("工作表", "")).strip(),
                "Excel列": row.get("Excel列", ""),
                "內容": str(row.get("顯示內容", "")).strip(),
                "資料JSON": str(row.get("資料JSON", "")).strip(),
            })
    return sorted(versions.values(), key=lambda version: (version["上傳時間"], version["版本ID"]))


def save_installation_excel_version(order_number, part_number, file_name, parsed_rows):
    """將單次 Excel 讀取結果以完整版本寫入 Google 試算表。"""
    excel_worksheet = get_installation_excel_worksheet()
    upload_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    version_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
    uploader = f"{st.session_state.user_name} ({st.session_state.user_role})"
    rows = [[
        version_id, "版本資訊", upload_time, uploader, order_number, part_number,
        file_name, "", "", "", "",
    ]]
    rows.extend([
        [
            version_id, "資料", upload_time, uploader, order_number, part_number,
            file_name, row["工作表"], row["Excel列"], row["資料JSON"], row["內容"],
        ]
        for row in parsed_rows
    ])
    # 直接擴充格線再附加，避免為了取得最後一列額外讀取整張工作表。
    excel_worksheet.resize(rows=excel_worksheet.row_count + len(rows) + 50)
    excel_worksheet.append_rows(rows, value_input_option="RAW")
    return {
        "版本ID": version_id,
        "上傳時間": upload_time,
        "上傳者": uploader,
        "檔名": file_name,
        "資料": parsed_rows,
    }


def prune_installation_excel_versions(
    order_number,
    part_number,
    keep_version_id,
    loaded_versions=None,
):
    """刪除指定資料流的舊版本列，只保留目前要作為下次基準的版本。"""
    rows_to_delete = []
    removed_version_ids = set()
    versions = loaded_versions
    if versions is None:
        versions = load_installation_excel_versions(order_number, part_number)
    for version in versions:
        version_id = str(version.get("版本ID", "")).strip()
        if version_id and version_id != keep_version_id:
            rows_to_delete.extend(
                int(row_number)
                for row_number in version.get("工作表列", [])
                if str(row_number).strip()
            )
            removed_version_ids.add(version_id)

    if not rows_to_delete:
        return {"刪除列數": 0, "刪除版本數": 0}

    excel_worksheet = get_installation_excel_worksheet()
    rows_to_delete = sorted(set(rows_to_delete))
    contiguous_ranges = []
    range_start = range_end = rows_to_delete[0]
    for row_number in rows_to_delete[1:]:
        if row_number == range_end + 1:
            range_end = row_number
        else:
            contiguous_ranges.append((range_start, range_end))
            range_start = range_end = row_number
    contiguous_ranges.append((range_start, range_end))

    # 由底部往上刪除，避免前方列號因刪除動作而位移。
    for range_start, range_end in reversed(contiguous_ranges):
        excel_worksheet.delete_rows(range_start, range_end)
    return {
        "刪除列數": len(rows_to_delete),
        "刪除版本數": len(removed_version_ids),
    }


def installation_excel_display_row(row):
    """將目前或已保存的版本資料還原為指定九欄。"""
    if all(field_name in row for field_name in INSTALLATION_ORDER_FIELDS):
        return {
            field_name: row.get(field_name, "")
            for field_name in INSTALLATION_ORDER_FIELDS
        }
    try:
        saved_data = json.loads(str(row.get("資料JSON", "")))
    except (TypeError, ValueError, json.JSONDecodeError):
        saved_data = {}
    if all(field_name in saved_data for field_name in INSTALLATION_ORDER_FIELDS):
        return {
            field_name: saved_data.get(field_name, "")
            for field_name in INSTALLATION_ORDER_FIELDS
        }
    return {
        "日期": "",
        "訂單": "",
        "客戶簡稱": "",
        "業務員名稱": "",
        "品號": "",
        "品名": str(row.get("內容", "")).strip(),
        "訂單數量": "",
        "未交數量": "",
        "已交數量": "",
    }


def compare_installation_excel_rows(previous_rows, current_rows):
    """區分新增、刪除與三個數量欄位的變更，並支援重複資料列。"""
    previous_buckets = {}
    current_buckets = {}

    def identity_key(row):
        display_row = installation_excel_display_row(row)
        return json.dumps(
            {field_name: display_row[field_name] for field_name in INSTALLATION_IDENTITY_FIELDS},
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )

    for row in previous_rows:
        previous_buckets.setdefault(identity_key(row), []).append(row)
    for row in current_rows:
        current_buckets.setdefault(identity_key(row), []).append(row)

    added_rows = []
    removed_rows = []
    quantity_changes = []
    for item_key in sorted(set(previous_buckets) | set(current_buckets)):
        previous_group = list(previous_buckets.get(item_key, []))
        current_group = list(current_buckets.get(item_key, []))

        # 先排除內容完全相同的資料，再將同一項目的剩餘資料配對為數量變更。
        current_by_signature = {}
        for row in current_group:
            current_by_signature.setdefault(row["資料JSON"], []).append(row)
        unmatched_previous = []
        for row in previous_group:
            matching_rows = current_by_signature.get(row["資料JSON"], [])
            if matching_rows:
                matching_rows.pop()
            else:
                unmatched_previous.append(row)
        unmatched_current = [
            row
            for matching_rows in current_by_signature.values()
            for row in matching_rows
        ]

        unmatched_previous.sort(
            key=lambda row: tuple(
                installation_excel_display_row(row)[field_name]
                for field_name in INSTALLATION_QUANTITY_FIELDS
            )
        )
        unmatched_current.sort(
            key=lambda row: tuple(
                installation_excel_display_row(row)[field_name]
                for field_name in INSTALLATION_QUANTITY_FIELDS
            )
        )
        paired_count = min(len(unmatched_previous), len(unmatched_current))
        for index in range(paired_count):
            previous_data = installation_excel_display_row(unmatched_previous[index])
            current_data = installation_excel_display_row(unmatched_current[index])
            quantity_changes.append({
                **{
                    field_name: current_data[field_name]
                    for field_name in INSTALLATION_IDENTITY_FIELDS
                },
                **{
                    f"原{field_name}": previous_data[field_name]
                    for field_name in INSTALLATION_QUANTITY_FIELDS
                },
                **{
                    f"新{field_name}": current_data[field_name]
                    for field_name in INSTALLATION_QUANTITY_FIELDS
                },
            })

        added_rows.extend(
            installation_excel_display_row(row)
            for row in unmatched_current[paired_count:]
        )
        removed_rows.extend(
            installation_excel_display_row(row)
            for row in unmatched_previous[paired_count:]
        )

    return added_rows, removed_rows, quantity_changes


@st.cache_resource(show_spinner=False)
def get_installation_comparison_worksheet():
    """取得或建立每次 Excel 比較結果的永久紀錄分頁。"""
    headers = [
        "比較ID", "紀錄類型", "比較時間", "比較者", "前版ID", "新版本ID",
        "前版檔名", "新版本檔名", "目標品號", "資料JSON",
        "處理狀態", "處理時間", "處理者",
    ]
    try:
        comparison_worksheet = sh.worksheet(INSTALLATION_COMPARISON_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        comparison_worksheet = sh.add_worksheet(
            title=INSTALLATION_COMPARISON_WORKSHEET_NAME,
            rows=2000,
            cols=len(headers),
        )

    if comparison_worksheet.col_count < len(headers):
        comparison_worksheet.resize(cols=len(headers))
    existing_headers = comparison_worksheet.row_values(1)
    if not existing_headers:
        comparison_worksheet.append_row(headers)
    elif existing_headers == headers[:len(existing_headers)]:
        if len(existing_headers) < len(headers):
            comparison_worksheet.update(
                range_name=f"A1:{chr(ord('A') + len(headers) - 1)}1",
                values=[headers],
            )
    else:
        raise ValueError(
            f"「{INSTALLATION_COMPARISON_WORKSHEET_NAME}」的欄位格式不符，請確認標題列。"
        )
    return comparison_worksheet


def save_installation_excel_comparison(
    previous_version,
    current_version,
    added_rows,
    removed_rows,
    quantity_changes,
):
    """將一次比較的新增、刪除與數量變更完整保存。"""
    comparison_worksheet = get_installation_comparison_worksheet()
    comparison_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    comparison_id = datetime.now().strftime("CMP%Y%m%d%H%M%S%f")
    comparer = f"{st.session_state.user_name} ({st.session_state.user_role})"
    base_columns = [
        comparison_id,
        "",
        comparison_time,
        comparer,
        previous_version["版本ID"],
        current_version["版本ID"],
        previous_version["檔名"],
        current_version["檔名"],
        INSTALLATION_TARGET_PART_NUMBER,
    ]
    summary = {
        "新增項目": len(added_rows),
        "刪除項目": len(removed_rows),
        "數量變更": len(quantity_changes),
    }
    rows = [[
        *base_columns[:1],
        "比較資訊",
        *base_columns[2:],
        json.dumps(summary, ensure_ascii=False, separators=(",", ":")),
        "", "", "",
    ]]
    for record_type, records in [
        ("新增項目", added_rows),
        ("刪除項目", removed_rows),
        ("數量變更", quantity_changes),
    ]:
        rows.extend([
            [
                *base_columns[:1],
                record_type,
                *base_columns[2:],
                json.dumps(record, ensure_ascii=False, separators=(",", ":")),
                "未紀錄" if record_type == "新增項目" else "",
                "",
                "",
            ]
            for record in records
        ])

    # 避免為了取得最後一列再次讀取整張比較紀錄表。
    comparison_worksheet.resize(
        rows=comparison_worksheet.row_count + len(rows) + 50
    )
    comparison_worksheet.append_rows(rows, value_input_option="RAW")
    load_unrecorded_installation_items.clear()
    return {
        "比較ID": comparison_id,
        "比較時間": comparison_time,
        "比較者": comparer,
        "前版ID": previous_version["版本ID"],
        "新版本ID": current_version["版本ID"],
        "前版檔名": previous_version["檔名"],
        "新版本檔名": current_version["檔名"],
        "新增": [
            {**record, "紀錄狀態": "未紀錄"}
            for record in added_rows
        ],
        "刪除": removed_rows,
        "數量變更": quantity_changes,
    }


@st.cache_data(ttl=60, show_spinner=False)
def load_installation_excel_comparisons():
    """載入所有已永久保存的 Excel 比較結果。"""
    comparison_worksheet = get_installation_comparison_worksheet()
    comparisons = {}
    record_type_mapping = {
        "新增項目": "新增",
        "刪除項目": "刪除",
        "數量變更": "數量變更",
    }
    for row in comparison_worksheet.get_all_records():
        comparison_id = str(row.get("比較ID", "")).strip()
        if not comparison_id:
            continue
        comparison = comparisons.setdefault(comparison_id, {
            "比較ID": comparison_id,
            "比較時間": str(row.get("比較時間", "")).strip(),
            "比較者": str(row.get("比較者", "")).strip(),
            "前版ID": str(row.get("前版ID", "")).strip(),
            "新版本ID": str(row.get("新版本ID", "")).strip(),
            "前版檔名": str(row.get("前版檔名", "")).strip(),
            "新版本檔名": str(row.get("新版本檔名", "")).strip(),
            "新增": [],
            "刪除": [],
            "數量變更": [],
        })
        target_key = record_type_mapping.get(str(row.get("紀錄類型", "")).strip())
        if not target_key:
            continue
        try:
            saved_record = json.loads(str(row.get("資料JSON", "")))
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
        if isinstance(saved_record, dict):
            if target_key == "新增":
                record_status = str(row.get("處理狀態", "")).strip()
                if record_status:
                    saved_record["紀錄狀態"] = record_status
            comparison[target_key].append(saved_record)
    return sorted(
        comparisons.values(),
        key=lambda comparison: (comparison["比較時間"], comparison["比較ID"]),
    )


@st.cache_data(ttl=30, show_spinner=False)
def load_unrecorded_installation_items():
    """載入比較後新增、且尚未標記為已記錄的項目。"""
    comparison_worksheet = get_installation_comparison_worksheet()
    unrecorded_items = []
    for sheet_row, row in enumerate(comparison_worksheet.get_all_records(), start=2):
        if (
            str(row.get("紀錄類型", "")).strip() != "新增項目"
            or str(row.get("處理狀態", "")).strip() != "未紀錄"
        ):
            continue
        try:
            saved_record = json.loads(str(row.get("資料JSON", "")))
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
        if not isinstance(saved_record, dict):
            continue
        unrecorded_items.append({
            "工作表列": sheet_row,
            "比較ID": str(row.get("比較ID", "")).strip(),
            "比較時間": str(row.get("比較時間", "")).strip(),
            "新版本檔名": str(row.get("新版本檔名", "")).strip(),
            **installation_excel_display_row(saved_record),
            "紀錄狀態": "未紀錄",
        })
    return unrecorded_items


def mark_installation_items_recorded(sheet_rows):
    """將指定比較紀錄列批次標記為已記錄。"""
    cleaned_rows = sorted({int(row_number) for row_number in sheet_rows})
    if not cleaned_rows:
        return 0
    comparison_worksheet = get_installation_comparison_worksheet()
    processed_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    processor = f"{st.session_state.user_name} ({st.session_state.user_role})"
    comparison_worksheet.batch_update(
        [
            {
                "range": f"K{row_number}:M{row_number}",
                "values": [["已記錄", processed_time, processor]],
            }
            for row_number in cleaned_rows
        ],
        value_input_option="USER_ENTERED",
    )
    load_unrecorded_installation_items.clear()
    load_installation_excel_comparisons.clear()
    st.session_state.installation_excel_last_result = None
    try:
        ws_log.append_row(
            [
                processed_time,
                processor,
                f"將 {len(cleaned_rows)} 筆客戶訂單新增項目標記為已記錄",
                "未紀錄",
                "已記錄",
            ],
            table_range="A:E",
        )
    except Exception:
        pass
    return len(cleaned_rows)


@st.dialog("📋 未紀錄項目", width="large")
def show_unrecorded_installation_items_dialog():
    """集中顯示未紀錄項目，供使用者勾選後批次完成記錄。"""
    unrecorded_items = load_unrecorded_installation_items()
    if not unrecorded_items:
        st.success("目前沒有未紀錄項目。")
        return

    st.caption(f"目前共有 {len(unrecorded_items)} 筆未紀錄項目，請勾選已完成記錄的資料。")
    editor_rows = []
    for item in unrecorded_items:
        editor_rows.append({
            "勾選": False,
            "工作表列": item["工作表列"],
            "比較時間": item["比較時間"],
            "來源檔案": item["新版本檔名"],
            "訂單": item["訂單"],
            "客戶簡稱": item["客戶簡稱"],
            "業務員名稱": item["業務員名稱"],
            "品號": item["品號"],
            "品名": item["品名"],
            "訂單數量": item["訂單數量"],
            "未交數量": item["未交數量"],
            "已交數量": item["已交數量"],
            "狀態": item["紀錄狀態"],
        })
    editor_df = pd.DataFrame(editor_rows)
    edited_df = st.data_editor(
        editor_df,
        hide_index=True,
        use_container_width=True,
        disabled=[column for column in editor_df.columns if column != "勾選"],
        column_config={
            "勾選": st.column_config.CheckboxColumn("勾選", default=False),
            "工作表列": None,
        },
        key=f"installation_unrecorded_editor_{st.session_state.installation_unrecorded_grid_key}",
    )
    selected_rows = edited_df.loc[edited_df["勾選"], "工作表列"].tolist()
    if st.button(
        f"✅ 將選取的 {len(selected_rows)} 筆改為已記錄",
        type="primary",
        use_container_width=True,
        disabled=not selected_rows,
        key=f"installation_unrecorded_save_{st.session_state.installation_unrecorded_grid_key}",
    ):
        try:
            updated_count = mark_installation_items_recorded(selected_rows)
            st.session_state.installation_unrecorded_flash = (
                f"已將 {updated_count} 筆項目改為「已記錄」。"
            )
            st.session_state.installation_unrecorded_grid_key += 1
            st.rerun()
        except Exception as error:
            st.error(f"狀態更新失敗：{error}")


def build_installation_comparison_excel(comparison):
    """將單次比較結果輸出為含摘要、三種差異分頁的 Excel。"""
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    title_format = workbook.add_format({
        "bold": True,
        "font_size": 16,
        "font_color": "#FFFFFF",
        "bg_color": "#1F4E78",
        "align": "center",
        "valign": "vcenter",
    })
    label_format = workbook.add_format({"bold": True, "bg_color": "#D9EAF7"})
    integer_format = workbook.add_format({"num_format": "#,##0", "valign": "top"})
    date_format = workbook.add_format({"num_format": "yyyy-mm-dd", "valign": "top"})
    wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})

    summary_sheet = workbook.add_worksheet("比較摘要")
    summary_sheet.hide_gridlines(2)
    summary_sheet.merge_range("A1:D2", "客戶訂單明細比較結果", title_format)
    summary_rows = [
        ("比較時間", comparison["比較時間"]),
        ("比較者", comparison["比較者"]),
        ("前一版本", comparison["前版檔名"]),
        ("目前版本", comparison["新版本檔名"]),
        ("目標品號", INSTALLATION_TARGET_PART_NUMBER),
        ("新增項目", len(comparison["新增"])),
        ("刪除項目", len(comparison["刪除"])),
        ("數量變更", len(comparison["數量變更"])),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.write(row_index, 0, label, label_format)
        if isinstance(value, int):
            summary_sheet.write_number(row_index, 1, value, integer_format)
        else:
            summary_sheet.write(row_index, 1, value)
    summary_sheet.set_column("A:A", 16)
    summary_sheet.set_column("B:B", 55)

    def write_detail_sheet(sheet_name, records, headers, header_color):
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.hide_gridlines(2)
        header_format = workbook.add_format({
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": header_color,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
        })
        for column_index, header in enumerate(headers):
            worksheet.write(0, column_index, header, header_format)
        for row_index, record in enumerate(records, start=1):
            for column_index, header in enumerate(headers):
                value = record.get(header, "")
                if header == "日期" and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(value)):
                    worksheet.write_datetime(
                        row_index,
                        column_index,
                        datetime.strptime(str(value), "%Y-%m-%d"),
                        date_format,
                    )
                elif "數量" in header and re.fullmatch(r"-?\d+(?:\.\d+)?", str(value)):
                    worksheet.write_number(row_index, column_index, float(value), integer_format)
                else:
                    worksheet.write(row_index, column_index, value, wrap_format)
        worksheet.freeze_panes(1, 0)
        if records:
            worksheet.autofilter(0, 0, len(records), len(headers) - 1)
        for column_index, header in enumerate(headers):
            if header == "品名":
                width = 55
            elif header == "訂單":
                width = 32
            elif header in {"客戶簡稱", "業務員名稱"}:
                width = 22
            elif "數量" in header:
                width = 14
            else:
                width = 16
            worksheet.set_column(column_index, column_index, width)

    write_detail_sheet(
        "新增項目",
        comparison["新增"],
        [*INSTALLATION_ORDER_FIELDS, "紀錄狀態"],
        "#548235",
    )
    write_detail_sheet("刪除項目", comparison["刪除"], INSTALLATION_ORDER_FIELDS, "#C00000")
    quantity_headers = [
        *INSTALLATION_IDENTITY_FIELDS,
        *[f"原{field_name}" for field_name in INSTALLATION_QUANTITY_FIELDS],
        *[f"新{field_name}" for field_name in INSTALLATION_QUANTITY_FIELDS],
    ]
    write_detail_sheet("數量變更", comparison["數量變更"], quantity_headers, "#BF9000")
    workbook.close()
    output.seek(0)
    return output.getvalue()


def render_installation_comparison_tables(comparison, key_prefix):
    """顯示一次比較的三類差異並提供 Excel 下載。"""
    metric_columns = st.columns(3)
    metric_columns[0].metric("新增項目", len(comparison["新增"]))
    metric_columns[1].metric("刪除項目", len(comparison["刪除"]))
    metric_columns[2].metric("數量變更", len(comparison["數量變更"]))

    added_tab, removed_tab, quantity_tab = st.tabs(["新增項目", "刪除項目", "數量變更"])
    with added_tab:
        if comparison["新增"]:
            st.dataframe(pd.DataFrame(comparison["新增"]), hide_index=True, use_container_width=True)
        else:
            st.info("沒有新增項目。")
    with removed_tab:
        if comparison["刪除"]:
            st.dataframe(pd.DataFrame(comparison["刪除"]), hide_index=True, use_container_width=True)
        else:
            st.info("沒有刪除項目。")
    with quantity_tab:
        if comparison["數量變更"]:
            st.dataframe(
                pd.DataFrame(comparison["數量變更"]),
                hide_index=True,
                use_container_width=True,
            )
        else:
            st.info("訂單數量、未交數量與已交數量都沒有變動。")

    safe_time = re.sub(r"\D", "", comparison["比較時間"])[:14]
    st.download_button(
        "📥 匯出這次比較結果",
        data=build_installation_comparison_excel(comparison),
        file_name=f"裝機訂單比較_{safe_time or comparison['比較ID']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"{key_prefix}_download",
    )


def render_installation_excel_version_area():
    """記錄客戶訂單明細版本，依 PM-0001-00 關聯訂單比較。"""
    st.divider()
    st.markdown("#### 客戶訂單明細 Excel 比較")
    st.caption(
        "先找出品號 PM-0001-00，再忽略訂單末四碼，記錄所有同訂單主號的項目；"
        "第二次起自動與上一份檔案比較。比較結果永久保存，原始解析資料只保留最新一版。"
    )

    order_number = "客戶訂單明細表"
    part_number = INSTALLATION_TARGET_PART_NUMBER

    try:
        versions = load_installation_excel_versions(order_number, part_number)
        comparison_history = load_installation_excel_comparisons()
        unrecorded_items = load_unrecorded_installation_items()
    except Exception as error:
        st.error(f"無法讀取 Excel 版本或比較紀錄：{error}")
        return

    if st.session_state.installation_unrecorded_flash:
        st.success(st.session_state.installation_unrecorded_flash)
        st.session_state.installation_unrecorded_flash = ""
    if st.button(
        f"📋 查看未紀錄項目（{len(unrecorded_items)}）",
        use_container_width=True,
        disabled=not unrecorded_items,
        key="installation_open_unrecorded_items",
    ):
        show_unrecorded_installation_items_dialog()

    if versions:
        history_df = pd.DataFrame([
            {
                "上傳時間": version["上傳時間"],
                "檔名": version["檔名"],
                "上傳者": version["上傳者"],
                "資料列數": len(version["資料"]),
            }
            for version in reversed(versions)
        ])
        with st.expander(f"歷史版本（{len(versions)} 版）", expanded=False):
            st.dataframe(history_df, hide_index=True, use_container_width=True)
            history_index = st.selectbox(
                "查看歷史版本內容",
                range(len(versions)),
                index=len(versions) - 1,
                format_func=lambda index: (
                    f"{versions[index]['上傳時間']}｜{versions[index]['檔名']}"
                ),
                key=f"installation_excel_history_{order_number}_{part_number}",
            )
            history_rows = versions[history_index]["資料"]
            if history_rows:
                st.dataframe(
                    pd.DataFrame([
                        {
                            **installation_excel_display_row(row),
                        }
                        for row in history_rows
                    ]),
                    hide_index=True,
                    use_container_width=True,
                )
    else:
        st.info("此訂單與品號尚未保存 Excel；第一次讀取會建立比較基準。")

    uploaded_excel = st.file_uploader(
        "選擇 Excel 檔案",
        type=["xlsx", "xlsm"],
        help="支援 .xlsx 與 .xlsm；程式會自動尋找包含所需九個欄位的工作表。",
        key=f"installation_excel_upload_{st.session_state.installation_excel_uploader_key}",
    )
    if st.button(
        "讀取、保存並比較",
        type="primary",
        use_container_width=True,
        disabled=uploaded_excel is None,
        key="installation_excel_save",
    ):
        try:
            if uploaded_excel.size > 10 * 1024 * 1024:
                raise ValueError("Excel 檔案不可超過 10 MB。")
            parsed_result = parse_installation_excel(uploaded_excel.getvalue())
            parsed_rows = parsed_result["資料"]
            if not parsed_rows:
                raise ValueError("找不到符合訂單主號規則的資料。")
            previous_version = versions[-1] if versions else None
            if previous_version:
                added_rows, removed_rows, quantity_changes = compare_installation_excel_rows(
                    previous_version["資料"],
                    parsed_rows,
                )
            else:
                added_rows, removed_rows, quantity_changes = [], [], []
            saved_version = save_installation_excel_version(
                order_number,
                part_number,
                uploaded_excel.name,
                parsed_rows,
            )
            saved_comparison = None
            cleanup_result = {"刪除列數": 0, "刪除版本數": 0}
            cleanup_error = ""
            if previous_version:
                saved_comparison = save_installation_excel_comparison(
                    previous_version,
                    saved_version,
                    added_rows,
                    removed_rows,
                    quantity_changes,
                )
                try:
                    cleanup_result = prune_installation_excel_versions(
                        order_number,
                        part_number,
                        saved_version["版本ID"],
                        loaded_versions=versions,
                    )
                except Exception as cleanup_exception:
                    cleanup_error = str(cleanup_exception)
            # 寫入完成後才清除短期快取，下一次重跑只需各讀取一次最新資料。
            load_installation_excel_versions.clear()
            if saved_comparison:
                load_installation_excel_comparisons.clear()
            st.session_state.installation_excel_last_result = {
                "訂單": order_number,
                "品號": part_number,
                "版本ID": saved_version["版本ID"],
                "檔名": uploaded_excel.name,
                "上一版本": previous_version["檔名"] if previous_version else "",
                "新增": added_rows,
                "刪除": removed_rows,
                "數量變更": quantity_changes,
                "完整資料": [
                    installation_excel_display_row(row)
                    for row in parsed_rows
                ],
                "目標品號筆數": parsed_result["目標品號筆數"],
                "訂單主號數": len(parsed_result["訂單主號"]),
                "資料工作表": parsed_result["資料工作表"],
                "清理結果": cleanup_result,
                "清理錯誤": cleanup_error,
            }
            if saved_comparison:
                st.session_state.installation_excel_last_result.update(saved_comparison)
            st.session_state.installation_excel_uploader_key += 1
            st.rerun()
        except Exception as error:
            st.error(f"Excel 讀取或保存失敗：{error}")

    result = st.session_state.installation_excel_last_result
    if result and result["訂單"] == order_number and result["品號"] == part_number:
        result.setdefault("數量變更", [])
        result.setdefault("比較ID", result.get("版本ID", "舊版暫存"))
        result.setdefault("比較時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        result.setdefault(
            "比較者",
            f"{st.session_state.user_name} ({st.session_state.user_role})",
        )
        result.setdefault("前版檔名", result.get("上一版本", ""))
        result.setdefault("新版本檔名", result.get("檔名", ""))
        st.success(
            f"已保存「{result['檔名']}」：找到 {result['目標品號筆數']} 筆 "
            f"{INSTALLATION_TARGET_PART_NUMBER}、{result['訂單主號數']} 個訂單主號，"
            f"共記錄 {len(result['完整資料'])} 筆同訂單項目。"
        )
        if result.get("清理錯誤"):
            st.warning(
                "比較結果與本次基準已保存，但舊基準資料刪除失敗；"
                f"下次比較時會再次清理。詳細錯誤：{result['清理錯誤']}"
            )
        elif result.get("上一版本") and "清理結果" in result:
            cleanup_result = result.get("清理結果", {})
            st.caption(
                f"已採用滾動基準：清除 {cleanup_result.get('刪除版本數', 0)} 個舊版本、"
                f"{cleanup_result.get('刪除列數', 0)} 列解析資料；目前只保留本次版本供下次比較。"
            )
        if result["上一版本"]:
            st.markdown(f"##### 與上一版「{result['上一版本']}」的比較")
            render_installation_comparison_tables(
                result,
                f"latest_comparison_{result.get('比較ID', result['版本ID'])}",
            )
        else:
            st.info("這是第一版，已建立比較基準；下次讀取會列出新增、刪除與數量變更。")

        with st.expander("查看本次讀取的完整資料", expanded=False):
            st.dataframe(pd.DataFrame(result["完整資料"]), hide_index=True, use_container_width=True)

    if comparison_history:
        st.divider()
        with st.expander(f"歷史比較紀錄（{len(comparison_history)} 次）", expanded=False):
            comparison_index = st.selectbox(
                "選擇比較紀錄",
                range(len(comparison_history)),
                index=len(comparison_history) - 1,
                format_func=lambda index: (
                    f"{comparison_history[index]['比較時間']}｜"
                    f"{comparison_history[index]['前版檔名']} → "
                    f"{comparison_history[index]['新版本檔名']}"
                ),
                key="installation_comparison_history_select",
            )
            selected_comparison = comparison_history[comparison_index]
            st.caption(
                f"比較者：{selected_comparison['比較者']}｜"
                f"比較編號：{selected_comparison['比較ID']}"
            )
            render_installation_comparison_tables(
                selected_comparison,
                f"history_comparison_{selected_comparison['比較ID']}",
            )


def render_installation_confirmation_area(can_download):
    """只顯示已由背鍋俠確認的開發資料。"""
    st.markdown("### 裝機確認區")
    st.caption("此區只顯示背鍋俠已確認的訂單、品號與開發附件。")

    confirmed_records = [
        (index, record)
        for index, record in enumerate(st.session_state.dev_development_records)
        if str(record.get("背鍋俠確認時間", "")).strip()
    ]
    if not confirmed_records:
        st.info("目前沒有已確認的裝機資料。")
        return

    installation_df = pd.DataFrame([
        {
            "訂單": record.get("訂單", ""),
            "品號": record.get("品號", ""),
            "案件名稱": record.get("案件", ""),
            "附件檔名": record.get("附件檔名", ""),
            "背鍋俠確認時間": record.get("背鍋俠確認時間", ""),
            "背鍋俠確認人": record.get("背鍋俠確認人", ""),
        }
        for _, record in confirmed_records
    ])
    st.dataframe(installation_df, hide_index=True, use_container_width=True)

    for record_index, record in reversed(confirmed_records):
        with st.expander(
            f"{record.get('訂單', '')}｜{record.get('品號', '')}｜"
            f"{record.get('背鍋俠確認時間', '')}",
            expanded=False,
        ):
            st.markdown(f"**確認人：** {record.get('背鍋俠確認人', '')}")
            attachment_name = str(record.get("附件檔名", "")).strip()
            if can_download:
                render_secure_attachment_download(record, f"installation_{record_index}")
            else:
                st.write(f"📎 {attachment_name or '（無附件名稱）'}")
                st.caption("目前帳號沒有下載附件的權限。")

            current_case = str(record.get("案件", "")).strip()
            matched_current_case = find_matching_case_name(current_case)
            case_options = list(st.session_state.dev_case_options)
            case_default_index = None
            if matched_current_case in case_options:
                case_default_index = case_options.index(matched_current_case)
            assigned_case = st.selectbox(
                "案件名稱 *",
                case_options,
                index=case_default_index,
                placeholder="選擇既有案件或直接輸入新案件名稱",
                accept_new_options=True,
                help="可替已確認資料補上或修改案件名稱；同名案件會帶入既有確認項目。",
                key=f"installation_case_{record_index}",
            )
            cleaned_case = str(assigned_case or "").strip()
            matched_case = find_matching_case_name(cleaned_case)
            if matched_case:
                matched_items = st.session_state.dev_case_checklists.get(matched_case, [])
                st.success(
                    f"找到同名案件「{matched_case}」，目前有 {len(matched_items)} 行確認項目設定。"
                )
                if matched_items:
                    with st.expander("預覽既有確認項目", expanded=False):
                        st.text("\n".join(matched_items))
            elif cleaned_case:
                st.info(f"尚無案件「{cleaned_case}」，確認後會建立新案件。")

            if st.button(
                "✅ 確認案件名稱並設定確認項目",
                type="primary",
                use_container_width=True,
                disabled=not bool(cleaned_case),
                key=f"installation_checklist_{record_index}",
            ):
                if not st.session_state.get("user_permissions", {}).get(
                    "handoff_access",
                    st.session_state.get("user_role") == "管理者",
                ):
                    st.error("目前帳號沒有設定裝機確認案件的權限。")
                else:
                    final_case = matched_case or cleaned_case
                    if not matched_case:
                        st.session_state.dev_case_options.append(final_case)
                        st.session_state.dev_case_options = sorted(
                            set(st.session_state.dev_case_options)
                        )
                        st.session_state.dev_case_checklists.setdefault(final_case, [])
                    record["案件"] = final_case
                    action_message = (
                        f"已將裝機確認資料指定為案件「{final_case}」："
                        f"訂單「{record.get('訂單', '')}」、品號「{record.get('品號', '')}」"
                    )
                    queue_dev_auto_sync(action_message)
                    log_dev_delete_action(
                        action_message,
                        current_case or "未命名",
                        final_case,
                    )
                    queue_checklist_navigation(final_case)
                    st.rerun()
            render_development_delete_button(record_index, "installation_confirmation")

# ===================================================================
# ==================== 2. 登入系統與權限驗證 ====================
if not st.session_state.logged_in:
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        # 🚀 透過自訂 CSS 容器包裹 st.image，利用 margin-left 完美對齊下方文字中心
        st.markdown(
            """
            <style>
            .centered-logo {
                display: flex;
                justify-content: center;
                margin-left: 20px; /* 根據視覺微調向右偏移 */
                margin-bottom: -10px;
            }
            </style>
            <div class="centered-logo">
            """,
            unsafe_allow_html=True
        )
        
        try:
            st.image("logo.jpg", width=300)
        except:
            st.warning("⚠️ 找不到 logo.jpg 圖片檔案。")
            
        st.markdown("</div>", unsafe_allow_html=True)
            
        st.markdown("<h2 style='text-align: center; margin-top: 5px;'>鴻伍裝機日報系統</h2>", unsafe_allow_html=True)
        st.markdown("### 🔐 系統登入")
        
        with st.form("login_form"):
            user_id = st.text_input("帳號 (ID)")
            user_pwd = st.text_input("密碼", type="password")
            submitted = st.form_submit_button("登入系統", use_container_width=True)
            
            if submitted:
                if not user_id or not user_pwd:
                    st.warning("請輸入帳號與密碼。")
                else:
                    ensure_account_permission_columns()
                    accounts_data = ws_accounts.get_all_records()
                    df_acc = pd.DataFrame(accounts_data)
                    
                    if not df_acc.empty and '帳號' in df_acc.columns and '密碼' in df_acc.columns:
                        df_acc['帳號'] = df_acc['帳號'].astype(str)
                        df_acc['密碼'] = df_acc['密碼'].astype(str)
                        
                        match = df_acc[(df_acc['帳號'] == user_id) & (df_acc['密碼'] == user_pwd)]
                        if not match.empty:
                            account_row = match.iloc[0].to_dict()
                            st.session_state.logged_in = True
                            st.session_state.user_id = user_id
                            st.session_state.user_name = account_row.get('姓名', '未知使用者')
                            st.session_state.user_role = account_row.get('權限', '未設定權限')
                            st.session_state.user_permissions = permissions_for_account(
                                account_row,
                                st.session_state.user_role,
                            )
                            st.rerun()
                        else:
                            st.error("帳號或密碼錯誤，請重新輸入。")
                    else:
                        st.error("『帳號管理』分頁中缺少必要欄位或無資料。")
        
        # 🚀 訪客快速登入按鈕
        if st.button("🚀 訪客快速登入 (公用權限)", use_container_width=True, type="secondary"):
            st.session_state.logged_in = True
            st.session_state.user_id = "guest"
            st.session_state.user_name = "訪客"
            st.session_state.user_role = "公用"
            st.session_state.user_permissions = permissions_for_account(role="公用")
            st.rerun()

    st.stop() # 阻擋未登入者往下執行

# 權限定義
can_edit = st.session_state.user_role == "管理者"
can_add = st.session_state.user_role in ["管理者", "工程師", "業務", "RD"]
can_access_sales = bool(st.session_state.user_permissions.get("sales_access", can_edit))
can_access_development = bool(
    st.session_state.user_permissions.get("development_access", can_edit)
)
can_access_handoff = bool(
    st.session_state.user_permissions.get("handoff_access", can_edit)
)
can_access_installation = bool(
    st.session_state.user_permissions.get("installation_access", can_edit)
)
can_upload_attachment = bool(
    st.session_state.user_permissions.get("attachment_upload", can_edit)
)
can_download_attachment = bool(
    st.session_state.user_permissions.get("attachment_download", can_edit)
)
can_delete_attachment = bool(
    st.session_state.user_permissions.get("attachment_delete", can_edit)
)

# 側邊欄狀態
with st.sidebar:
    try:
        st.image("logo.jpg", width=200)
    except:
        pass
    st.divider()
    st.header("👤 帳號資訊")
    st.markdown(f"**姓名：** {st.session_state.user_name}")
    st.markdown(f"**權限：** {st.session_state.user_role}")
    st.divider()
    if st.button("🚪 登出系統", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.user_id = ""
        st.session_state.user_name = ""
        st.session_state.user_role = ""
        st.session_state.user_permissions = {}
        st.rerun()

st.title("📊 鴻伍裝機日報系統 (Web 雲端版)")

# ==================== 3. 彈出視窗功能 (Dialog) ====================
def parse_checklist_summary(summary):
    items = [item.strip() for item in str(summary or "").split("、") if item.strip()]
    completed = [item.removeprefix("✅").strip() for item in items if item.startswith("✅")]
    incomplete = [item.removeprefix("❌").strip() for item in items if item.startswith("❌")]
    return completed, incomplete


def format_incomplete_reason(reason):
    """將 1.、2.、A.、B. 等分項標記整理成各自一行。"""
    reason_text = str(reason or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not reason_text:
        return ""

    # 標記後必須接非數字內容，避免將 1.25、2026.08.14 等數值或日期拆開。
    item_marker = re.compile(
        r"(\d{1,2}[.、)](?=\s*[^\d.\s])|(?<![A-Za-z])[A-Z][.、)](?=\s*[^.\s]))"
    )
    formatted_text = item_marker.sub(lambda match: f"\n{match.group(1)} ", reason_text)
    return "\n".join(
        line.strip()
        for line in formatted_text.splitlines()
        if line.strip()
    )


def parse_checklist_definition(checklist_lines):
    """解析 [主分類]、[[子分類]] 與其確認項目。"""
    groups = []
    current_group = None
    current_subgroup = None

    def ensure_group():
        nonlocal current_group
        if current_group is None:
            current_group = {"name": "", "items": [], "subgroups": []}
            groups.append(current_group)
        return current_group

    for raw_line in checklist_lines:
        line = str(raw_line).strip()
        if not line:
            continue
        if line.startswith("[[") and line.endswith("]]") and len(line) > 4:
            parent_group = ensure_group()
            current_subgroup = {"name": line[2:-2].strip(), "items": []}
            parent_group["subgroups"].append(current_subgroup)
        elif line.startswith("[") and line.endswith("]") and len(line) > 2:
            current_group = {"name": line[1:-1].strip(), "items": [], "subgroups": []}
            groups.append(current_group)
            current_subgroup = None
        else:
            parent_group = ensure_group()
            if current_subgroup is not None:
                current_subgroup["items"].append(line)
            else:
                parent_group["items"].append(line)

    return [
        group for group in groups
        if group["items"] or any(subgroup["items"] for subgroup in group["subgroups"])
    ]


def normalize_checklist_definition_lines(checklist_text):
    """整理確認項目，僅移除同一分類路徑內的重複項目。"""
    cleaned_lines = []
    seen_item_paths = set()
    current_group = ""
    current_subgroup = ""

    for raw_line in str(checklist_text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("[[") and line.endswith("]]") and len(line) > 4:
            current_subgroup = line[2:-2].strip()
            cleaned_lines.append(f"[[{current_subgroup}]]")
        elif line.startswith("[") and line.endswith("]") and len(line) > 2:
            current_group = line[1:-1].strip()
            current_subgroup = ""
            cleaned_lines.append(f"[{current_group}]")
        else:
            item_path = (current_group, current_subgroup, line)
            if item_path not in seen_item_paths:
                cleaned_lines.append(line)
                seen_item_paths.add(item_path)

    return cleaned_lines


def format_checklist_progress(summary):
    completed, incomplete = parse_checklist_summary(summary)
    total = len(completed) + len(incomplete)
    return f"已施工 {len(completed)}/{total}" if total else "未設定"


def group_checklist_paths(items):
    """將 A / B / C 格式整理為主分類、子分類及項目。"""
    grouped = {}
    for raw_item in items:
        path_parts = [part.strip() for part in str(raw_item).split(" / ") if part.strip()]
        if len(path_parts) >= 3:
            category_name = path_parts[0]
            subcategory_name = path_parts[1]
            item_name = " / ".join(path_parts[2:])
        elif len(path_parts) == 2:
            category_name = path_parts[0]
            subcategory_name = ""
            item_name = path_parts[1]
        else:
            category_name = "其他項目"
            subcategory_name = ""
            item_name = path_parts[0] if path_parts else str(raw_item)

        grouped.setdefault(category_name, {}).setdefault(subcategory_name, []).append(item_name)
    return grouped


def render_grouped_checklist(items, icon):
    grouped_items = group_checklist_paths(items)
    for category_name, subcategories in grouped_items.items():
        category_count = sum(len(group_items) for group_items in subcategories.values())
        st.markdown(f"**📂 {category_name}（{category_count}）**")
        for subcategory_name, group_items in subcategories.items():
            compact_items = "　".join(f"{icon} {item_name}" for item_name in group_items)
            if subcategory_name:
                st.markdown(f"**↳ {subcategory_name}：** {compact_items}")
            elif category_name == "其他項目":
                for item_name in group_items:
                    st.markdown(f"- {icon} {item_name}")
            else:
                st.markdown(compact_items)


def get_checklist_export_labels(checklist_lines):
    """依主分類、子分類順序產生 Excel 欄位名稱。"""
    labels = []
    for category_group in parse_checklist_definition(checklist_lines):
        category_name = category_group["name"]
        for item_name in category_group["items"]:
            labels.append(f"{category_name} / {item_name}" if category_name else item_name)
        for subgroup in category_group["subgroups"]:
            for item_name in subgroup["items"]:
                path_parts = [category_name, subgroup["name"], item_name]
                labels.append(" / ".join(part for part in path_parts if part))
    return labels


def toggle_checklist_group_state(widget_keys, default_values=None):
    """切換一個分類的全部核取狀態；已全選時再次點擊會全部取消。"""
    default_values = default_values or {}
    all_checked = bool(widget_keys) and all(
        bool(st.session_state.get(widget_key, default_values.get(widget_key, False)))
        for widget_key in widget_keys
    )
    for widget_key in widget_keys:
        st.session_state[widget_key] = not all_checked


def checklist_group_is_selected(widget_keys, default_values=None):
    """判斷分類內的項目是否已全部勾選。"""
    default_values = default_values or {}
    return bool(widget_keys) and all(
        bool(st.session_state.get(widget_key, default_values.get(widget_key, False)))
        for widget_key in widget_keys
    )


def render_checklist_editor(checklist_lines, current_summary, key_prefix):
    """依案件設定呈現分類勾選項目，並帶入既有施工狀態。"""
    completed_items, _ = parse_checklist_summary(current_summary)
    completed_set = {str(item).strip() for item in completed_items}
    checklist_results = {}
    item_index = 0

    for category_group in parse_checklist_definition(checklist_lines):
        category_name = category_group["name"]
        category_item_count = len(category_group["items"]) + sum(
            len(subgroup["items"]) for subgroup in category_group["subgroups"]
        )
        category_widget_keys = [
            f"{key_prefix}_{index}"
            for index in range(item_index, item_index + category_item_count)
        ]
        category_defaults = {}
        default_index = item_index
        for item_name in category_group["items"]:
            item_label = f"{category_name} / {item_name}" if category_name else item_name
            category_defaults[f"{key_prefix}_{default_index}"] = (
                item_label in completed_set or item_name in completed_set
            )
            default_index += 1
        for subgroup in category_group["subgroups"]:
            for item_name in subgroup["items"]:
                path_parts = [category_name, subgroup["name"], item_name]
                item_label = " / ".join(part for part in path_parts if part)
                category_defaults[f"{key_prefix}_{default_index}"] = (
                    item_label in completed_set or item_name in completed_set
                )
                default_index += 1

        def render_group_items():
            nonlocal item_index
            for item_name in category_group["items"]:
                item_label = f"{category_name} / {item_name}" if category_name else item_name
                checklist_results[item_label] = st.checkbox(
                    item_name,
                    value=item_label in completed_set or item_name in completed_set,
                    key=f"{key_prefix}_{item_index}",
                )
                item_index += 1

            for subgroup in category_group["subgroups"]:
                subgroup_widget_keys = [
                    f"{key_prefix}_{index}"
                    for index in range(item_index, item_index + len(subgroup["items"]))
                ]
                subgroup_defaults = {
                    widget_key: category_defaults.get(widget_key, False)
                    for widget_key in subgroup_widget_keys
                }
                subgroup_all_checked = checklist_group_is_selected(
                    subgroup_widget_keys,
                    subgroup_defaults,
                )
                subgroup_title_col, subgroup_action_col = st.columns([5, 1])
                with subgroup_action_col:
                    st.form_submit_button(
                        "☑ 已全選" if subgroup_all_checked else "☐ 全選",
                        key=f"{key_prefix}_select_subgroup_{item_index}",
                        use_container_width=True,
                        on_click=toggle_checklist_group_state,
                        args=(subgroup_widget_keys, subgroup_defaults),
                    )
                with subgroup_title_col:
                    st.markdown(f"**↳ {subgroup['name']}**")
                for item_name in subgroup["items"]:
                    path_parts = [category_name, subgroup["name"], item_name]
                    item_label = " / ".join(part for part in path_parts if part)
                    checklist_results[item_label] = st.checkbox(
                        item_name,
                        value=item_label in completed_set or item_name in completed_set,
                        key=f"{key_prefix}_{item_index}",
                    )
                    item_index += 1

        if category_name:
            category_all_checked = checklist_group_is_selected(
                category_widget_keys,
                category_defaults,
            )
            category_col, category_action_col = st.columns([5, 1])
            with category_action_col:
                st.form_submit_button(
                    "☑ 已全選" if category_all_checked else "☐ 全選",
                    key=f"{key_prefix}_select_category_{item_index}",
                    use_container_width=True,
                    on_click=toggle_checklist_group_state,
                    args=(category_widget_keys, category_defaults),
                )
            with category_col:
                with st.expander(f"📂 {category_name}", expanded=False):
                    render_group_items()
        else:
            category_all_checked = checklist_group_is_selected(
                category_widget_keys,
                category_defaults,
            )
            uncategorized_title_col, uncategorized_action_col = st.columns([5, 1])
            with uncategorized_action_col:
                st.form_submit_button(
                    "☑ 已全選" if category_all_checked else "☐ 全選",
                    key=f"{key_prefix}_select_uncategorized_{item_index}",
                    use_container_width=True,
                    on_click=toggle_checklist_group_state,
                    args=(category_widget_keys, category_defaults),
                )
            with uncategorized_title_col:
                st.markdown("**其他項目**")
            render_group_items()

    return checklist_results


def render_export_checklist_selector(checklist_lines, key_prefix):
    """依輸入時的主分類／子分類顯示 Excel 匯出項目選擇器。"""
    selected_labels = []
    item_index = 0
    for category_group in parse_checklist_definition(checklist_lines):
        category_name = category_group["name"]
        category_item_count = len(category_group["items"]) + sum(
            len(subgroup["items"]) for subgroup in category_group["subgroups"]
        )
        category_widget_keys = [
            f"{key_prefix}_{index}"
            for index in range(item_index, item_index + category_item_count)
        ]
        category_defaults = {widget_key: False for widget_key in category_widget_keys}

        def render_group_items():
            nonlocal item_index
            for item_name in category_group["items"]:
                item_label = f"{category_name} / {item_name}" if category_name else item_name
                if st.checkbox(
                    item_name,
                    value=False,
                    key=f"{key_prefix}_{item_index}",
                ):
                    selected_labels.append(item_label)
                item_index += 1

            for subgroup in category_group["subgroups"]:
                subgroup_widget_keys = [
                    f"{key_prefix}_{index}"
                    for index in range(item_index, item_index + len(subgroup["items"]))
                ]
                subgroup_defaults = {widget_key: False for widget_key in subgroup_widget_keys}
                subgroup_all_checked = checklist_group_is_selected(
                    subgroup_widget_keys,
                    subgroup_defaults,
                )
                subgroup_title_col, subgroup_action_col = st.columns([5, 1])
                with subgroup_title_col:
                    st.markdown(f"**↳ {subgroup['name']}**")
                with subgroup_action_col:
                    st.button(
                        "☑ 已全選" if subgroup_all_checked else "☐ 全選",
                        key=f"{key_prefix}_select_subgroup_{item_index}",
                        use_container_width=True,
                        on_click=toggle_checklist_group_state,
                        args=(subgroup_widget_keys, subgroup_defaults),
                    )
                for item_name in subgroup["items"]:
                    path_parts = [category_name, subgroup["name"], item_name]
                    item_label = " / ".join(part for part in path_parts if part)
                    if st.checkbox(
                        item_name,
                        value=False,
                        key=f"{key_prefix}_{item_index}",
                    ):
                        selected_labels.append(item_label)
                    item_index += 1

        if category_name:
            category_all_checked = checklist_group_is_selected(
                category_widget_keys,
                category_defaults,
            )
            category_col, category_action_col = st.columns([5, 1])
            with category_col:
                with st.expander(f"📂 {category_name}", expanded=False):
                    render_group_items()
            with category_action_col:
                st.button(
                    "☑ 已全選" if category_all_checked else "☐ 全選",
                    key=f"{key_prefix}_select_category_{item_index - category_item_count}",
                    use_container_width=True,
                    on_click=toggle_checklist_group_state,
                    args=(category_widget_keys, category_defaults),
                )
        else:
            category_all_checked = checklist_group_is_selected(
                category_widget_keys,
                category_defaults,
            )
            uncategorized_title_col, uncategorized_action_col = st.columns([5, 1])
            with uncategorized_title_col:
                st.markdown("**其他項目**")
            with uncategorized_action_col:
                st.button(
                    "☑ 已全選" if category_all_checked else "☐ 全選",
                    key=f"{key_prefix}_select_uncategorized_{item_index}",
                    use_container_width=True,
                    on_click=toggle_checklist_group_state,
                    args=(category_widget_keys, category_defaults),
                )
            render_group_items()

    return selected_labels


def build_dev_excel_export(machine_records, selected_checklist_items):
    """建立多機台確認 Excel；內部用完整路徑判斷，標題只顯示末層項目。"""
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    sheet = workbook.add_worksheet("裝機確認")

    common_format = {
        "font_name": "新細明體",
        "font_size": 11,
        "border": 1,
        "align": "center",
        "valign": "vcenter",
    }
    header_format = workbook.add_format({**common_format, "text_wrap": True})
    cell_format = workbook.add_format(common_format)
    date_format = workbook.add_format({**common_format, "num_format": "m/d/yy"})
    completed_format = workbook.add_format({**common_format, "bg_color": "#C6EFCE"})
    pending_format = workbook.add_format({**common_format, "bg_color": "#FFEB9C"})

    checklist_header_names = [
        item_label.split(" / ")[-1].strip()
        for item_label in selected_checklist_items
    ]
    headers = [
        "日期",
        "廠別",
        "案件名稱",
        "機台名",
        *checklist_header_names,
        "狀態",
    ]
    sheet.write_row(0, 0, headers, header_format)

    for row_index, record in enumerate(machine_records, start=1):
        date_value = pd.to_datetime(record.get("日期", ""), errors="coerce")
        if pd.notna(date_value):
            sheet.write_datetime(row_index, 0, date_value.to_pydatetime(), date_format)
        else:
            sheet.write(row_index, 0, str(record.get("日期", "")), cell_format)

        sheet.write(row_index, 1, record.get("廠別", ""), cell_format)
        sheet.write(row_index, 2, record.get("案件", ""), cell_format)
        sheet.write(row_index, 3, record.get("機台名稱", ""), cell_format)

        completed_items, _ = parse_checklist_summary(record.get("項目確認", ""))
        completed_set = set(completed_items)
        legacy_completed_names = {
            item.strip() for item in completed_items if " / " not in item
        }
        for item_offset, item_label in enumerate(selected_checklist_items, start=4):
            item_leaf_name = item_label.split(" / ")[-1].strip()
            is_completed = item_label in completed_set or item_leaf_name in legacy_completed_names
            status_format = completed_format if is_completed else pending_format
            sheet.write_blank(row_index, item_offset, None, status_format)

        status_column = 4 + len(selected_checklist_items)
        sheet.write(row_index, status_column, record.get("狀態", ""), cell_format)
        incomplete_reason = format_incomplete_reason(
            record.get("未完成或缺貨原因", "")
        )
        if incomplete_reason:
            sheet.write_comment(
                row_index,
                status_column,
                incomplete_reason,
                {
                    "author": "裝機日報",
                    "width": 300,
                    "height": 120,
                    "visible": False,
                },
            )
        sheet.set_row(row_index, 28)

    last_column = len(headers) - 1
    sheet.autofilter(0, 0, max(len(machine_records), 1), last_column)
    sheet.freeze_panes(1, 4)
    sheet.set_row(0, 40)
    sheet.set_column(0, 0, 12)
    sheet.set_column(1, 1, 12)
    sheet.set_column(2, 3, 18)
    if selected_checklist_items:
        sheet.set_column(4, 3 + len(selected_checklist_items), 18)
    status_column = 4 + len(selected_checklist_items)
    sheet.set_column(status_column, status_column, 12)

    workbook.close()
    output.seek(0)
    return output.getvalue()


def reset_dialog_selection():
    reset_key = st.session_state.get("active_dialog_reset_key")
    if reset_key and reset_key in st.session_state:
        st.session_state[reset_key] += 1


def adjust_photo_rotation(rotation_key, degrees):
    """調整照片顯示角度，不修改 Drive 原始檔案。"""
    st.session_state[rotation_key] = (
        int(st.session_state.get(rotation_key, 0)) + degrees
    ) % 360


def toggle_photo_mirror(mirror_key):
    """切換照片水平鏡像顯示。"""
    st.session_state[mirror_key] = not bool(st.session_state.get(mirror_key, False))


def transform_photo_for_display(photo_bytes, rotation_degrees=0, mirrored=False):
    """套用 EXIF 方向、旋轉與鏡像後回傳畫面用圖片。"""
    with Image.open(io.BytesIO(photo_bytes)) as source_image:
        display_image = ImageOps.exif_transpose(source_image)
        if mirrored:
            display_image = ImageOps.mirror(display_image)
        if rotation_degrees:
            display_image = display_image.rotate(rotation_degrees, expand=True)
        return display_image.copy()


def render_installation_photo_popover(row_data, installation_photos):
    """在詳細資料對話框上方顯示照片浮動面板，保留原視窗。"""
    with st.popover(
        f"📷 查看裝機照片（{len(installation_photos)}）",
        use_container_width=True,
    ):
        if not st.session_state.get("user_permissions", {}).get(
            "attachment_download",
            st.session_state.get("user_role") == "管理者",
        ):
            st.warning("目前帳號沒有查看照片的權限。")
            return

        for photo_index, photo_record in enumerate(installation_photos, start=1):
            photo_name = str(photo_record.get("附件檔名", "")).strip() or f"裝機照片 {photo_index}"
            file_id = str(photo_record.get("附件ID", "")).strip() or extract_drive_file_id(
                photo_record.get("附件連結", "")
            )
            photo_key = re.sub(
                r"[^A-Za-z0-9_-]",
                "_",
                file_id or f"{row_data.get('紀錄ID', 'record')}_{photo_index}",
            )
            rotation_key = f"installation_photo_rotation_{photo_key}"
            mirror_key = f"installation_photo_mirror_{photo_key}"
            st.markdown(f"**{photo_index}. {photo_name}**")
            control_col1, control_col2, control_col3 = st.columns(3)
            with control_col1:
                st.button(
                    "↶ 左轉 90°",
                    key=f"photo_rotate_left_{photo_key}",
                    use_container_width=True,
                    on_click=adjust_photo_rotation,
                    args=(rotation_key, 90),
                )
            with control_col2:
                st.button(
                    "↷ 右轉 90°",
                    key=f"photo_rotate_right_{photo_key}",
                    use_container_width=True,
                    on_click=adjust_photo_rotation,
                    args=(rotation_key, -90),
                )
            with control_col3:
                st.button(
                    "⇋ 水平翻轉",
                    key=f"photo_mirror_{photo_key}",
                    use_container_width=True,
                    on_click=toggle_photo_mirror,
                    args=(mirror_key,),
                )
            try:
                photo_file = download_private_drive_file(file_id)
                display_image = transform_photo_for_display(
                    photo_file["bytes"],
                    st.session_state.get(rotation_key, 0),
                    st.session_state.get(mirror_key, False),
                )
                st.image(display_image, caption=photo_name, use_container_width=True)
            except Exception as e:
                st.error(f"照片載入失敗：{e}")
            if photo_index < len(installation_photos):
                st.divider()


@st.dialog("📝 詳細資料檢視", on_dismiss=reset_dialog_selection)
def show_details_dialog(row_data, reset_key):
    st.session_state.active_dialog_reset_key = reset_key
    data_source = str(row_data.get("資料來源", "")).strip()
    if data_source:
        if data_source == "舊版" or "唯讀" in data_source:
            st.info(f"資料來源：{data_source}（原始資料唯讀；修改時會另存為新版）")
        else:
            st.caption(f"資料來源：{data_source}")
    st.markdown(f"### 🏭 廠別：{row_data.get('廠別', '')}")
    st.markdown(f"### 💻 機台名稱：{row_data.get('機台名稱', '')}")
    st.markdown(f"**📅 日期：** {row_data.get('日期', '')}")
    st.markdown(f"**📂 案件：** {row_data.get('案件', '')}")
    st.markdown(f"**👷 安裝人員：** {row_data.get('安裝人員', '')}")
    st.markdown(f"**🚦 狀態：** {row_data.get('狀態', '')}")
    if row_data.get('建立時間', ''):
        st.markdown(f"**🕒 建立時間：** {row_data.get('建立時間', '')}")
    if row_data.get('項目確認', ''):
        st.markdown("### ✅ 項目確認：")
        completed_items, incomplete_items = parse_checklist_summary(row_data.get('項目確認', ''))
        total_items = len(completed_items) + len(incomplete_items)
        if total_items:
            st.caption(
                f"已施工 {len(completed_items)} 項｜待施工 {len(incomplete_items)} 項"
            )
            if incomplete_items:
                with st.expander(f"❌ 待施工項目（{len(incomplete_items)}）", expanded=False):
                    render_grouped_checklist(incomplete_items, "❌")
            if completed_items:
                with st.expander(f"✅ 已施工項目（{len(completed_items)}）", expanded=False):
                    render_grouped_checklist(completed_items, "✅")
        else:
            st.write(row_data.get('項目確認', ''))
    if row_data.get('未完成或缺貨原因', ''):
        st.markdown("### ⚠️ 未完成原因：")
        st.text(format_incomplete_reason(row_data.get('未完成或缺貨原因', '')))
    installation_photos = installation_photos_from_record(row_data)
    if installation_photos:
        render_installation_photo_popover(row_data, installation_photos)
    st.markdown("---")
    st.markdown("### 📝 Remark (備忘內容)：")
    
    remark_text = str(row_data.get('Remark', '')).replace('\n', '<br>')
    if not remark_text.strip():
        remark_text = "（無備忘內容）"
        
    st.markdown(
        f"<div style='font-size: 1.3em; line-height: 1.6; background-color: #f0f2f6; padding: 15px; border-radius: 8px; color: #333; margin-bottom: 20px;'>{remark_text}</div>",
        unsafe_allow_html=True
    )
    
    if st.button("❌ 關閉視窗並取消選取", type="primary", use_container_width=True):
        st.session_state[reset_key] += 1  
        st.rerun()


def dismiss_dev_reason_dialog():
    st.session_state.dev_pending_preview = None


def prepare_dev_preview(record):
    """依狀態決定直接預覽，或先要求輸入原因。"""
    if record.get("狀態") == "未完成":
        st.session_state.dev_add_preview = None
        st.session_state.dev_pending_preview = record
    else:
        record["未完成或缺貨原因"] = ""
        st.session_state.dev_pending_preview = None
        st.session_state.dev_add_preview = record


@st.dialog("⚠️ 記錄原因", on_dismiss=dismiss_dev_reason_dialog)
def show_dev_reason_dialog():
    pending_record = st.session_state.dev_pending_preview
    if not pending_record:
        st.rerun()

    status = pending_record.get("狀態", "")
    st.warning(f"目前狀態為「{status}」，請記錄原因後再繼續。")
    reason = st.text_area(
        "原因 *",
        placeholder="請輸入未完成的原因",
        height=140,
        key=f"dev_reason_{st.session_state.dev_reason_dialog_key}",
    )

    confirm_col, cancel_col = st.columns(2)
    with confirm_col:
        if st.button("確認並產生預覽", type="primary", use_container_width=True):
            cleaned_reason = reason.strip()
            if not cleaned_reason:
                st.error("請輸入原因。")
            else:
                pending_record["未完成或缺貨原因"] = cleaned_reason
                st.session_state.dev_add_preview = pending_record
                st.session_state.dev_pending_preview = None
                st.session_state.dev_reason_dialog_key += 1
                st.rerun()
    with cancel_col:
        if st.button("取消", use_container_width=True):
            st.session_state.dev_pending_preview = None
            st.session_state.dev_reason_dialog_key += 1
            st.rerun()


def dismiss_previous_record_dialog():
    st.session_state.dev_pending_previous_record = None


@st.dialog("🔎 發現未完成紀錄", on_dismiss=dismiss_previous_record_dialog)
def show_previous_record_dialog():
    pending_data = st.session_state.dev_pending_previous_record
    if not pending_data:
        st.rerun()

    identity_data = dict(pending_data["基本資料"])
    previous_record = pending_data["上次資料"]
    st.warning("相同廠別、案件與機台名稱有未完成紀錄，是否帶入上次資料？")
    st.markdown(f"**資料來源：** {previous_record.get('資料來源', '開發區裝機紀錄')}")
    st.markdown(f"**廠別：** {identity_data.get('廠別', '')}")
    st.markdown(f"**案件：** {identity_data.get('案件', '')}")
    st.markdown(f"**機台名稱：** {identity_data.get('機台名稱', '')}")
    st.markdown(
        f"**上次狀態：** "
        f"{previous_record.get('原始狀態', previous_record.get('狀態', ''))}"
    )
    st.markdown("**上次確認項目：**")
    st.write(previous_record.get("項目確認", ""))
    st.markdown("**上次備註：**")
    st.write(previous_record.get("Remark", "") or "（無備註）")

    bring_col, keep_col = st.columns(2)
    with bring_col:
        if st.button("帶入上次資料", type="primary", use_container_width=True):
            st.session_state.dev_identity_draft = identity_data
            st.session_state.dev_loaded_case = identity_data["案件"]
            st.session_state.dev_previous_prefill = {
                "項目確認": previous_record.get("項目確認", ""),
                "狀態": previous_record.get("狀態", "未完成"),
                "Remark": previous_record.get("Remark", ""),
                "資料來源": previous_record.get("資料來源", "開發區裝機紀錄"),
            }
            st.session_state.dev_pending_previous_record = None
            st.session_state.dev_add_form_key += 1
            st.session_state.dev_checklist_key += 1
            st.rerun()
    with keep_col:
        if st.button("不帶入，繼續輸入", use_container_width=True):
            st.session_state.dev_identity_draft = identity_data
            st.session_state.dev_loaded_case = identity_data["案件"]
            st.session_state.dev_previous_prefill = None
            st.session_state.dev_pending_previous_record = None
            st.session_state.dev_add_form_key += 1
            st.session_state.dev_checklist_key += 1
            st.rerun()


def dismiss_dev_delete_dialog():
    st.session_state.dev_pending_delete = None


def log_dev_delete_action(action, old_value, new_value="已刪除"):
    """將開發流程操作寫入既有修改紀錄。"""
    try:
        ws_log.append_row(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                f"{st.session_state.user_name} ({st.session_state.user_role})",
                action,
                old_value,
                new_value,
            ],
            table_range="A:E",
        )
    except Exception:
        pass


@st.dialog("🗑️ 確認刪除", on_dismiss=dismiss_dev_delete_dialog)
def show_dev_delete_dialog():
    pending_delete = st.session_state.dev_pending_delete
    if not pending_delete:
        st.rerun()

    can_delete = st.session_state.get("user_permissions", {}).get(
        "attachment_delete",
        st.session_state.get("user_role") == "管理者",
    )
    if not can_delete:
        st.error("目前帳號沒有刪除資料或附件的權限。")
        return

    record_type = pending_delete.get("type")
    if record_type == "installation":
        record = dict(pending_delete.get("record") or {})
        if not str(record.get("紀錄ID", "")).strip():
            st.error("找不到要刪除的新版裝機資料，可能已被其他操作更新。")
            return

        installation_photos = installation_photos_from_record(record)
        st.markdown(f"**日期：** {record.get('日期', '')}")
        st.markdown(f"**廠別：** {record.get('廠別', '')}")
        st.markdown(f"**案件：** {record.get('案件', '')}")
        st.markdown(f"**機台名稱：** {record.get('機台名稱', '')}")
        st.markdown(f"**照片：** {len(installation_photos)} 張")
        st.error("確認後將刪除新版裝機資料，照片也會移至 Google Drive 垃圾桶。")
        if str(record.get("來源版本", "")).strip() == "舊版轉換":
            st.warning("此筆由舊版轉換；刪除新版後，原本的唯讀舊版紀錄會重新出現在搜尋結果。")

        confirm_col, cancel_col = st.columns(2)
        with confirm_col:
            if st.button(
                "確認刪除資料與照片",
                type="primary",
                use_container_width=True,
                key=f"confirm_installation_delete_{st.session_state.dev_delete_dialog_key}",
            ):
                deleted_photo_names = []
                try:
                    if installation_photos:
                        with st.spinner("正在將裝機照片移至 Google Drive 垃圾桶..."):
                            for photo_record in installation_photos:
                                file_id = str(photo_record.get("附件ID", "")).strip() or extract_drive_file_id(
                                    photo_record.get("附件連結", "")
                                )
                                if not file_id:
                                    raise ValueError(
                                        f"照片「{photo_record.get('附件檔名', '')}」缺少 Drive 檔案 ID。"
                                    )
                                delete_dev_attachment(file_id)
                                deleted_photo_names.append(
                                    str(photo_record.get("附件檔名", "")).strip()
                                )

                    delete_new_installation_record(record)
                    try:
                        download_private_drive_file.clear()
                    except Exception:
                        pass
                    action_message = (
                        f"已刪除新版裝機紀錄：{record.get('廠別', '')}／"
                        f"{record.get('案件', '')}／{record.get('機台名稱', '')}"
                    )
                    if deleted_photo_names:
                        action_message += f"，並刪除 {len(deleted_photo_names)} 張照片"
                    log_dev_delete_action(action_message, str(record))
                    st.session_state.dev_pending_delete = None
                    st.session_state.dev_delete_dialog_key += 1
                    st.session_state.dev_results_edit_form_key += 1
                    st.session_state.dev_results_grid_key += 1
                    st.session_state.dev_flash_level = "success"
                    st.session_state.dev_flash_message = action_message
                    st.rerun()
                except Exception as e:
                    partial_message = (
                        f"已移除 {len(deleted_photo_names)} 張照片，但裝機資料尚未刪除。"
                        if deleted_photo_names else "裝機資料與照片均未刪除。"
                    )
                    st.error(f"刪除失敗。{partial_message} 詳細錯誤：{e}")

        with cancel_col:
            if st.button(
                "取消",
                use_container_width=True,
                key=f"cancel_installation_delete_{st.session_state.dev_delete_dialog_key}",
            ):
                st.session_state.dev_pending_delete = None
                st.session_state.dev_delete_dialog_key += 1
                st.rerun()
        return

    record_index = int(pending_delete.get("index", -1))
    records = (
        st.session_state.dev_sales_records
        if record_type == "sales"
        else st.session_state.dev_development_records
    )
    if record_index < 0 or record_index >= len(records):
        st.error("找不到要刪除的資料，可能已被其他操作更新。")
        return

    record = records[record_index]
    order_number = str(record.get("訂單", "")).strip()
    part_number = str(record.get("品號", "")).strip()
    st.markdown(f"**訂單：** {order_number}")
    st.markdown(f"**品號：** {part_number}")

    if record_type == "sales":
        related_records = [
            item for item in st.session_state.dev_development_records
            if str(item.get("訂單", "")).strip() == order_number
            and str(item.get("品號", "")).strip() == part_number
        ]
        if related_records:
            st.error(
                f"此業務資料仍有 {len(related_records)} 筆開發紀錄，"
                "請先刪除相關開發紀錄後再刪除。"
            )
            if st.button("關閉", use_container_width=True):
                st.session_state.dev_pending_delete = None
                st.rerun()
            return
        st.warning("刪除後會從業務專區移除，此動作無法在 App 中復原。")
    else:
        attachment_name = str(record.get("附件檔名", "")).strip()
        st.markdown(f"**附件：** {attachment_name or '（無附件名稱）'}")
        st.warning("確認後會將 Drive 附件移到垃圾桶，並刪除這筆開發紀錄。")

    confirm_col, cancel_col = st.columns(2)
    with confirm_col:
        if st.button(
            "確認刪除",
            type="primary",
            use_container_width=True,
            key=f"confirm_dev_delete_{st.session_state.dev_delete_dialog_key}",
        ):
            try:
                if record_type == "development":
                    file_id = str(record.get("附件ID", "")).strip() or extract_drive_file_id(
                        record.get("附件連結", "")
                    )
                    with st.spinner("正在將附件移至 Google Drive 垃圾桶..."):
                        delete_dev_attachment(file_id)
                    deleted_record = st.session_state.dev_development_records.pop(record_index)
                    action_message = (
                        f"已刪除開發紀錄：訂單「{order_number}」、品號「{part_number}」"
                    )
                    log_dev_delete_action(action_message, deleted_record.get("附件檔名", ""))
                else:
                    deleted_record = st.session_state.dev_sales_records.pop(record_index)
                    action_message = (
                        f"已刪除業務紀錄：訂單「{order_number}」、品號「{part_number}」"
                    )
                    log_dev_delete_action(action_message, str(deleted_record))

                queue_dev_auto_sync(action_message)
                st.session_state.dev_pending_delete = None
                st.session_state.dev_delete_dialog_key += 1
                st.rerun()
            except Exception as e:
                st.error(f"刪除失敗，原紀錄仍保留。詳細錯誤：{e}")

    with cancel_col:
        if st.button(
            "取消",
            use_container_width=True,
            key=f"cancel_dev_delete_{st.session_state.dev_delete_dialog_key}",
        ):
            st.session_state.dev_pending_delete = None
            st.session_state.dev_delete_dialog_key += 1
            st.rerun()


if st.session_state.dev_pending_delete:
    show_dev_delete_dialog()

# ==================== 4. 建立功能分頁 ====================
installers_list = ["鍾博宇", "黃政欽", "張智偉", "林嬴燦", "吳建華", "何乙霆"]

tab_specs = [
    ("morning", "🌅 晨會當日動態"),
    ("add", "📝 新增裝機紀錄"),
    ("search", "🔍 歷史搜尋與修改"),
    ("tracking", "📌 待追蹤清單 (更新狀態)"),
]
if can_access_sales and not can_edit:
    tab_specs.append(("sales", "💼 業務專區"))
if can_access_development and not can_edit:
    tab_specs.append(("development", "🛠️ 開發專區"))
if can_edit:
    tab_specs.append(("report", "📊 報告專區"))
    tab_specs.append(("dev_admin", "🧪 開發測試區"))

tabs = st.tabs([label for _, label in tab_specs])
tab_map = {tab_key: tab for (tab_key, _), tab in zip(tab_specs, tabs)}
tab1 = tab_map["morning"]
tab_report = tab_map.get("report")
tab2 = tab_map["add"]
tab3 = tab_map["search"]
tab4 = tab_map["tracking"]
tab_sales = tab_map.get("sales")
tab_development = tab_map.get("development")
tab_dev = tab_map.get("dev_admin")

# ==================== 分頁 1：晨會當日動態 ====================
with tab1:
    st.subheader("查詢晨會動態")
    target_date = st.date_input("選擇日期", datetime.now(), key="morning_date")
    
    if st.button("🔍 查詢當日動態", key="btn_morning"):
        with st.spinner('讀取雲端資料中...'):
            data = load_production_installation_records()
            df = pd.DataFrame(data)
            
            if not df.empty:
                df['日期_temp'] = pd.to_datetime(df['日期'], format='mixed', errors='coerce').dt.date
                filtered_df = df[df['日期_temp'] == target_date].copy()
                
                if not filtered_df.empty:
                    filtered_df['日期'] = filtered_df['日期'].astype(str)
                    st.session_state.tab1_filtered_df = filtered_df
                    st.session_state.tab1_search_active = True
                else:
                    st.session_state.tab1_filtered_df = pd.DataFrame()
                    st.session_state.tab1_search_active = True
            else:
                st.warning("試算表中尚無任何資料。")
                st.session_state.tab1_search_active = False

    if st.session_state.tab1_search_active:
        filtered_df = st.session_state.tab1_filtered_df
        if not filtered_df.empty:
            st.success(f"找到 {len(filtered_df)} 筆紀錄 (🖱️ 提示：點擊任意列可彈出詳細資訊)")
            display_cols = ["廠別", "案件", "機台名稱", "安裝人員", "狀態", "Remark"]
            display_cols = [col for col in display_cols if col in filtered_df.columns]
            
            event = st.dataframe(
                filtered_df[display_cols], hide_index=True, use_container_width=True, 
                on_select="rerun", selection_mode="single-row", key=f"tab1_grid_{st.session_state.tab1_grid_key}"
            )
            if event.selection.rows:
                selected_idx = event.selection.rows[0]
                show_details_dialog(filtered_df.iloc[selected_idx], 'tab1_grid_key')
        else:
            st.info(f"📅 該日尚無裝機紀錄。")

# ==================== 管理員專屬：報告專區 ====================
if can_edit and tab_report is not None:
    with tab_report:
        render_report_area()

# ==================== 分頁 2：新增裝機紀錄 ====================
with tab2:
    st.subheader("填寫裝機資訊")
    
    if not can_add:
        st.warning(
            f"⚠️ 目前為「{st.session_state.user_role}」身分，"
            "此帳號僅能使用已授權的專區，無法新增裝機紀錄。"
        )
    else:
        k_suffix = st.session_state.add_form_key
        col1, col2 = st.columns(2)
        
        with col1:
            input_date = st.date_input("裝機日期", datetime.now(), key=f"add_date_{k_suffix}")
            plant = st.text_input("廠別:", key=f"add_plant_{k_suffix}")
            case = st.text_input("案件:", key=f"add_case_{k_suffix}")
            machine = st.text_input("機台名稱:", key=f"add_machine_{k_suffix}")
            
        with col2:
            status = st.selectbox("狀態:", ["未完成", "缺料", "已完成"], key=f"add_status_{k_suffix}")
            installers = st.multiselect("安裝人員 (可複選):", installers_list, key=f"add_installers_{k_suffix}")
            remark = st.text_area("Remark (備忘):", height=130, key=f"add_remark_{k_suffix}")

        b_col1, b_col2 = st.columns(2)
        with b_col1:
            btn_submit = st.button("💾 新增紀錄", type="primary", key="btn_add")
        with b_col2:
            if st.button("🗑️ 清空欄位", key="btn_clear_form"):
                st.session_state.add_form_key += 1
                st.rerun()

        if btn_submit:
            if not plant or not machine:
                st.error("「廠別」與「機台名稱」為必填欄位！")
            else:
                with st.spinner('寫入雲端中...'):
                    installer_str = "\n".join(installers)
                    date_str = input_date.strftime("%Y-%m-%d")
                    
                    headers = worksheet.row_values(1)
                    new_row = [""] * len(headers)
                    def fill_col(col_name, val):
                        if col_name in headers:
                            new_row[headers.index(col_name)] = val

                    fill_col("日期", date_str)
                    fill_col("安裝人員", installer_str)
                    fill_col("廠別", plant)
                    fill_col("案件", case)
                    fill_col("機台名稱", machine)
                    fill_col("狀態", status)
                    fill_col("Remark", remark)
                    
                    worksheet.append_row(new_row)
                    load_production_installation_records.clear()
                    
                    log_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws_log.append_row(
                        [log_time, f"{st.session_state.user_name} ({st.session_state.user_role})", f"新增機台: {machine} (廠別:{plant})", "", "建立新紀錄"],
                        table_range="A:E",
                    )
                    
                    st.success(f"✅ 成功將機台【{machine}】新增至雲端！")

# ==================== 分頁 3：歷史紀錄搜尋與修改 ====================
with tab3:
    st.subheader("🔍 進階條件篩選與修改")
    data = load_production_installation_records()
    df_search = pd.DataFrame(data)
    
    if not df_search.empty:
        df_search = df_search.fillna("")
        df_search['Sheet_Row'] = df_search.index + 2
        
        unique_plants = ["(全部)"] + sorted(
            set(str(x).strip() for x in df_search['廠別'] if str(x).strip()),
            key=natural_plant_sort_key,
        )
        unique_installers = ["(全部)"] + installers_list
        
        st.markdown("##### 1. 設定搜尋條件 (設定完畢後請點擊下方搜尋按鈕)")
        col_s1, col_s2, col_s3, col_s4, col_s5 = st.columns(5)
        
        with col_s1: date_range = st.date_input("選擇日期區間:", [])
        with col_s2: search_plant = st.selectbox("廠別:", unique_plants)
            
        if search_plant != "(全部)":
            target_plant_df = df_search[df_search['廠別'].astype(str).str.strip() == search_plant]
            unique_cases = ["(全部)"] + sorted(list(set([str(x).strip() for x in target_plant_df['案件'] if str(x).strip()])))
            is_case_disabled = False
        else:
            unique_cases = ["(請先選擇廠別)"]
            is_case_disabled, target_plant_df = True, pd.DataFrame()
            
        with col_s3: search_case = st.selectbox("案件 (廠別確定後解鎖):", unique_cases, disabled=is_case_disabled)
            
        if search_plant != "(全部)":
            if search_case not in ["(全部)", "(請先選擇廠別)"]:
                target_case_df = target_plant_df[target_plant_df['案件'].astype(str).str.strip() == search_case]
                unique_machines = ["(全部)"] + sorted(list(set([str(x).strip() for x in target_case_df['機台名稱'] if str(x).strip()])))
            else:
                unique_machines = ["(全部)"] + sorted(list(set([str(x).strip() for x in target_plant_df['機台名稱'] if str(x).strip()])))
            is_machine_disabled = False
        else:
            unique_machines, is_machine_disabled = ["(請先選擇廠別)"], True
            
        with col_s4: search_machine = st.selectbox("機台名稱 (依廠別、案件篩選):", unique_machines, disabled=is_machine_disabled)
        with col_s5: search_installer = st.selectbox("安裝人員:", unique_installers)
            
        if st.button("🔍 開始搜尋", type="primary", key="btn_execute_search"):
            with st.spinner("搜尋中..."):
                filtered_df = df_search.copy()
                if len(date_range) == 2:
                    filtered_df['日期_temp'] = pd.to_datetime(filtered_df['日期'], format='mixed', errors='coerce').dt.date
                    filtered_df = filtered_df[(filtered_df['日期_temp'] >= date_range[0]) & (filtered_df['日期_temp'] <= date_range[1])].drop(columns=['日期_temp'])
                elif len(date_range) == 1:
                    filtered_df['日期_temp'] = pd.to_datetime(filtered_df['日期'], format='mixed', errors='coerce').dt.date
                    filtered_df = filtered_df[filtered_df['日期_temp'] == date_range[0]].drop(columns=['日期_temp'])
                    
                if search_plant != "(全部)": filtered_df = filtered_df[filtered_df['廠別'].astype(str).str.strip() == search_plant]
                if search_machine not in ["(全部)", "(請先選擇廠別)"]: filtered_df = filtered_df[filtered_df['機台名稱'].astype(str).str.strip() == search_machine]
                if search_case not in ["(全部)", "(請先選擇廠別)"]: filtered_df = filtered_df[filtered_df['案件'].astype(str).str.strip() == search_case]
                if search_installer != "(全部)": filtered_df = filtered_df[filtered_df['安裝人員'].astype(str).str.contains(search_installer)]
                
                st.session_state.tab3_filtered_df = filtered_df
                st.session_state.tab3_search_active = True
                st.session_state.tab3_edit_requested = False
                st.session_state.tab3_edit_confirmed = False
                st.rerun()
                
        st.divider()
        
        if st.session_state.tab3_search_active:
            filtered_df = st.session_state.tab3_filtered_df
            st.markdown(f"##### 2. 搜尋結果 (共計 <span style='color:red;'>{len(filtered_df)}</span> 筆)", unsafe_allow_html=True)
            
            if not filtered_df.empty:
                view_cols = ["日期", "廠別", "案件", "機台名稱", "安裝人員", "狀態", "Remark"]
                view_cols = [col for col in view_cols if col in filtered_df.columns]
                
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    filtered_df[view_cols].to_excel(writer, sheet_name='裝機搜尋結果', index=False)
                    workbook, worksheet_excel = writer.book, writer.sheets['裝機搜尋結果']
                    wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top'})
                    default_format = workbook.add_format({'valign': 'top'})
                    for idx, col_name in enumerate(view_cols):
                        width = 45 if col_name == 'Remark' else (20 if col_name == '安裝人員' else 18)
                        worksheet_excel.set_column(idx, idx, width, wrap_format if col_name in ['Remark', '安裝人員'] else default_format)
                buffer.seek(0)
                
                st.download_button(label="📥 匯出搜尋結果為 Excel", data=buffer, file_name=f"鴻伍裝機搜尋結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
                if not st.session_state.tab3_edit_confirmed:
                    event = st.dataframe(filtered_df[view_cols], hide_index=True, use_container_width=True, on_select="rerun", selection_mode="single-row", key=f"tab3_grid_{st.session_state.tab3_grid_key}")
                    if event.selection.rows:
                        show_details_dialog(filtered_df.iloc[event.selection.rows[0]], 'tab3_grid_key')
                    
                    if not st.session_state.tab3_edit_requested:
                        if can_edit:
                            if st.button("✏️ 開啟修改模式"):
                                st.session_state.tab3_edit_requested = True
                                st.rerun()
                        else:
                            st.info(f"💡 您的權限 ({st.session_state.user_role}) 僅供查詢與檢視，修改功能僅限「管理者」。")
                    else:
                        st.warning("⚠️ 即將進入修改，請確認")
                        c1, c2, c3 = st.columns([1, 1, 4])
                        with c1:
                            if st.button("✅ 確認修改", type="primary"):
                                st.session_state.tab3_edit_confirmed = True
                                st.session_state.tab3_edit_requested = False
                                st.rerun()
                        with c2:
                            if st.button("❌ 取消"):
                                st.session_state.tab3_edit_requested = False
                                st.rerun()
                else:
                    st.info("✏️ 編輯模式已開啟，請直接在下方表格修改內容。")
                    edited_df = st.data_editor(filtered_df[view_cols], hide_index=True, use_container_width=True, key="search_editor")
                    
                    if st.button("💾 儲存表格上的所有修改", type="primary"):
                        with st.spinner("正在批次同步更新並寫入日誌..."):
                            changed_cells = []
                            log_entries = []
                            headers = worksheet.row_values(1)
                            
                            for i in range(len(edited_df)):
                                orig_row = filtered_df[view_cols].iloc[i]
                                new_row = edited_df.iloc[i]
                                sheet_row_idx = filtered_df.iloc[i]['Sheet_Row']
                                machine_name = new_row['機台名稱']
                                
                                for col in view_cols:
                                    if str(orig_row[col]).strip() != str(new_row[col]).strip():
                                        if col in headers:
                                            col_idx = headers.index(col) + 1
                                            changed_cells.append(gspread.Cell(int(sheet_row_idx), col_idx, str(new_row[col])))
                                            log_entries.append([
                                                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                                f"{st.session_state.user_name} ({st.session_state.user_role})",
                                                f"{machine_name} (Row {sheet_row_idx}) - {col}",
                                                str(orig_row[col]),
                                                str(new_row[col])
                                            ])
                                            
                            if changed_cells:
                                worksheet.update_cells(changed_cells)
                                load_production_installation_records.clear()
                                ws_log.append_rows(log_entries, table_range="A:E")
                                st.success(f"✅ 成功更新資料，並已記錄 {len(log_entries)} 筆修改日誌！")
                                st.session_state.tab3_search_active = False
                                st.session_state.tab3_edit_confirmed = False
                                st.rerun()
                            else:
                                st.info("沒有偵測到任何修改內容。")
            else:
                st.warning("⚠️ 找不到符合您設定條件的紀錄。")
    else:
        st.info("試算表中尚無資料。")

# ==================== 分頁 4：待追蹤清單與狀態更新 ====================
with tab4:
    st.subheader("📌 待追蹤機台與狀態更新")
    data = load_production_installation_records()
    df = pd.DataFrame(data)
    
    if not df.empty:
        df = df.fillna("")
        df['Sheet_Row'] = df.index + 2
        pending_df = df[(df['狀態'].astype(str).str.strip() != '已完成') & (df['狀態'].astype(str).str.strip() != '') & (df['機台名稱'].astype(str).str.strip() != '')]
        st.markdown(f"**目前待追蹤數量： <span style='color:red;'>{len(pending_df)}</span> 筆**", unsafe_allow_html=True)
        
        if not pending_df.empty:
            display_cols = ["日期", "廠別", "案件", "機台名稱", "安裝人員", "狀態", "Remark"]
            display_cols = [col for col in display_cols if col in pending_df.columns]
            
            event = st.dataframe(pending_df[display_cols], hide_index=True, use_container_width=True, on_select="rerun", selection_mode="single-row", key=f"tab4_grid_{st.session_state.tab4_grid_key}")
            if event.selection.rows:
                show_details_dialog(pending_df.iloc[event.selection.rows[0]], 'tab4_grid_key')
            
            st.divider()
            st.markdown("#### ✏️ 更新機台狀態")
            
            if not can_edit:
                st.info(f"💡 您的權限 ({st.session_state.user_role}) 僅供檢視，狀態更新功能僅限「管理者」操作。")
            else:
                options = pending_df['Sheet_Row'].tolist()
                upd_col1, upd_col2 = st.columns([2, 1])
                with upd_col1:
                    selected_item_row = st.selectbox(
                        "選擇要更新的機台：", options, 
                        format_func=lambda r: f"{pending_df.loc[pending_df['Sheet_Row'] == r, '日期'].values[0]} | {pending_df.loc[pending_df['Sheet_Row'] == r, '廠別'].values[0]} - {pending_df.loc[pending_df['Sheet_Row'] == r, '機台名稱'].values[0]} (目前: {pending_df.loc[pending_df['Sheet_Row'] == r, '狀態'].values[0]})"
                    )
                with upd_col2:
                    new_status = st.selectbox("修改為新狀態：", ["未完成", "缺料", "已完成"], index=2)
                    
                if st.button("送出狀態更新", type="primary", key="btn_update"):
                    with st.spinner("同步至雲端並記錄日誌中..."):
                        row_idx = int(selected_item_row)
                        headers = worksheet.row_values(1)
                        if "狀態" in headers and "Remark" in headers:
                            status_col_idx = headers.index("狀態") + 1
                            remark_col_idx = headers.index("Remark") + 1
                            
                            old_status = pending_df.loc[pending_df['Sheet_Row'] == row_idx, '狀態'].values[0]
                            target_machine = pending_df.loc[pending_df['Sheet_Row'] == row_idx, '機台名稱'].values[0]
                            
                            worksheet.update_cell(row_idx, status_col_idx, new_status)
                            
                            old_remark = worksheet.cell(row_idx, remark_col_idx).value or ""
                            today_str = datetime.now().strftime("%Y-%m-%d")
                            append_text = f"[{today_str} 更新狀態: {new_status}]"
                            new_remark = (str(old_remark).strip() + "\n" + append_text).strip()
                            worksheet.update_cell(row_idx, remark_col_idx, new_remark)
                            load_production_installation_records.clear()
                            
                            log_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            ws_log.append_row(
                                [log_time, f"{st.session_state.user_name} ({st.session_state.user_role})", f"更新機台狀態: {target_machine}", old_status, new_status],
                                table_range="A:E",
                            )
                            
                            st.success(f"✅ 更新成功！該機台已標記為「{new_status}」。")
                            st.rerun()
                        else:
                            st.error("找不到「狀態」或「Remark」欄位。")
        else:
            st.success("🎉 所有機台皆已完工，沒有待追蹤項目。")


# ==================== 依帳號權限顯示業務／開發專區 ====================
if tab_sales is not None:
    with tab_sales:
        initialize_dev_cloud_data()
        if can_access_sales:
            render_sales_area()
        else:
            st.error("目前帳號沒有進入業務專區的權限。")

if tab_development is not None:
    with tab_development:
        initialize_dev_cloud_data()
        if can_access_development:
            render_development_area(can_upload_attachment, can_download_attachment)
        else:
            st.error("目前帳號沒有進入開發專區的權限。")

# ==================== 管理員專屬：開發測試區 ====================
if can_edit and tab_dev is not None:
    with tab_dev:
        st.subheader("🧪 開發測試區")
        st.info(
            f"此區僅供管理員測試；選項與專區資料保存至「{DEV_WORKSHEET_NAME}」，"
            f"新版裝機資料保存至「{NEW_INSTALLATION_WORKSHEET_NAME}」。"
        )

        initialize_dev_cloud_data()
        try:
            migrated_installation_count = migrate_saved_dev_installation_records()
            if migrated_installation_count:
                st.success(
                    f"已將 {migrated_installation_count} 筆既有開發裝機資料轉入「{NEW_INSTALLATION_WORKSHEET_NAME}」。"
                )
        except Exception as e:
            st.error(f"新版裝機資料庫初始化失敗：{e}")

        requested_checklist_case = str(
            st.session_state.dev_checklist_navigation_case or ""
        ).strip()
        if requested_checklist_case:
            # 清除舊的分頁選擇，讓 default 能可靠地導向確認項目管理頁。
            st.session_state.pop("dev_active_tab", None)

        (
            dev_form_tab,
            dev_options_tab,
            dev_sales_tab,
            dev_development_tab,
            dev_handoff_tab,
            dev_installation_tab,
            dev_results_tab,
            dev_excel_tab,
        ) = st.tabs([
            "📝 新版新增裝機",
            "⚙️ 下拉選項管理",
            "💼 業務專區",
            "🛠️ 開發專區",
            "🧰 背鍋俠專區",
            "✅ 裝機確認區",
            "📋 新版搜尋與修改",
            "📥 Excel 匯出",
        ],
            default="⚙️ 下拉選項管理" if requested_checklist_case else None,
            key="dev_active_tab",
            on_change="rerun",
        )

        with dev_handoff_tab:
            render_handoff_area(can_download_attachment)

        with dev_installation_tab:
            render_installation_confirmation_area(can_download_attachment)

        with dev_form_tab:
            st.markdown("### 新版新增裝機紀錄（原型）")
            st.caption("請先完成廠別、案件與機台名稱識別，系統檢查舊紀錄後才會開放其餘欄位。")

            dev_key = st.session_state.dev_add_form_key

            if not st.session_state.dev_identity_draft:
                st.markdown("#### 1. 識別裝機資料")
                identity_col1, identity_col2, identity_col3 = st.columns(3)
                with identity_col1:
                    plant_choices = sorted(
                        st.session_state.dev_plant_options,
                        key=natural_plant_sort_key,
                    ) or ["（請先至選項管理新增廠別）"]
                    identity_plant = st.selectbox(
                        "廠別 *",
                        plant_choices,
                        disabled=not st.session_state.dev_plant_options,
                        key=f"dev_identity_plant_{dev_key}",
                    )
                with identity_col2:
                    case_choices = st.session_state.dev_case_options or ["（請先至選項管理新增案件）"]
                    identity_case = st.selectbox(
                        "案件 *",
                        case_choices,
                        disabled=not st.session_state.dev_case_options,
                        key=f"dev_identity_case_{dev_key}",
                    )
                with identity_col3:
                    identity_machine = st.text_input(
                        "機台名稱 *",
                        placeholder="輸入機台名稱",
                        key=f"dev_identity_machine_{dev_key}",
                    )

                if st.button(
                    "檢查未完成紀錄並繼續",
                    type="primary",
                    use_container_width=True,
                    key=f"dev_check_identity_{dev_key}",
                ):
                    identity_missing = []
                    if not st.session_state.dev_plant_options:
                        identity_missing.append("廠別")
                    if not st.session_state.dev_case_options:
                        identity_missing.append("案件")
                    if not identity_machine.strip():
                        identity_missing.append("機台名稱")

                    if identity_missing:
                        st.error(f"請填寫必填欄位：{'、'.join(identity_missing)}")
                    else:
                        identity_data = {
                            "廠別": identity_plant,
                            "案件": identity_case,
                            "機台名稱": identity_machine.strip(),
                        }
                        target_identity = (
                            identity_plant.strip().casefold(),
                            identity_case.strip().casefold(),
                            identity_machine.strip().casefold(),
                        )
                        previous_unfinished = find_latest_unfinished_new_record(
                            load_new_installation_records(),
                            target_identity,
                        )

                        if previous_unfinished:
                            previous_unfinished = dict(previous_unfinished)
                            previous_unfinished.setdefault("資料來源", "新版裝機紀錄")
                        else:
                            previous_unfinished = find_latest_unfinished_production_record(
                                load_production_installation_records(),
                                target_identity,
                            )

                        if previous_unfinished:
                            st.session_state.dev_pending_previous_record = {
                                "基本資料": identity_data,
                                "上次資料": dict(previous_unfinished),
                            }
                            show_previous_record_dialog()
                        else:
                            st.session_state.dev_identity_draft = identity_data
                            st.session_state.dev_loaded_case = identity_case
                            st.session_state.dev_previous_prefill = None
                            st.session_state.dev_pending_previous_record = None
                            st.session_state.dev_add_form_key += 1
                            st.session_state.dev_checklist_key += 1
                            st.rerun()
            else:
                identity_data = st.session_state.dev_identity_draft
                loaded_case = identity_data["案件"]
                previous_prefill = st.session_state.dev_previous_prefill or {}

                identity_info_col, identity_action_col = st.columns([4, 1])
                with identity_info_col:
                    st.success(
                        f"已確認：{identity_data['廠別']}／{loaded_case}／{identity_data['機台名稱']}"
                    )
                    if previous_prefill.get("資料來源"):
                        st.caption(f"已帶入來源：{previous_prefill['資料來源']}（唯讀參考）")
                with identity_action_col:
                    if st.button(
                        "重新選擇",
                        use_container_width=True,
                        key=f"dev_reset_identity_{dev_key}",
                    ):
                        st.session_state.dev_identity_draft = None
                        st.session_state.dev_previous_prefill = None
                        st.session_state.dev_loaded_case = None
                        st.session_state.dev_add_preview = None
                        st.session_state.dev_pending_preview = None
                        st.session_state.dev_pending_previous_record = None
                        st.session_state.dev_add_form_key += 1
                        st.session_state.dev_checklist_key += 1
                        st.rerun()

                checklist_items = st.session_state.dev_case_checklists.get(loaded_case, [])
                checklist_key = st.session_state.dev_checklist_key
                previous_checklist = str(previous_prefill.get("項目確認", ""))
                status_options = ["未完成", "已完成"]
                previous_status = str(previous_prefill.get("狀態", ""))
                status_index = status_options.index(previous_status) if previous_status in status_options else 0

                with st.form(f"dev_add_installation_form_{dev_key}"):
                    st.markdown("#### 2. 裝機日期")
                    dev_date = st.date_input(
                        "裝機日期 *",
                        datetime.now(),
                        key=f"dev_date_{dev_key}",
                    )

                    st.markdown("#### 3. 項目確認")
                    checklist_results = {}
                    if not checklist_items:
                        st.info("此案件尚未設定確認項目，請至「下拉選項管理」新增。")
                    else:
                        checklist_results = render_checklist_editor(
                            checklist_items,
                            previous_checklist,
                            f"dev_check_{dev_key}_{checklist_key}",
                        )

                    st.markdown("#### 4. 執行資訊")
                    work_col1, work_col2 = st.columns(2)
                    with work_col1:
                        dev_status = st.selectbox(
                            "目前狀態",
                            status_options,
                            index=status_index,
                            key=f"dev_status_{dev_key}",
                        )
                    with work_col2:
                        dev_installers = st.multiselect(
                            "安裝人員",
                            installers_list,
                            key=f"dev_installers_{dev_key}",
                        )

                    st.markdown("#### 5. 備註")
                    dev_remark = st.text_area(
                        "Remark",
                        value=str(previous_prefill.get("Remark", "")),
                        placeholder="輸入進度、缺料項目或其他注意事項",
                        height=120,
                        key=f"dev_remark_{dev_key}",
                    )

                    st.markdown("#### 6. 裝機照片")
                    dev_photos = st.file_uploader(
                        "上傳照片（可多選）",
                        type=["jpg", "jpeg", "png", "webp", "heic"],
                        accept_multiple_files=True,
                        disabled=not can_upload_attachment,
                        help="最多 10 張，每張不可超過 10 MB；按下確認加入時才會上傳至私人 Google Drive。",
                        key=f"dev_photos_{dev_key}",
                    )
                    if not can_upload_attachment:
                        st.caption("目前帳號沒有上傳照片的權限。")

                    preview_submitted = st.form_submit_button(
                        "產生送出預覽",
                        type="primary",
                        use_container_width=True,
                    )

                if preview_submitted:
                    photo_error = ""
                    if len(dev_photos) > 10:
                        photo_error = "一次最多只能上傳 10 張照片。"
                    oversized_photos = [
                        photo.name for photo in dev_photos
                        if len(photo.getvalue()) > 10 * 1024 * 1024
                    ]
                    if oversized_photos:
                        photo_error = (
                            "下列照片超過 10 MB：" + "、".join(oversized_photos)
                        )
                    if photo_error:
                        st.error(photo_error)
                        st.session_state.dev_add_preview = None
                        st.session_state.dev_pending_preview = None

                    checklist_summary = "、".join(
                        f"{'✅' if checked else '❌'} {item_name}"
                        for item_name, checked in checklist_results.items()
                    ) or "未設定確認項目"
                    current_preview = {
                        "日期": dev_date.strftime("%Y-%m-%d"),
                        "廠別": identity_data["廠別"],
                        "案件": loaded_case,
                        "機台名稱": identity_data["機台名稱"],
                        "項目確認": checklist_summary,
                        "安裝人員": "、".join(dev_installers) if dev_installers else "未指定",
                        "狀態": dev_status,
                        "Remark": dev_remark.strip(),
                        "_待上傳照片": [
                            {
                                "name": photo.name,
                                "type": photo.type or "image/jpeg",
                                "data": photo.getvalue(),
                            }
                            for photo in dev_photos
                        ],
                    }
                    if not photo_error:
                        prepare_dev_preview(current_preview)
                        if st.session_state.dev_pending_preview:
                            show_dev_reason_dialog()

                if st.session_state.dev_pending_preview and not preview_submitted:
                    show_dev_reason_dialog()

                if st.session_state.dev_add_preview:
                    st.divider()
                    st.markdown("### 送出前預覽")
                    preview_record = {
                        key: value
                        for key, value in st.session_state.dev_add_preview.items()
                        if not str(key).startswith("_")
                    }
                    pending_photo_names = [
                        photo.get("name", "")
                        for photo in st.session_state.dev_add_preview.get("_待上傳照片", [])
                    ]
                    preview_record["照片"] = "、".join(pending_photo_names) or "（未上傳照片）"
                    if preview_record.get("未完成或缺貨原因"):
                        preview_record["未完成或缺貨原因"] = format_incomplete_reason(
                            preview_record["未完成或缺貨原因"]
                        )
                    preview_df = pd.DataFrame([preview_record]).rename(
                        columns={"未完成或缺貨原因": "未完成原因"}
                    )
                    st.dataframe(preview_df, hide_index=True, use_container_width=True)
                    st.success("請確認內容；確認後可加入測試結果清單。")

                    if st.button(
                        "確認加入測試結果",
                        type="primary",
                        use_container_width=True,
                        key="dev_confirm_test_record",
                    ):
                        test_record = dict(st.session_state.dev_add_preview)
                        pending_photos = test_record.pop("_待上傳照片", [])
                        test_record["建立時間"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        uploaded_photos = []
                        record_saved = False
                        try:
                            for pending_photo in pending_photos:
                                uploaded_photos.append(
                                    upload_installation_photo(
                                        pending_photo,
                                        test_record.get("廠別", ""),
                                        test_record.get("案件", ""),
                                        test_record.get("機台名稱", ""),
                                    )
                                )
                            test_record["照片檔名"] = json.dumps(
                                [photo["name"] for photo in uploaded_photos],
                                ensure_ascii=False,
                            ) if uploaded_photos else ""
                            test_record["照片連結"] = json.dumps(
                                [photo["url"] for photo in uploaded_photos],
                                ensure_ascii=False,
                            ) if uploaded_photos else ""
                            test_record["照片ID"] = json.dumps(
                                [photo["id"] for photo in uploaded_photos],
                                ensure_ascii=False,
                            ) if uploaded_photos else ""
                            test_record["來源版本"] = "新版輸入"
                            new_record_id = append_new_installation_record(test_record)
                            record_saved = True
                            if test_record.get("狀態") == "已完成":
                                try:
                                    complete_matching_new_installation_records(
                                        test_record,
                                        exclude_record_id=new_record_id,
                                    )
                                except Exception:
                                    pass
                            st.session_state.dev_add_preview = None
                            st.session_state.dev_previous_prefill = None
                            st.session_state.dev_identity_draft = None
                            st.session_state.dev_loaded_case = None
                            st.session_state.dev_add_form_key += 1
                            st.session_state.dev_checklist_key += 1
                            st.session_state.dev_flash_level = "success"
                            st.session_state.dev_flash_message = (
                                f"已新增至「{NEW_INSTALLATION_WORKSHEET_NAME}」"
                                f"，並上傳 {len(uploaded_photos)} 張照片。"
                            )
                            st.rerun()
                        except Exception as e:
                            if not record_saved:
                                for uploaded_photo in uploaded_photos:
                                    try:
                                        delete_dev_attachment(uploaded_photo.get("id", ""))
                                    except Exception:
                                        pass
                            st.error(f"新版裝機紀錄儲存失敗：{e}")

                if st.button("清空原型表單", key="dev_clear_add_form"):
                    st.session_state.dev_add_form_key += 1
                    st.session_state.dev_checklist_key += 1
                    st.session_state.dev_loaded_case = None
                    st.session_state.dev_identity_draft = None
                    st.session_state.dev_previous_prefill = None
                    st.session_state.dev_add_preview = None
                    st.session_state.dev_pending_preview = None
                    st.session_state.dev_pending_previous_record = None
                    st.rerun()

        with dev_options_tab:
            st.markdown("### 下拉選項管理")
            st.caption("管理新版表單使用的廠別與案件選項；新增、修改與刪除後會自動保存至 Google Sheets。")

            def render_option_manager(title, state_key, widget_prefix):
                options = st.session_state[state_key]
                if state_key == "dev_plant_options":
                    options.sort(key=natural_plant_sort_key)
                manager_key = st.session_state.dev_option_manager_key

                with st.container(border=True):
                    st.markdown(f"#### {title}")

                    add_col1, add_col2 = st.columns([3, 1])
                    with add_col1:
                        new_option = st.text_input(
                            f"新增{title}",
                            placeholder=f"輸入新的{title}名稱",
                            key=f"{widget_prefix}_new_{manager_key}",
                        )
                    with add_col2:
                        st.write("")
                        st.write("")
                        add_clicked = st.button(
                            "新增選項",
                            use_container_width=True,
                            key=f"{widget_prefix}_add_{manager_key}",
                        )

                    if add_clicked:
                        cleaned_name = new_option.strip()
                        if not cleaned_name:
                            st.error(f"請輸入{title}名稱。")
                        elif cleaned_name in options:
                            st.warning(f"此{title}已存在。")
                        else:
                            options.append(cleaned_name)
                            options.sort(
                                key=natural_plant_sort_key
                                if state_key == "dev_plant_options"
                                else None
                            )
                            if state_key == "dev_case_options":
                                st.session_state.dev_case_checklists.setdefault(cleaned_name, [])
                            queue_dev_auto_sync(f"已新增{title}「{cleaned_name}」")
                            st.session_state.dev_option_manager_key += 1
                            st.session_state.dev_add_form_key += 1
                            st.rerun()

                    if options:
                        edit_col1, edit_col2 = st.columns(2)
                        with edit_col1:
                            selected_option = st.selectbox(
                                f"選擇要修改的{title}",
                                options,
                                key=f"{widget_prefix}_selected_{manager_key}",
                            )
                        with edit_col2:
                            renamed_option = st.text_input(
                                f"新的{title}名稱",
                                placeholder="輸入新名稱",
                                key=f"{widget_prefix}_rename_value_{manager_key}",
                            )

                        action_col1, action_col2 = st.columns(2)
                        with action_col1:
                            rename_clicked = st.button(
                                "儲存名稱修改",
                                use_container_width=True,
                                key=f"{widget_prefix}_rename_{manager_key}",
                            )
                        with action_col2:
                            delete_clicked = st.button(
                                "刪除選項",
                                use_container_width=True,
                                key=f"{widget_prefix}_delete_{manager_key}",
                            )

                        if rename_clicked:
                            cleaned_rename = renamed_option.strip()
                            if not cleaned_rename:
                                st.error(f"請輸入新的{title}名稱。")
                            elif cleaned_rename != selected_option and cleaned_rename in options:
                                st.warning(f"新的{title}名稱已存在。")
                            else:
                                option_index = options.index(selected_option)
                                options[option_index] = cleaned_rename
                                options.sort(
                                    key=natural_plant_sort_key
                                    if state_key == "dev_plant_options"
                                    else None
                                )
                                if state_key == "dev_case_options" and cleaned_rename != selected_option:
                                    existing_items = st.session_state.dev_case_checklists.pop(selected_option, [])
                                    st.session_state.dev_case_checklists[cleaned_rename] = existing_items
                                    if st.session_state.dev_loaded_case == selected_option:
                                        st.session_state.dev_loaded_case = cleaned_rename
                                queue_dev_auto_sync(
                                    f"已將{title}「{selected_option}」修改為「{cleaned_rename}」"
                                )
                                st.session_state.dev_option_manager_key += 1
                                st.session_state.dev_add_form_key += 1
                                st.session_state.dev_add_preview = None
                                st.rerun()

                        if delete_clicked:
                            options.remove(selected_option)
                            if state_key == "dev_case_options":
                                st.session_state.dev_case_checklists.pop(selected_option, None)
                                if st.session_state.dev_loaded_case == selected_option:
                                    st.session_state.dev_loaded_case = None
                            queue_dev_auto_sync(f"已刪除{title}「{selected_option}」")
                            st.session_state.dev_option_manager_key += 1
                            st.session_state.dev_add_form_key += 1
                            st.session_state.dev_add_preview = None
                            st.rerun()
                    else:
                        st.info(f"目前沒有{title}選項，請先新增。")

            with st.container(border=True):
                st.markdown("#### 案件確認項目")
                st.caption("使用 [主分類] 與 [[子分類]] 建立階層；分類只供整理，不需勾選。")

                if requested_checklist_case:
                    st.success(f"目前正在設定案件「{requested_checklist_case}」的確認項目。")

                if st.session_state.dev_case_options:
                    checklist_manager_key = st.session_state.dev_option_manager_key
                    checklist_default_index = 0
                    if requested_checklist_case in st.session_state.dev_case_options:
                        checklist_default_index = st.session_state.dev_case_options.index(
                            requested_checklist_case
                        )
                    checklist_case = st.selectbox(
                        "選擇案件",
                        st.session_state.dev_case_options,
                        index=checklist_default_index,
                        key=f"dev_checklist_case_{checklist_manager_key}",
                    )
                    if requested_checklist_case:
                        st.session_state.dev_checklist_navigation_case = ""
                    checklist_case_index = st.session_state.dev_case_options.index(checklist_case)
                    existing_checklist = st.session_state.dev_case_checklists.get(checklist_case, [])
                    checklist_text = st.text_area(
                        "分類與確認項目",
                        value="\n".join(existing_checklist),
                        placeholder="例如：\n[電力系統]\n電源確認\n[[接地系統]]\n接地線確認\n接地阻抗確認",
                        height=180,
                        key=f"dev_checklist_text_{checklist_manager_key}_{checklist_case_index}",
                    )

                    if st.button(
                        "儲存案件確認項目",
                        type="primary",
                        use_container_width=True,
                        key=f"dev_save_checklist_{checklist_manager_key}",
                    ):
                        cleaned_items = normalize_checklist_definition_lines(checklist_text)

                        st.session_state.dev_case_checklists[checklist_case] = cleaned_items
                        queue_dev_auto_sync(f"已更新案件「{checklist_case}」的確認項目")
                        st.session_state.dev_option_manager_key += 1
                        st.session_state.dev_checklist_key += 1
                        st.session_state.dev_add_preview = None
                        st.rerun()
                else:
                    st.info("請先新增案件，才能設定確認項目。")

            render_option_manager("廠別", "dev_plant_options", "dev_plant")
            render_option_manager("案件", "dev_case_options", "dev_case")

        with dev_sales_tab:
            st.markdown("### 業務專區")
            st.caption("輸入訂單與品號後會立即記錄，並自動保存至 Google Sheets。")

            sales_form_key = st.session_state.dev_sales_form_key
            with st.form(f"dev_sales_form_{sales_form_key}"):
                sales_col1, sales_col2 = st.columns(2)
                with sales_col1:
                    sales_order = st.text_input(
                        "訂單 *",
                        placeholder="輸入訂單編號",
                        key=f"dev_sales_order_{sales_form_key}",
                    )
                with sales_col2:
                    sales_part_number = st.text_input(
                        "品號 *",
                        placeholder="輸入品號",
                        key=f"dev_sales_part_{sales_form_key}",
                    )

                sales_submitted = st.form_submit_button(
                    "新增業務紀錄",
                    type="primary",
                    use_container_width=True,
                )

            if sales_submitted:
                cleaned_order = sales_order.strip()
                cleaned_part_number = sales_part_number.strip()
                missing_sales_fields = []
                if not cleaned_order:
                    missing_sales_fields.append("訂單")
                if not cleaned_part_number:
                    missing_sales_fields.append("品號")

                if missing_sales_fields:
                    st.error(f"請填寫必填欄位：{'、'.join(missing_sales_fields)}")
                else:
                    st.session_state.dev_sales_records.append({
                        "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "訂單": cleaned_order,
                        "品號": cleaned_part_number,
                        "建立者": f"{st.session_state.user_name} ({st.session_state.user_role})",
                    })
                    queue_dev_auto_sync(
                        f"已新增業務紀錄：訂單「{cleaned_order}」、品號「{cleaned_part_number}」"
                    )
                    st.session_state.dev_sales_form_key += 1
                    st.rerun()

            st.divider()
            st.markdown("#### 已記錄資料")
            if st.session_state.dev_sales_records:
                sales_df = pd.DataFrame(st.session_state.dev_sales_records)
                sales_columns = [
                    column for column in ["建立時間", "訂單", "品號", "建立者"]
                    if column in sales_df.columns
                ]
                st.dataframe(
                    sales_df[sales_columns],
                    hide_index=True,
                    use_container_width=True,
                )
                st.info(f"目前共有 {len(sales_df)} 筆業務資料。")
                render_sales_delete_controls("admin_sales")
            else:
                st.info("目前沒有業務資料。")

        with dev_development_tab:
            st.markdown("### 開發專區")
            st.caption("從業務專區選擇訂單與品號，上傳附件後建立開發紀錄。")

            sales_pairs = []
            for sales_record in st.session_state.dev_sales_records:
                pair = (
                    str(sales_record.get("訂單", "")).strip(),
                    str(sales_record.get("品號", "")).strip(),
                )
                if all(pair) and pair not in sales_pairs:
                    sales_pairs.append(pair)

            if not sales_pairs:
                st.info("業務專區目前沒有可選擇的訂單與品號，請先新增業務資料。")
            else:
                development_orders = list(dict.fromkeys(order for order, _ in sales_pairs))
                selected_development_order = st.selectbox(
                    "選擇訂單 *",
                    development_orders,
                    key="dev_development_order",
                )
                development_parts = [
                    part_number for order, part_number in sales_pairs
                    if order == selected_development_order
                ]
                selected_development_part = st.selectbox(
                    "選擇品號 *",
                    development_parts,
                    key=(
                        f"dev_development_part_{st.session_state.dev_development_form_key}_"
                        f"{selected_development_order}"
                    ),
                )

                development_form_key = st.session_state.dev_development_form_key
                with st.form(f"dev_development_form_{development_form_key}"):
                    development_file = st.file_uploader(
                        "附加檔案 *",
                        accept_multiple_files=False,
                        key=f"dev_development_files_{development_form_key}",
                    )
                    st.caption("每筆紀錄附加一個檔案，單檔上限 10 MB；附件會保存到指定的 Google Drive 資料夾。")
                    development_submitted = st.form_submit_button(
                        "上傳附件並建立紀錄",
                        type="primary",
                        use_container_width=True,
                    )

                if development_submitted:
                    if development_file is None:
                        st.error("請附加一個檔案。")
                    elif development_file.size > 10 * 1024 * 1024:
                        st.error("附件不可超過 10 MB。")
                    else:
                        try:
                            with st.spinner("正在上傳附件並建立開發紀錄..."):
                                uploaded_drive_file = upload_dev_attachment(
                                    development_file,
                                    selected_development_order,
                                    selected_development_part,
                                )

                                st.session_state.dev_development_records.append({
                                    "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "訂單": selected_development_order,
                                    "品號": selected_development_part,
                                    "案件": "",
                                    "附件檔名": uploaded_drive_file["原始檔名"],
                                    "附件連結": uploaded_drive_file["連結"],
                                    "附件ID": uploaded_drive_file["id"],
                                    "背鍋俠確認時間": "",
                                    "背鍋俠確認人": "",
                                    "建立者": (
                                        f"{st.session_state.user_name} "
                                        f"({st.session_state.user_role})"
                                    ),
                                })
                                queue_dev_auto_sync(
                                    f"已建立訂單「{selected_development_order}」、"
                                    f"品號「{selected_development_part}」的開發紀錄"
                                )
                                st.session_state.dev_development_form_key += 1
                                st.rerun()
                        except Exception as e:
                            st.error(f"附件上傳或紀錄建立失敗，本次資料未保存。詳細錯誤：{e}")

            st.divider()
            st.markdown("#### 已建立的開發紀錄")
            if st.session_state.dev_development_records:
                development_df = pd.DataFrame(st.session_state.dev_development_records)
                development_display = development_df.copy()
                development_display["附件數量"] = development_display["附件檔名"].apply(
                    lambda value: len([name for name in str(value).splitlines() if name.strip()])
                )
                development_columns = [
                    column for column in ["建立時間", "訂單", "品號", "附件數量", "建立者"]
                    if column in development_display.columns
                ]
                st.dataframe(
                    development_display[development_columns],
                    hide_index=True,
                    use_container_width=True,
                )

                for record_index, development_record in reversed(
                    list(enumerate(st.session_state.dev_development_records, start=1))
                ):
                    with st.expander(
                        f"{development_record.get('訂單', '')}｜"
                        f"{development_record.get('品號', '')}｜"
                        f"{development_record.get('建立時間', '')}",
                        expanded=False,
                    ):
                        attachment_names = [
                            name.strip()
                            for name in str(development_record.get("附件檔名", "")).splitlines()
                            if name.strip()
                        ]
                        attachment_links = [
                            link.strip()
                            for link in str(development_record.get("附件連結", "")).splitlines()
                            if link.strip()
                        ]
                        for attachment_index, attachment_name in enumerate(attachment_names):
                            if attachment_index < len(attachment_links):
                                st.link_button(
                                    f"📎 {attachment_name}",
                                    attachment_links[attachment_index],
                                    use_container_width=True,
                                    key=(
                                        f"dev_attachment_{record_index}_{attachment_index}_"
                                        f"{st.session_state.dev_development_form_key}"
                                    ),
                                )
                            else:
                                st.write(f"📎 {attachment_name}")
                        render_development_delete_button(record_index - 1, "admin_development")
                st.info(f"目前共有 {len(development_df)} 筆開發資料。")
            else:
                st.info("目前沒有開發資料。")

        with dev_results_tab:
            st.markdown("### 新版裝機資料搜尋與修改")
            st.caption("此處合併顯示新版與舊版裝機資料；舊版修改後會轉存為新版，原始資料保持不變。")

            new_results_records = load_new_installation_records()
            converted_legacy_keys = {
                str(record.get("來源鍵", "")).strip()
                for record in new_results_records
                if str(record.get("來源鍵", "")).startswith("OLD-")
            }
            legacy_results_records = [
                record for record in legacy_records_for_new_interface()
                if str(record.get("來源鍵", "")).strip() not in converted_legacy_keys
            ]
            if new_results_records or legacy_results_records:
                combined_results_records = [
                    {
                        **dict(record),
                        "資料來源": "新版",
                        "_record_version": "new",
                    }
                    for record in new_results_records
                ]
                combined_results_records.extend([
                    {
                        **dict(record),
                        "資料來源": "舊版",
                        "_record_version": "legacy",
                    }
                    for record in legacy_results_records
                ])
                results_df = pd.DataFrame(combined_results_records)
                st.markdown("#### 搜尋裝機資料")
                search_row1_col1, search_row1_col2, search_row1_col3 = st.columns(3)
                with search_row1_col1:
                    dev_search_dates = st.date_input(
                        "日期區間",
                        [],
                        key="dev_results_search_dates",
                    )
                with search_row1_col2:
                    dev_search_plants = ["（全部）"] + sorted(
                        {
                            str(record.get("廠別", "")).strip()
                            for record in combined_results_records
                            if str(record.get("廠別", "")).strip()
                        },
                        key=natural_plant_sort_key,
                    )
                    dev_search_plant = st.selectbox(
                        "廠別",
                        dev_search_plants,
                        key="dev_results_search_plant",
                    )
                plant_filtered_records = [
                    record for record in combined_results_records
                    if dev_search_plant == "（全部）"
                    or str(record.get("廠別", "")).strip() == dev_search_plant
                ]
                with search_row1_col3:
                    dev_search_cases = sorted({
                        str(record.get("案件", "")).strip()
                        for record in plant_filtered_records
                        if str(record.get("案件", "")).strip()
                    })
                    dev_search_case = st.multiselect(
                        "案件（可多選）",
                        dev_search_cases,
                        placeholder="未選擇代表全部案件",
                        key="dev_results_search_cases_multi",
                    )

                case_filtered_records = [
                    record for record in plant_filtered_records
                    if not dev_search_case
                    or str(record.get("案件", "")).strip() in dev_search_case
                ]
                search_row2_col1, search_row2_col2, search_row2_col3 = st.columns(3)
                with search_row2_col1:
                    dev_search_machines = ["（全部）"] + sorted({
                        str(record.get("機台名稱", "")).strip()
                        for record in case_filtered_records
                        if str(record.get("機台名稱", "")).strip()
                    })
                    dev_search_machine = st.selectbox(
                        "機台名稱",
                        dev_search_machines,
                        key="dev_results_search_machine",
                    )
                with search_row2_col2:
                    dev_search_installers = ["（全部）"] + sorted({
                        installer.strip()
                        for record in combined_results_records
                        for installer in re.split(",|、", str(record.get("安裝人員", "")))
                        if installer.strip() and installer.strip() != "未指定"
                    })
                    dev_search_installer = st.selectbox(
                        "安裝人員",
                        dev_search_installers,
                        key="dev_results_search_installer",
                    )
                with search_row2_col3:
                    available_result_statuses = sorted({
                        str(record.get("狀態", "")).strip()
                        for record in combined_results_records
                        if str(record.get("狀態", "")).strip()
                    })
                    dev_search_status = st.selectbox(
                        "狀態",
                        ["（全部）", *available_result_statuses],
                        key="dev_results_search_status",
                    )

                filtered_record_indices = []
                for record_index, record in enumerate(combined_results_records):
                    if (
                        dev_search_plant != "（全部）"
                        and str(record.get("廠別", "")).strip() != dev_search_plant
                    ):
                        continue
                    if (
                        dev_search_case
                        and str(record.get("案件", "")).strip() not in dev_search_case
                    ):
                        continue
                    if (
                        dev_search_machine != "（全部）"
                        and str(record.get("機台名稱", "")).strip() != dev_search_machine
                    ):
                        continue
                    if (
                        dev_search_installer != "（全部）"
                        and dev_search_installer not in {
                            installer.strip()
                            for installer in re.split(",|、", str(record.get("安裝人員", "")))
                            if installer.strip()
                        }
                    ):
                        continue
                    if (
                        dev_search_status != "（全部）"
                        and str(record.get("狀態", "")).strip() != dev_search_status
                    ):
                        continue
                    record_date = pd.to_datetime(record.get("日期", ""), errors="coerce")
                    if len(dev_search_dates) == 2 and (
                        pd.isna(record_date)
                        or not (dev_search_dates[0] <= record_date.date() <= dev_search_dates[1])
                    ):
                        continue
                    if len(dev_search_dates) == 1 and (
                        pd.isna(record_date) or record_date.date() != dev_search_dates[0]
                    ):
                        continue
                    filtered_record_indices.append(record_index)

                filtered_results_df = results_df.iloc[filtered_record_indices].copy()
                preferred_columns = [
                    "資料來源",
                    "建立時間",
                    "日期",
                    "廠別",
                    "案件",
                    "機台名稱",
                    "項目確認",
                    "安裝人員",
                    "狀態",
                    "未完成或缺貨原因",
                    "Remark",
                ]
                result_columns = [column for column in preferred_columns if column in filtered_results_df.columns]
                display_results_df = filtered_results_df[result_columns].copy()
                display_results_df = display_results_df.rename(
                    columns={"未完成或缺貨原因": "未完成原因"}
                )
                if "未完成原因" in display_results_df.columns:
                    display_results_df["未完成原因"] = display_results_df["未完成原因"].apply(
                        format_incomplete_reason
                    )
                if "項目確認" in display_results_df.columns:
                    display_results_df["項目確認"] = display_results_df["項目確認"].apply(
                        format_checklist_progress
                    )
                st.markdown(f"#### 搜尋結果（{len(display_results_df)} 筆）")
                if display_results_df.empty:
                    st.info("沒有符合目前搜尋條件的裝機資料。")
                else:
                    results_event = st.dataframe(
                        display_results_df,
                        hide_index=True,
                        use_container_width=True,
                        on_select="rerun",
                        selection_mode="single-row",
                        key=f"dev_results_grid_{st.session_state.dev_results_grid_key}",
                    )
                    if (
                        results_event.selection.rows
                        and not st.session_state.dev_pending_preview
                        and not st.session_state.dev_pending_previous_record
                    ):
                        selected_display_index = results_event.selection.rows[0]
                        selected_combined_index = filtered_record_indices[selected_display_index]
                        show_details_dialog(
                            pd.Series(combined_results_records[selected_combined_index]),
                            "dev_results_grid_key",
                        )

                    st.markdown("#### 修改裝機資料")
                    edit_record_options = filtered_record_indices or [None]
                    edit_record_index = st.selectbox(
                        "選擇要修改的紀錄",
                        edit_record_options,
                        format_func=lambda index: (
                            "（目前沒有可修改的紀錄）"
                            if index is None else
                            f"[{combined_results_records[index].get('資料來源', '')}]｜"
                            f"{combined_results_records[index].get('日期', '')}｜"
                            f"{combined_results_records[index].get('廠別', '')}｜"
                            f"{combined_results_records[index].get('案件', '')}｜"
                            f"{combined_results_records[index].get('機台名稱', '')}"
                        ),
                        disabled=not filtered_record_indices,
                        key="dev_results_edit_record_v2",
                    )
                    edit_record = (
                        combined_results_records[edit_record_index]
                        if edit_record_index is not None else {}
                    )
                    if edit_record_index is not None:
                        if edit_record.get("_record_version") == "legacy":
                            st.info("這是舊版紀錄；儲存後會轉成新版並寫入「新版裝機紀錄」，舊版原始資料不會改動。")
                        else:
                            st.caption("這是新版紀錄；儲存後會直接更新新版資料庫。")
                    parsed_edit_date = pd.to_datetime(edit_record.get("日期", ""), errors="coerce")
                    edit_date_default = (
                        parsed_edit_date.date()
                        if pd.notna(parsed_edit_date)
                        else datetime.now().date()
                    )
                    edit_form_key = st.session_state.dev_results_edit_form_key
                    edit_case_options = sorted(
                        set(st.session_state.dev_case_options)
                        | {str(edit_record.get("案件", "")).strip()}
                    )
                    edited_case = st.selectbox(
                        "案件 *",
                        edit_case_options,
                        index=edit_case_options.index(str(edit_record.get("案件", "")).strip()),
                        key=f"dev_edit_case_{edit_form_key}_{edit_record_index}",
                        help="切換案件後，下方會立即帶入該案件已建立的確認項目。",
                    )
                    edit_checklist_definition = st.session_state.dev_case_checklists.get(
                        edited_case,
                        [],
                    )
                    with st.form(f"dev_results_edit_form_{edit_form_key}_{edit_record_index}"):
                        edit_col1, edit_col2, edit_col3 = st.columns(3)
                        with edit_col1:
                            edited_date = st.date_input("裝機日期 *", edit_date_default)
                            edit_plant_options = sorted(
                                set(st.session_state.dev_plant_options)
                                | {str(edit_record.get("廠別", "")).strip()},
                                key=natural_plant_sort_key,
                            )
                            edited_plant = st.selectbox(
                                "廠別 *",
                                edit_plant_options,
                                index=edit_plant_options.index(str(edit_record.get("廠別", "")).strip()),
                            )
                        with edit_col2:
                            edited_machine = st.text_input(
                                "機台名稱 *",
                                value=str(edit_record.get("機台名稱", "")),
                            )
                        with edit_col3:
                            edited_installers = st.text_input(
                                "安裝人員",
                                value=str(edit_record.get("安裝人員", "")),
                                help="多人請使用頓號「、」分隔。",
                            )
                            edit_status_options = ["未完成", "已完成"]
                            current_edit_status = str(edit_record.get("狀態", "")).strip()
                            normalized_edit_status = (
                                "已完成" if current_edit_status in {"完成", "已完成"} else "未完成"
                            )
                            edited_status = st.selectbox(
                                "狀態 *",
                                edit_status_options,
                                index=edit_status_options.index(normalized_edit_status),
                            )

                        st.markdown("#### 項目確認")
                        if edit_checklist_definition:
                            st.caption(
                                f"已帶入案件「{edited_case}」的確認項目；勾選代表已施工，未勾選代表待施工。"
                            )
                            checklist_current_summary = (
                                str(edit_record.get("項目確認", ""))
                                if edited_case == str(edit_record.get("案件", "")).strip()
                                else ""
                            )
                            edited_checklist_results = render_checklist_editor(
                                edit_checklist_definition,
                                checklist_current_summary,
                                f"dev_edit_checklist_{edit_form_key}_{edit_record_index}_{edited_case}",
                            )
                            edited_checklist = ""
                        else:
                            edited_checklist_results = None
                            st.info(f"案件「{edited_case}」尚未建立確認項目，暫時保留文字輸入。")
                            edited_checklist = st.text_area(
                                "項目確認",
                                value=str(edit_record.get("項目確認", "")),
                                height=180,
                                help="保留每個項目前方的 ✅ 或 ❌，以及完整分類路徑。",
                            )
                        edited_reason = st.text_area(
                            "未完成原因",
                            value=str(edit_record.get("未完成或缺貨原因", "")),
                            height=100,
                        )
                        edited_remark = st.text_area(
                            "Remark",
                            value=str(edit_record.get("Remark", "")),
                            height=120,
                        )
                        edit_submitted = st.form_submit_button(
                            "儲存修改",
                            type="primary",
                            use_container_width=True,
                            disabled=edit_record_index is None,
                        )

                    if (
                        edit_record_index is not None
                        and edit_record.get("_record_version") == "new"
                    ):
                        if st.button(
                            "🗑️ 刪除此筆新版裝機資料",
                            use_container_width=True,
                            disabled=not can_delete_dev_data(),
                            help="若紀錄包含照片，確認後會一併移至 Google Drive 垃圾桶。",
                            key=f"delete_new_installation_{edit_record.get('紀錄ID', edit_record_index)}",
                        ):
                            st.session_state.dev_pending_delete = {
                                "type": "installation",
                                "record": dict(edit_record),
                            }
                            st.session_state.dev_delete_dialog_key += 1
                            st.session_state.dev_results_grid_key += 1
                            st.rerun()
                        if not can_delete_dev_data():
                            st.caption("目前帳號沒有刪除裝機資料與照片的權限。")
                    elif edit_record_index is not None:
                        st.caption("舊版資料為唯讀，不能在新版搜尋區刪除。")

                    if edit_submitted and edit_record_index is not None:
                        missing_edit_fields = []
                        if not edited_plant.strip():
                            missing_edit_fields.append("廠別")
                        if not edited_case.strip():
                            missing_edit_fields.append("案件")
                        if not edited_machine.strip():
                            missing_edit_fields.append("機台名稱")
                        if edited_status == "未完成" and not edited_reason.strip():
                            missing_edit_fields.append("未完成原因")

                        if missing_edit_fields:
                            st.error(f"請填寫必填欄位：{'、'.join(missing_edit_fields)}")
                        else:
                            old_record = dict(edit_record)
                            updated_record = dict(edit_record)
                            if edited_checklist_results is not None:
                                edited_checklist = "、".join(
                                    f"{'✅' if checked else '❌'} {item_name}"
                                    for item_name, checked in edited_checklist_results.items()
                                ) or "未設定確認項目"
                            updated_record.update({
                                "日期": edited_date.strftime("%Y-%m-%d"),
                                "廠別": edited_plant.strip(),
                                "案件": edited_case.strip(),
                                "機台名稱": edited_machine.strip(),
                                "項目確認": edited_checklist.strip(),
                                "安裝人員": edited_installers.strip() or "未指定",
                                "狀態": edited_status,
                                "未完成或缺貨原因": (
                                    edited_reason.strip() if edited_status == "未完成" else ""
                                ),
                                "Remark": edited_remark.strip(),
                            })
                            try:
                                if edit_record.get("_record_version") == "legacy":
                                    updated_record["來源版本"] = "舊版轉換"
                                    updated_record["來源鍵"] = edit_record.get("來源鍵", "")
                                    saved_record_id = append_new_installation_record(updated_record)
                                    action_message = (
                                        f"已將舊版裝機紀錄轉為新版：{edited_plant}／"
                                        f"{edited_case}／{edited_machine}"
                                    )
                                else:
                                    update_new_installation_record(edit_record, updated_record)
                                    saved_record_id = edit_record.get("紀錄ID", "")
                                    action_message = (
                                        f"已修改新版裝機紀錄：{edited_plant}／"
                                        f"{edited_case}／{edited_machine}"
                                    )

                                related_completed_count = 0
                                if edited_status == "已完成":
                                    related_completed_count = complete_matching_new_installation_records(
                                        updated_record,
                                        exclude_record_id=saved_record_id,
                                    )
                                if related_completed_count:
                                    action_message += (
                                        f"，並完成 {related_completed_count} 筆相同機台的新版紀錄"
                                    )
                                log_dev_delete_action(
                                    action_message,
                                    str(old_record),
                                    str(updated_record),
                                )
                                st.session_state.dev_flash_level = "success"
                                st.session_state.dev_flash_message = action_message
                                st.session_state.dev_results_edit_form_key += 1
                                st.session_state.dev_results_grid_key += 1
                                st.rerun()
                            except Exception as e:
                                st.error(f"新版裝機紀錄儲存失敗：{e}")

                st.info(
                    f"新版 {len(new_results_records)} 筆｜"
                    f"尚未轉換的舊版 {len(legacy_results_records)} 筆｜"
                    f"目前搜尋符合 {len(filtered_record_indices)} 筆。"
                )
            else:
                st.info("目前沒有新版或舊版裝機資料。")

            st.divider()
            st.success(
                f"☁️ 新版裝機資料會直接保存至「{NEW_INSTALLATION_WORKSHEET_NAME}」；"
                "舊版資料只讀取、不覆寫。"
            )

        with dev_excel_tab:
            st.markdown("### 多機台確認項目 Excel")
            st.caption("依範例格式，每台機台一列；已施工使用綠色儲存格，待施工使用黃色儲存格。")

            export_records = load_new_installation_records()
            if not export_records:
                st.info("目前沒有可匯出的新版裝機資料，請先新增或轉換裝機資料。")
            else:
                export_plants = sorted({
                    str(record.get("廠別", "")).strip()
                    for record in export_records
                    if str(record.get("廠別", "")).strip()
                }, key=natural_plant_sort_key)
                export_plant = st.selectbox(
                    "選擇廠別",
                    export_plants,
                    key="dev_export_plant",
                )

                plant_records = [
                    record for record in export_records
                    if str(record.get("廠別", "")).strip() == export_plant
                ]
                export_cases = sorted({
                    str(record.get("案件", "")).strip()
                    for record in plant_records
                    if str(record.get("案件", "")).strip()
                })
                export_case = st.selectbox(
                    "選擇案件",
                    export_cases,
                    key=f"dev_export_case_{export_plant}",
                )

                case_records = [
                    record for record in plant_records
                    if str(record.get("案件", "")).strip() == export_case
                ]
                latest_machine_records = {}
                for record in case_records:
                    machine_name = str(record.get("機台名稱", "")).strip()
                    if machine_name:
                        latest_machine_records[machine_name] = record

                machine_options = sorted(latest_machine_records)
                selected_machines = st.multiselect(
                    "選擇要輸出的機台（可複選）",
                    machine_options,
                    default=[],
                    key=f"dev_export_machines_{export_plant}_{export_case}",
                )

                checklist_definition = st.session_state.dev_case_checklists.get(export_case, [])
                checklist_options = get_checklist_export_labels(checklist_definition)
                st.markdown("#### 選擇要生成的確認項目")
                st.caption("選項預設全部不勾選，請依主分類與子分類展開挑選。")
                selected_export_items = render_export_checklist_selector(
                    checklist_definition,
                    key_prefix=f"dev_export_checklist_{export_plant}_{export_case}",
                )
                st.caption("🟩 已施工　🟨 待施工")

                if not checklist_options:
                    st.warning("此案件尚未建立確認項目，請先至「下拉選項管理」設定。")
                elif not selected_machines:
                    st.warning("請至少選擇一台機台。")
                elif not selected_export_items:
                    st.warning("請至少選擇一個確認項目。")
                else:
                    selected_machine_records = [
                        latest_machine_records[machine_name]
                        for machine_name in selected_machines
                    ]
                    excel_data = build_dev_excel_export(
                        selected_machine_records,
                        selected_export_items,
                    )
                    unsafe_filename = f"{export_plant}_{export_case}_裝機確認表_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    safe_filename = "".join(
                        "_" if char in '\\/:*?\"<>|' else char
                        for char in unsafe_filename
                    )

                    st.success(
                        f"將輸出 {len(selected_machine_records)} 台機台、"
                        f"{len(selected_export_items)} 個確認項目。"
                    )
                    st.download_button(
                        "下載 Excel",
                        data=excel_data,
                        file_name=safe_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="dev_download_excel",
                    )
